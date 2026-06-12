
import re
import csv
import datetime
import calendar
from datetime import timedelta
from ftplib import FTP
from io import BytesIO
import pandas as pd
import mysql.connector
import numpy as np
import streamlit as st
import logging
import glob
import os
import tempfile
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders
import openpyxl
from playwright.sync_api import sync_playwright, TimeoutError as PlaywrightTimeoutError
import time
import pytz
import math

IST = pytz.timezone('Asia/Kolkata')
DEFAULT_HYBRID_AVC_CAP = 200.0
SCHEDULE_SOLAR_AVC_OVERRIDE_MW = 94.325

def current_time_ist():
    """Return current timezone-aware IST datetime."""
    return datetime.datetime.now(IST)

def today_ist():
    """Return today's date in IST."""
    return current_time_ist().date()


def to_float(value, default=0.0):
    """Safely convert values to float for schedule/AVC math."""
    try:
        if pd.isna(value):
            return default
    except Exception:
        pass

    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def is_missing_scalar(value):
    """Return True only for scalar missing values, not pandas containers."""
    if value is None:
        return True

    if isinstance(value, (pd.Series, pd.DataFrame, np.ndarray, list, tuple, dict, set)):
        return False

    try:
        return bool(pd.isna(value))
    except (TypeError, ValueError):
        return False


def round_down_one_decimal(value):
    """Round down toward zero at one decimal place for non-negative values."""
    numeric_value = max(to_float(value), 0.0)
    return math.floor(numeric_value * 10.0) / 10.0


def normalize_obligation_dataframe(df):
    """
    Keep backward compatibility with legacy obligation rows that only stored
    the total market value. In copy2, Linde GDAM is out of scope, so we
    preserve only a single manual GDAM input path backed by market_fp.
    """
    if df is None:
        return pd.DataFrame(columns=["date", "block", "market_fp", "market"])

    obligation_df = df.copy()

    for column in ["market", "market_fp", "market_linde"]:
        if column not in obligation_df.columns:
            obligation_df[column] = 0.0

    obligation_df["market"] = pd.to_numeric(
        obligation_df["market"].apply(to_float),
        errors="coerce",
    ).fillna(0.0)
    obligation_df["market_fp"] = pd.to_numeric(
        obligation_df["market_fp"].apply(to_float),
        errors="coerce",
    ).fillna(0.0)
    obligation_df["market_linde"] = pd.to_numeric(
        obligation_df["market_linde"].apply(to_float),
        errors="coerce",
    ).fillna(0.0)

    legacy_mask = (
        obligation_df["market"].abs() > 1e-9
    ) & (
        obligation_df["market_fp"].abs() <= 1e-9
    ) & (
        obligation_df["market_linde"].abs() <= 1e-9
    )

    if legacy_mask.any():
        obligation_df.loc[legacy_mask, "market_fp"] = obligation_df.loc[legacy_mask, "market"]

    obligation_df["market_linde"] = 0.0
    obligation_df["market"] = obligation_df["market_fp"].round(1)

    return obligation_df


def compute_hybrid_avc_distribution(wind_schedule, solar_schedule, wind_avc, solar_avc, hybrid_cap):
    """
    Cap total hybrid AVC and redistribute the capped value across wind and solar.

    This mirrors the workbook's AVC redistribution behavior: when the combined
    component AVC exceeds the allowed hybrid AVC cap, the capped total is first
    distributed in proportion to schedule and any overflow beyond a component's
    own AVC is pushed to the other component.
    """
    wind_schedule = max(to_float(wind_schedule), 0.0)
    solar_schedule = max(to_float(solar_schedule), 0.0)
    wind_avc = max(to_float(wind_avc), 0.0)
    solar_avc = max(to_float(solar_avc), 0.0)
    hybrid_cap = max(to_float(hybrid_cap, DEFAULT_HYBRID_AVC_CAP), 0.0)

    total_input_avc = wind_avc + solar_avc
    effective_hybrid_avc = min(total_input_avc, hybrid_cap)

    if effective_hybrid_avc <= 0:
        return {
            "hybrid_avc_input": 0.0,
            "hybrid_avc": 0.0,
            "wind_avc_capped": 0.0,
            "solar_avc_capped": 0.0,
        }

    if total_input_avc <= hybrid_cap + 1e-9:
        return {
            "hybrid_avc_input": round(total_input_avc, 1),
            "hybrid_avc": round(total_input_avc, 1),
            "wind_avc_capped": round(wind_avc, 1),
            "solar_avc_capped": round(solar_avc, 1),
        }

    total_schedule = wind_schedule + solar_schedule
    if total_schedule > 0:
        wind_base = effective_hybrid_avc * (wind_schedule / total_schedule)
        solar_base = effective_hybrid_avc * (solar_schedule / total_schedule)
    else:
        wind_base = effective_hybrid_avc * (wind_avc / total_input_avc) if total_input_avc else 0.0
        solar_base = effective_hybrid_avc * (solar_avc / total_input_avc) if total_input_avc else 0.0

    wind_capped = min(wind_avc, wind_base + max(solar_base - solar_avc, 0.0))
    solar_capped = min(solar_avc, solar_base + max(wind_base - wind_avc, 0.0))

    remaining = effective_hybrid_avc - (wind_capped + solar_capped)
    if remaining > 1e-9:
        wind_room = max(wind_avc - wind_capped, 0.0)
        solar_room = max(solar_avc - solar_capped, 0.0)

        if wind_room >= solar_room and wind_room > 0:
            add = min(remaining, wind_room)
            wind_capped += add
            remaining -= add
            wind_room -= add

        if remaining > 1e-9 and solar_room > 0:
            add = min(remaining, solar_room)
            solar_capped += add
            remaining -= add
            solar_room -= add

        if remaining > 1e-9 and wind_room > 0:
            wind_capped += min(remaining, wind_room)

    wind_capped = round(min(wind_capped, wind_avc), 1)
    solar_capped = round(min(solar_capped, solar_avc), 1)
    hybrid_avc = round(wind_capped + solar_capped, 1)

    rounding_gap = round(round(effective_hybrid_avc, 1) - hybrid_avc, 1)
    if rounding_gap > 0:
        if round(wind_avc - wind_capped, 1) >= rounding_gap:
            wind_capped = round(wind_capped + rounding_gap, 1)
        elif round(solar_avc - solar_capped, 1) >= rounding_gap:
            solar_capped = round(solar_capped + rounding_gap, 1)
        hybrid_avc = round(wind_capped + solar_capped, 1)

    return {
        "hybrid_avc_input": round(total_input_avc, 1),
        "hybrid_avc": hybrid_avc,
        "wind_avc_capped": wind_capped,
        "solar_avc_capped": solar_capped,
    }


def resolve_component_avc(schedule_value, avc_value):
    """
    Use the reported AVC when present; if it is missing/zero, fall back to the
    input schedule so the hybrid builder can still create a schedule.
    """
    schedule_value = round(max(to_float(schedule_value), 0.0), 1)
    avc_value = round(max(to_float(avc_value), 0.0), 1)
    if avc_value <= 0 and schedule_value > 0:
        return schedule_value
    return avc_value


def compute_hybrid_schedule_from_validation_limit(
    solar_input,
    wind_input,
    solar_avc,
    wind_avc,
    hybrid_cap,
    schedule_cap=None,
    wind_priority_pct=100.0,
    solar_priority_pct=0.0,
):
    """
    Build the hybrid schedule by rounding generation down to the effective
    hybrid validation limit.

    This keeps the workbook behavior in the live Streamlit path:
    - start from the raw solar/wind generation inputs,
    - use component AVC where available,
    - cap the combined hybrid schedule by the validation limit,
    - and preserve the component split with priority when a cut is needed.
    """
    solar_input = round(max(to_float(solar_input), 0.0), 1)
    wind_input = round(max(to_float(wind_input), 0.0), 1)
    solar_avc = resolve_component_avc(solar_input, solar_avc)
    wind_avc = resolve_component_avc(wind_input, wind_avc)

    declared_plant_schedule = round(solar_input + wind_input, 1)
    plant_avc_input = round(solar_avc + wind_avc, 1)
    validation_limit = round(max(to_float(hybrid_cap, DEFAULT_HYBRID_AVC_CAP), 0.0), 1)
    effective_hybrid_cap = round(min(plant_avc_input, validation_limit), 1)

    total_cap = effective_hybrid_cap
    if schedule_cap is not None:
        total_cap = round(min(total_cap, max(to_float(schedule_cap), 0.0)), 1)

    final_total = round(min(declared_plant_schedule, total_cap), 1)

    if declared_plant_schedule <= 0 or final_total <= 0:
        return {
            "declared_plant_schedule": declared_plant_schedule,
            "plant_avc_input": plant_avc_input,
            "solar_schedule_final": 0.0,
            "wind_schedule_final": 0.0,
            "hybrid_schedule_final": 0.0,
            "effective_hybrid_cap": effective_hybrid_cap,
            "validation_limit": validation_limit,
        }

    if wind_priority_pct == solar_priority_pct:
        wind_final = round(min(wind_input * final_total / declared_plant_schedule, wind_avc), 1)
        solar_final = round(min(final_total - wind_final, solar_avc), 1)
    elif wind_priority_pct >= solar_priority_pct:
        wind_final = round(min(wind_input, wind_avc, final_total), 1)
        solar_final = round(min(solar_input, solar_avc, final_total - wind_final), 1)
    else:
        solar_final = round(min(solar_input, solar_avc, final_total), 1)
        wind_final = round(min(wind_input, wind_avc, final_total - solar_final), 1)

    allocated_total = round(wind_final + solar_final, 1)
    gap = round(final_total - allocated_total, 1)
    if gap > 0:
        wind_room = round(min(wind_input, wind_avc) - wind_final, 1)
        solar_room = round(min(solar_input, solar_avc) - solar_final, 1)

        if wind_priority_pct >= solar_priority_pct and wind_room > 0:
            add = min(gap, wind_room)
            wind_final = round(wind_final + add, 1)
            gap = round(gap - add, 1)
        if gap > 0 and solar_room > 0:
            add = min(gap, solar_room)
            solar_final = round(solar_final + add, 1)
            gap = round(gap - add, 1)
        if gap > 0 and wind_room > 0:
            add = min(gap, max(round(min(wind_input, wind_avc) - wind_final, 1), 0.0))
            wind_final = round(wind_final + add, 1)

    return {
        "declared_plant_schedule": declared_plant_schedule,
        "plant_avc_input": plant_avc_input,
        "solar_schedule_final": round(max(solar_final, 0.0), 1),
        "wind_schedule_final": round(max(wind_final, 0.0), 1),
        "hybrid_schedule_final": round(max(wind_final + solar_final, 0.0), 1),
        "effective_hybrid_cap": effective_hybrid_cap,
        "validation_limit": validation_limit,
    }


def normalize_template_token(value):
    return re.sub(r"[^a-z0-9]+", "", str(value).strip().lower())


def find_template_row_index(template_df, label):
    target = normalize_template_token(label)
    for row_idx in range(len(template_df)):
        for col_idx in range(len(template_df.columns)):
            if normalize_template_token(template_df.iloc[row_idx, col_idx]) == target:
                return row_idx
    return None


def set_template_value_next_to_label(template_df, label, value, logger=None):
    target = normalize_template_token(label)
    for row_idx in range(len(template_df)):
        for col_idx in range(len(template_df.columns) - 1):
            if normalize_template_token(template_df.iloc[row_idx, col_idx]) == target:
                template_df.iloc[row_idx, col_idx + 1] = format_template_cell_value(value)
                if logger:
                    logger.info(f"Updated template label '{label}' with value '{value}'")
                return True
    return False


def classify_template_avc_columns(template_df, header_row_idx, energy_type_row_idx, capacity_row_idx=None):
    """
    Identify each AVC column in the REMC template and infer whether it belongs
    to the WIND or SOLAR schedule section by looking at the next schedule block
    on the right.
    """
    avc_columns = []
    if header_row_idx is None:
        return avc_columns

    total_columns = len(template_df.columns)
    for col_idx in range(total_columns):
        if normalize_template_token(template_df.iloc[header_row_idx, col_idx]) != "avc":
            continue

        section_energy_type = ""
        if energy_type_row_idx is not None:
            for look_ahead_idx in range(col_idx + 1, total_columns):
                next_header_token = normalize_template_token(
                    template_df.iloc[header_row_idx, look_ahead_idx]
                )
                if next_header_token == "avc":
                    break

                energy_value = str(
                    template_df.iloc[energy_type_row_idx, look_ahead_idx]
                ).strip().upper()
                if energy_value in {"WIND", "SOLAR"}:
                    section_energy_type = energy_value
                    break

        avc_columns.append(
            {
                "col_idx": col_idx,
                "energy_type": section_energy_type,
                "capacity_value": (
                    template_df.iloc[capacity_row_idx, col_idx]
                    if capacity_row_idx is not None and capacity_row_idx < len(template_df)
                    else ""
                ),
            }
        )

    return avc_columns


def is_intraday_template_filename(filename):
    filename_lower = str(filename).strip().lower()
    return filename_lower.endswith(".csv") and (
        "_id_" in filename_lower or "intraday" in filename_lower
    )


def is_dayahead_template_filename(filename):
    filename_lower = str(filename).strip().lower()
    return filename_lower.endswith(".csv") and (
        "_da_" in filename_lower
        or "dayahead" in filename_lower
        or "day_ahead" in filename_lower
        or "day-ahead" in filename_lower
    )


def get_ftp_modified_timestamp(ftp, remote_filename):
    try:
        response = ftp.sendcmd(f"MDTM {remote_filename}")
        parts = str(response).split()
        if len(parts) >= 2:
            return parts[1]
    except Exception:
        return ""
    return ""


def ftp_template_has_hybrid_schema(ftp, remote_filename):
    try:
        buffer = BytesIO()
        ftp.retrbinary(f"RETR {remote_filename}", buffer.write)
        content = buffer.getvalue().decode("utf-8", errors="replace")
        checks = [
            "Hybrid Validation Limit",
            "Declared Hybrid Schedule",
            "Total Hybrid Solar",
            "Total Hybrid Wind",
            "Plant_AvC",
            "Hybrid_AvC",
        ]
        return all(token in content for token in checks)
    except Exception:
        return False


def pick_latest_template_file(ftp, template_files, match_fn, logger=None):
    matches = [file for file in template_files if match_fn(file)]
    if not matches:
        return None

    scored_matches = []
    for remote_filename in matches:
        modified_ts = get_ftp_modified_timestamp(ftp, remote_filename)
        has_hybrid_schema = ftp_template_has_hybrid_schema(ftp, remote_filename)
        score = (
            1 if has_hybrid_schema else 0,
            modified_ts,
            remote_filename,
        )
        scored_matches.append((score, remote_filename, has_hybrid_schema, modified_ts))

    scored_matches.sort(reverse=True)

    if logger:
        for _, remote_filename, has_hybrid_schema, modified_ts in scored_matches:
            logger.info(
                f"Template candidate: {remote_filename} | "
                f"hybrid_schema={has_hybrid_schema} | mdtm={modified_ts or 'N/A'}"
            )

    return scored_matches[0][1]


def format_template_cell_value(value):
    """Convert values to CSV-friendly strings for REMC template writes."""
    if is_missing_scalar(value):
        return ""

    if isinstance(value, (int, float, np.integer, np.floating)):
        numeric_value = float(value)
        rounded = round(numeric_value, 3)
        if abs(rounded - round(rounded)) < 1e-9:
            return str(int(round(rounded)))
        return f"{rounded:.3f}".rstrip("0").rstrip(".")

    return str(value)

# -------------------------------
# Logging Configuration & Functions
# -------------------------------

def setup_logging():
    """Setup logging configuration for bifurcation process"""
    # Create logs directory if it doesn't exist
    if not os.path.exists('logs'):
        os.makedirs('logs')

    # Setup main logger
    logger = logging.getLogger('bifurcation')
    logger.setLevel(logging.INFO)

    # Clear existing handlers
    logger.handlers.clear()

    # Create file handler with timestamp (IST)
    import pytz
    ist = pytz.timezone('Asia/Kolkata')
    timestamp = datetime.datetime.now(ist).strftime("%Y%m%d_%H%M%S")
    log_file = f'logs/bifurcation_{timestamp}.log'

    file_handler = logging.FileHandler(log_file)
    file_handler.setLevel(logging.INFO)

    # Create formatter
    formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
    file_handler.setFormatter(formatter)

    # Add handler to logger
    logger.addHandler(file_handler)

    return logger, log_file

def log_output_results(analysis_type, date, results_summary, detailed_results=None):
    """Log output results for Intra-day and Day Ahead with 7-day retention in tabular format"""
    # Create output_logs directory if it doesn't exist
    if not os.path.exists('output_logs'):
        os.makedirs('output_logs')

    # Create log file name with date
    log_file = f'output_logs/{analysis_type}_{date.strftime("%Y%m%d")}.log'

    # Setup logger for output
    output_logger = logging.getLogger(f'{analysis_type}_output')
    output_logger.setLevel(logging.INFO)

    # Clear existing handlers
    output_logger.handlers.clear()

    # Create file handler
    file_handler = logging.FileHandler(log_file, mode='a')
    file_handler.setLevel(logging.INFO)

    # Create formatter (no timestamp for cleaner tables)
    formatter = logging.Formatter('%(message)s')
    file_handler.setFormatter(formatter)

    # Add handler to logger
    output_logger.addHandler(file_handler)

    # Get current timestamp (IST)
    import pytz
    ist = pytz.timezone('Asia/Kolkata')
    timestamp = datetime.datetime.now(ist).strftime("%Y-%m-%d %H:%M:%S IST")

    # Log the results in tabular format with run timestamp
    output_logger.info("")
    output_logger.info("=" * 150)
    output_logger.info(f"  {analysis_type} ANALYSIS RESULTS - {date} - RUN AT {timestamp}")
    output_logger.info("=" * 150)

    # Summary Table
    output_logger.info("")
    output_logger.info("📊 SUMMARY STATISTICS")
    output_logger.info("-" * 80)
    output_logger.info(f"{'Metric':<30} {'Value':<20} {'Unit':<10}")
    output_logger.info("-" * 80)
    output_logger.info(f"{'Total Blocks':<30} {results_summary.get('total_blocks', 0):<20} {'blocks':<10}")
    output_logger.info(f"{'Total Hybrid Schedule':<30} {results_summary.get('total_hybrid_schedule', 0):<20.1f} {'MW':<10}")
    output_logger.info(f"{'Total GDAM':<30} {results_summary.get('total_gdam', 0):<20.1f} {'MW':<10}")
    output_logger.info(f"{'Total Adjusted Schedule':<30} {results_summary.get('total_adjusted_schedule', 0):<20.1f} {'MW':<10}")
    output_logger.info(f"{'Total RTM Solar':<30} {results_summary.get('total_rtm_solar', 0):<20.1f} {'MW':<10}")
    if 'total_gdam_solar_output' in results_summary:
        output_logger.info(f"{'G-DAM Solar Output':<30} {results_summary.get('total_gdam_solar_output', 0):<20.1f} {'MW':<10}")
    output_logger.info(f"{'Total RTM Wind':<30} {results_summary.get('total_rtm_wind', 0):<20.1f} {'MW':<10}")
    if 'total_gdam_wind_output' in results_summary:
        output_logger.info(f"{'G-DAM Wind Output':<30} {results_summary.get('total_gdam_wind_output', 0):<20.1f} {'MW':<10}")
    output_logger.info("-" * 80)

    # Buyer Allocations Table
    buyer_allocations = results_summary.get('buyer_allocations', {})
    if buyer_allocations:
        output_logger.info("")
        output_logger.info("🏭 BUYER ALLOCATIONS")
        output_logger.info("-" * 100)
        output_logger.info(f"{'Buyer Name':<25} {'Solar (MW)':<15} {'Wind (MW)':<15} {'Total (MW)':<15} {'Percentage':<15}")
        output_logger.info("-" * 100)

        total_allocation = sum(alloc.get('total', 0) for alloc in buyer_allocations.values())

        for buyer, allocation in buyer_allocations.items():
            solar = allocation.get('solar', 0)
            wind = allocation.get('wind', 0)
            total = allocation.get('total', 0)
            percentage = (total / total_allocation * 100) if total_allocation > 0 else 0

            output_logger.info(f"{buyer:<25} {solar:<15.1f} {wind:<15.1f} {total:<15.1f} {percentage:<15.1f}%")

        output_logger.info("-" * 100)
        output_logger.info(f"{'TOTAL':<25} {sum(alloc.get('solar', 0) for alloc in buyer_allocations.values()):<15.1f} {sum(alloc.get('wind', 0) for alloc in buyer_allocations.values()):<15.1f} {total_allocation:<15.1f} {'100.0%':<15}")
        output_logger.info("-" * 100)

    # Detailed block-wise results if provided
    if detailed_results is not None and not detailed_results.empty:
        output_logger.info("")
        output_logger.info(f"📋 DETAILED BLOCK-WISE RESULTS (All {len(detailed_results)} blocks)")
        output_logger.info("-" * 150)

        # Show all blocks
        sample_results = detailed_results

        # Get buyer names from columns
        buyer_names = []
        for col in sample_results.columns:
            if '_Solar' in col:
                buyer_name = col.replace('_Solar', '')
                if buyer_name not in buyer_names:
                    buyer_names.append(buyer_name)

        # Create simplified header
        header = f"{'Block':<6}{'Solar':<8}{'Wind':<8}{'Total':<8}{'GDAM':<8}{'Adj':<8}"
        for buyer in buyer_names[:3]:  # Limit to first 3 buyers for readability
            header += f"{buyer.replace('Linde ', ''):<12}"
        show_gdam_output_columns = analysis_type == "DAY_AHEAD"
        header += f"{'RTM':<8}{'RTM_S':<8}{'RTM_W':<8}"
        if show_gdam_output_columns:
            header += f"{'GD_S':<8}{'GD_W':<8}"

        output_logger.info(header)
        output_logger.info("-" * 150)

        # Print data rows
        for _, row in sample_results.iterrows():
            data_row = f"{int(row['Block']):<6}"
            data_row += f"{row.get('Solar_Sch', 0):<8.1f}"
            data_row += f"{row.get('Wind_Sch', 0):<8.1f}"
            data_row += f"{row.get('Hyb_Sch', 0):<8.1f}"
            data_row += f"{row.get('GDAM', 0):<8.1f}"
            data_row += f"{row.get('Adjusted_Sch', 0):<8.1f}"

            # Add buyer totals (Solar + Wind)
            for buyer in buyer_names[:3]:
                solar_col = f"{buyer}_Solar"
                wind_col = f"{buyer}_Wind"
                buyer_total = row.get(solar_col, 0) + row.get(wind_col, 0)
                data_row += f"{buyer_total:<12.1f}"

            data_row += f"{row.get('RTM_Available', 0):<8.1f}"
            data_row += f"{row.get('RTM_Solar', 0):<8.1f}"
            data_row += f"{row.get('RTM_Wind', 0):<8.1f}"
            if show_gdam_output_columns:
                data_row += f"{to_float(row.get('G-DAM_Solar_Output', 0)):<8.1f}"
                data_row += f"{to_float(row.get('G-DAM_Wind_Output', 0)):<8.1f}"

            output_logger.info(data_row)

        output_logger.info("-" * 150)
        output_logger.info(f"... (showing all {len(detailed_results)} blocks)")
        output_logger.info("-" * 150)

        # Add a separate detailed buyer breakdown table for all blocks
        output_logger.info("")
        output_logger.info(f"🔍 BUYER BREAKDOWN (All {len(detailed_results)} blocks)")
        output_logger.info("-" * 120)

        buyer_header = f"{'Block':<6}"
        for buyer in buyer_names:
            buyer_header += f"{buyer.replace('Linde ', '') + '_Sol':<10}{buyer.replace('Linde ', '') + '_Wnd':<10}"

        output_logger.info(buyer_header)
        output_logger.info("-" * 120)

        # Show all blocks in buyer breakdown
        for _, row in sample_results.iterrows():
            buyer_row = f"{int(row['Block']):<6}"
            for buyer in buyer_names:
                solar_col = f"{buyer}_Solar"
                wind_col = f"{buyer}_Wind"
                buyer_row += f"{row.get(solar_col, 0):<10.1f}{row.get(wind_col, 0):<10.1f}"

            output_logger.info(buyer_row)

        output_logger.info("-" * 120)

    output_logger.info("")
    output_logger.info("=" * 150)
    output_logger.info(f"  END OF {analysis_type} ANALYSIS - {timestamp}")
    output_logger.info("=" * 150)
    output_logger.info("")

    # Clean up old log files (keep only 7 days)
    cleanup_old_output_logs(analysis_type)

def cleanup_old_output_logs(analysis_type):
    """Clean up output log files older than 7 days"""
    try:
        import pytz
        ist = pytz.timezone('Asia/Kolkata')
        cutoff_date = datetime.datetime.now(ist) - datetime.timedelta(days=7)
        pattern = f'output_logs/{analysis_type}_*.log'

        for log_file in glob.glob(pattern):
            # Extract date from filename
            filename = os.path.basename(log_file)
            date_str = filename.split('_')[1].split('.')[0]  # Extract YYYYMMDD

            try:
                file_date = datetime.datetime.strptime(date_str, '%Y%m%d')
                if file_date < cutoff_date:
                    os.remove(log_file)
                    print(f"Removed old log file: {log_file}")
            except ValueError:
                # Skip files that don't match the expected date format
                continue
    except Exception as e:
        print(f"Error cleaning up old log files: {e}")

# -------------------------------
# FTP Configuration & Functions
# -------------------------------

FTP_HOST = "15.207.32.135"
FTP_PORT = 21
FTP_USER = "partner"
FTP_PASS = "jEm9P6182x89"

# Remote FTP folder paths (Updated paths)
FTP_PATH_SCHEDULE = "/WIND/SCHEDULE/KA/Kudligi_HYB_FP/"
FTP_PATH_LOAD = "/WIND/Kudligi/IV_Partner/Load/"
FTP_PATH_OBLIGATION = "/WIND/Kudligi/IV_Partner/Obligation/"
FTP_PATH_GDAM_RTM_RATIO = "/WIND/Kudligi/IV_Partner/Obligation/GDAM_RTM_Ratio/"
FTP_PATH_FINAL_BUYER_SCHEDULE = "/WIND/SCHEDULE/KA/Kudligi_HYB_FP/Final_Buyer_Schedule/"
FTP_PATH_FINAL_BUYER_SCHEDULE_FALLBACK = "/arch/scada_2020/wind/WIND/SCHEDULE/KA/Kudligi_HYB_FP/Final_Buyer_Schedule"
FTP_PATH_RTM_TEMPLATE = "/WIND/Kudligi/IV_Partner/SCADA/"
FTP_PATH_REMC_TEMPLATE = "/WIND/SCHEDULE/KA/Kudligi_HYB_FP/REMC_TEMPLATE/"
FTP_PATH_LOW_PRIORITY_BUYERS = "/WIND/Kudligi/IV_Partner/Low_priority_Buyers"

# -------------------------------
# Email Configuration
# -------------------------------

FTP_LOW_PRIORITY_FILENAME = "Low_Priority_FP.xlsx"

SMTP_SERVER = "mx1.50hertz.in"
SMTP_PORT = 465
EMAIL_USER = "mal.rldc@manikarananalytics.in"
EMAIL_PASSWORD = "$%^M@L*@CP"

# Email recipients for RTM
EMAIL_TO = ["rtm@manikaranpowerltd.in", "rtmmpl@gmail.com"]
EMAIL_CC = [
    "bd.d1@manikaranpowerltd.in",
    "bd.d18@manikaranpowerltd.in",
    "indranil@manikaranpowerltd.in",
    "headops@manikarananalytics.in",
    "headoperations@manikaranpowerltd.in",
    "harsh.sharma@manikaranpowerltd.in",
    "shailendra.pandey@fourthpartner.co",
    "gaurav.srivastava@fourthpartner.co",
    "aman.yadav@fourthpartner.co",
    "priyajit.mallick@fourthpartner.co",
    "mal.rldc@manikarananalytics.in"
]

# Email recipients for REMC Templates
REMC_EMAIL_TO = ["mal.rldc@manikarananalytics.in"]
REMC_EMAIL_CC = [
    "shailendra.pandey@fourthpartner.co",
   "gaurav.srivastava@fourthpartner.co",
   "aman.yadav@fourthpartner.co",
   "priyajit.mallick@fourthpartner.co"
]

# Prevent accidental duplicate email sends from rapid repeated clicks/reruns.
EMAIL_SEND_COOLDOWN_SECONDS = 300

# RTM template file name
RTM_TEMPLATE_FILENAME = "S1ZE1MAN2483_Kudligi_HYB_FP_HIRIYUR_ZREPL_W RTM.xlsm"
WIND_RTM_PORTFOLIO_ID = "S1ZE1MAN2483"
SOLAR_RTM_PORTFOLIO_ID = "S1ZR1MAN2625"


def _get_email_lock_store():
    """Get/create session-scoped email send lock store."""
    if 'email_send_locks' not in st.session_state:
        st.session_state.email_send_locks = {}
    return st.session_state.email_send_locks


def get_email_cooldown_remaining(email_key):
    """Return remaining cooldown seconds for a specific email key."""
    lock_store = _get_email_lock_store()
    last_sent_ts = lock_store.get(email_key)

    if last_sent_ts is None:
        return 0

    elapsed = time.time() - float(last_sent_ts)
    return max(0, int(EMAIL_SEND_COOLDOWN_SECONDS - elapsed))


def mark_email_sent(email_key):
    """Mark an email key as sent now (starts cooldown)."""
    lock_store = _get_email_lock_store()
    lock_store[email_key] = time.time()
    st.session_state.email_send_locks = lock_store


def format_cooldown(seconds):
    """Format cooldown seconds as MM:SS."""
    minutes, secs = divmod(max(0, int(seconds)), 60)
    return f"{minutes:02d}:{secs:02d}"

def get_rtm_template_from_ftp():
    """
    Get RTM template file from FTP server into memory.

    Returns:
        BytesIO: File content in memory, or None if failed
    """
    try:
        from io import BytesIO

        # Connect to FTP
        ftp = connect_ftp()

        # Navigate to RTM template directory
        ftp.cwd(FTP_PATH_RTM_TEMPLATE)

        # Check if file exists on FTP
        files = ftp.nlst()
        if RTM_TEMPLATE_FILENAME not in files:
            st.error(f"❌ RTM template file not found on FTP: {FTP_PATH_RTM_TEMPLATE}{RTM_TEMPLATE_FILENAME}")
            ftp.quit()
            return None

        # Download file into memory
        st.info(f"📥 Loading RTM template from FTP: {FTP_PATH_RTM_TEMPLATE}{RTM_TEMPLATE_FILENAME}")
        file_data = BytesIO()
        ftp.retrbinary(f'RETR {RTM_TEMPLATE_FILENAME}', file_data.write)

        ftp.quit()

        # Reset file pointer to beginning
        file_data.seek(0)

        # Verify we got data
        if file_data.getbuffer().nbytes > 0:
            st.success(f"✅ RTM template loaded from FTP ({file_data.getbuffer().nbytes:,} bytes)")
            return file_data
        else:
            st.error(f"❌ RTM template is empty on FTP")
            return None
    except Exception as e:
        import traceback
        error_msg = f"Error loading RTM template from FTP: {str(e)}\n{traceback.format_exc()}"
        st.error(f"❌ {error_msg}")
        # Try to log if logger is available
        try:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(error_msg)
        except:
            pass
        return None


def get_rtm_portfolio_id(energy_label):
    energy_token = str(energy_label).strip().lower()
    if energy_token == "solar":
        return SOLAR_RTM_PORTFOLIO_ID
    return WIND_RTM_PORTFOLIO_ID


def replace_rtm_portfolio_in_workbook(workbook, target_portfolio_id):
    """
    Replace the wind portfolio placeholder across workbook cells for alternate
    RTM variants such as Solar, while leaving the original wind file unchanged.
    """
    source_portfolio_id = WIND_RTM_PORTFOLIO_ID
    target_portfolio_id = str(target_portfolio_id).strip()

    if not target_portfolio_id or target_portfolio_id == source_portfolio_id:
        return 0

    replacement_count = 0
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                cell_value = cell.value
                if isinstance(cell_value, str) and source_portfolio_id in cell_value:
                    cell.value = cell_value.replace(source_portfolio_id, target_portfolio_id)
                    replacement_count += 1

    return replacement_count

def get_low_priority_buyers_from_excel(ftp):
    """
    Get low priority buyers list from Excel file on FTP.
    Returns a list of buyer names.
    """
    try:
        from io import BytesIO
        import pandas as pd

        # Navigate to low priority buyers directory
        ftp.cwd(FTP_PATH_LOW_PRIORITY_BUYERS)

        # Check if file exists on FTP
        files = ftp.nlst()
        if FTP_LOW_PRIORITY_FILENAME not in files:
            st.warning(f"⚠️ Low priority buyers file not found on FTP: {FTP_PATH_LOW_PRIORITY_BUYERS}/{FTP_LOW_PRIORITY_FILENAME}. Defaulting to no low priority buyers.")
            return []

        # Download file into memory
        file_data = BytesIO()
        ftp.retrbinary(f'RETR {FTP_LOW_PRIORITY_FILENAME}', file_data.write)
        file_data.seek(0)

        # Read Excel file
        df = pd.read_excel(file_data)

        if 'Low Priority Buyers' not in df.columns:
            st.error(f"❌ Column 'Low Priority Buyers' not found in {FTP_LOW_PRIORITY_FILENAME}")
            return []

        # Get unique buyer names, drop NaNs
        low_priority_buyers = df['Low Priority Buyers'].dropna().unique().tolist()
        # Clean whitespace
        low_priority_buyers = [str(b).strip() for b in low_priority_buyers]

        return low_priority_buyers

    except Exception as e:
        st.error(f"❌ Error reading low priority buyers from FTP: {str(e)}")
        return []

def connect_ftp():
    ftp = FTP()
    ftp.connect(FTP_HOST, FTP_PORT)
    ftp.login(FTP_USER, FTP_PASS)

    return ftp

def download_ftp_file(ftp, remote_folder, remote_filename):
    """
    Downloads a file from the FTP server and returns a BytesIO stream.
    """
    ftp.cwd(remote_folder)
    bio = BytesIO()
    try:
        ftp.retrbinary(f"RETR {remote_filename}", bio.write)
        bio.seek(0)
        return bio
    except Exception as e:
        st.error(f"Error downloading {remote_filename} from {remote_folder}: {e}")
        return None

def upload_ftp_file(ftp, remote_folder, remote_filename, file_data):
    """
    Uploads a file to the FTP server.

    Args:
        ftp: FTP connection object
        remote_folder: Remote folder path
        remote_filename: Name of the file on the FTP server
        file_data: BytesIO stream or file-like object containing the data to upload

    Returns:
        bool: True if upload successful, False otherwise
    """
    try:
        ftp.cwd(remote_folder)
        file_data.seek(0)  # Ensure we're at the beginning of the file
        ftp.storbinary(f"STOR {remote_filename}", file_data)
        return True
    except Exception as e:
        st.error(f"Error uploading {remote_filename} to {remote_folder}: {e}")
        return False

def find_file_for_date(ftp, remote_folder, target_date, file_type='load'):
    """
    For Load and Obligation: target file is yyyymmdd.xlsx; if not available then yyyymmdd.csv.
    """
    ftp.cwd(remote_folder)
    target_filename_xlsx = target_date.strftime("%Y%m%d") + ".xlsx"
    target_filename_csv = target_date.strftime("%Y%m%d") + ".csv"

    files = ftp.nlst()
    if file_type in ['load', 'obligation']:
        if target_filename_xlsx in files:
            return target_filename_xlsx
        elif target_filename_csv in files:
            return target_filename_csv
    return None

def find_latest_modified_file(ftp, remote_folder, file_extensions=None):
    """
    Find the latest modified file in the given FTP folder.

    Args:
        ftp: FTP connection object
        remote_folder: Remote folder path
        file_extensions: List of allowed file extensions (e.g., ['.xlsx', '.csv'])
                        If None, all files are considered

    Returns:
        Filename of the latest modified file, or None if no files found
    """
    try:
        ftp.cwd(remote_folder)
    except Exception as e:
        st.error(f"FTP Error accessing {remote_folder}: {str(e)}")
        return None

    try:
        files = ftp.nlst()
        if not files:
            return None

        # Filter by file extensions if specified
        if file_extensions:
            files = [f for f in files if any(f.lower().endswith(ext.lower()) for ext in file_extensions)]

        if not files:
            return None

        # Get modification times for all files
        file_times = []
        for filename in files:
            try:
                # Get file modification time
                ftp.voidcmd("TYPE I")  # Set binary mode
                response = ftp.sendcmd(f"MDTM {filename}")
                # MDTM response format: "213 YYYYMMDDHHMMSS"
                time_str = response.split()[1]
                file_times.append((filename, time_str))
            except Exception:
                # If MDTM fails, skip this file
                continue

        if not file_times:
            # If MDTM is not supported, just return the first file
            return files[0]

        # Sort by modification time (newest first)
        file_times.sort(key=lambda x: x[1], reverse=True)
        return file_times[0][0]

    except Exception as e:
        st.error(f"Error finding latest file in {remote_folder}: {str(e)}")
        return None

def find_latest_revision_file(ftp, remote_folder, target_date, schedule_type="DA"):
    """
    Find the latest revision file for a given date and schedule type.

    Returns:
        tuple: (filename, revision_number) or (None, None) if not found
    """
    try:
        ftp.cwd(remote_folder)
    except Exception as e:
        st.error(f"FTP Error: {str(e)}")
        return None, None

    files = ftp.nlst()
    date_str_dot = target_date.strftime("%d.%m.%Y")
    date_str_dash = target_date.strftime("%d-%m-%Y")

    if schedule_type.upper() == "DA":
        patterns = [
            re.compile(
                r"Kudligi\s+HYB\s+FP_DA_" + re.escape(date_str_dot) +
                r"_DA\s+RDA(\d+)\.csv",
                re.IGNORECASE
            ),
            re.compile(
                r"Kudligi\s+ZREPL\s+SPS_DA_" + re.escape(date_str_dash) +
                r"_\s*DA\s+RDA(\d+)\.csv",
                re.IGNORECASE
            ),
        ]
    else:
        patterns = [
            re.compile(
                r"Kudligi\s+HYB\s+FP_IntraDay_" + re.escape(date_str_dot) +
                r"_IntraDay\s+RID(\d+)\.csv",
                re.IGNORECASE
            ),
            re.compile(
                r"Kudligi\s+ZREPL\s+SPS_IntraDay_" + re.escape(date_str_dash) +
                r"_\s*IntraDay\s+RID(\d+)\.csv",
                re.IGNORECASE
            ),
        ]

    max_revision = -1
    selected_file = None

    for f in files:
        for pattern in patterns:
            m = pattern.search(f)
            if not m:
                continue
            rev = int(m.group(1))
            if rev > max_revision:
                max_revision = rev
                selected_file = f
            break

    if selected_file and max_revision >= 0:
        return selected_file, str(max_revision)
    return None, None


def load_schedule_revision_dataframe(bio, filename):
    """Load a schedule revision file into a raw dataframe without assuming a fixed header."""
    bio.seek(0)
    filename_lower = str(filename).lower()

    if filename_lower.endswith(".csv"):
        raw_text = bio.read().decode("utf-8", errors="replace")
        csv_lines = raw_text.splitlines()
        parsed_rows = list(csv.reader(csv_lines))
        max_columns = max((len(row) for row in parsed_rows), default=0)
        padded_rows = [row + [""] * (max_columns - len(row)) for row in parsed_rows]
        return pd.DataFrame(padded_rows).astype(object)
    if filename_lower.endswith((".xlsx", ".xls")):
        return pd.read_excel(bio, header=None, engine="openpyxl")

    raise ValueError(f"Unsupported schedule file format: {filename}")


def build_schedule_component_df(target_date, blocks, schedules, avcs):
    """Create a normalized schedule dataframe for MySQL inserts."""
    component_df = pd.DataFrame(
        {
            "Date": [target_date] * len(blocks),
            "Block": pd.to_numeric(blocks, errors="coerce"),
            "Sch": pd.to_numeric(schedules, errors="coerce").fillna(0.0),
            "AvC": pd.to_numeric(avcs, errors="coerce").fillna(0.0),
        }
    )
    component_df.dropna(subset=["Block"], inplace=True)
    component_df["Block"] = component_df["Block"].astype(int)
    component_df = component_df[(component_df["Block"] >= 1) & (component_df["Block"] <= 96)].copy()
    component_df["Date"] = target_date
    component_df["Sch"] = component_df["Sch"].apply(to_float)
    component_df["AvC"] = component_df["AvC"].apply(to_float)
    component_df.reset_index(drop=True, inplace=True)
    return component_df


def parse_schedule_revision_components(bio, filename, target_date):
    """
    Parse schedule revision files into separate solar and wind schedule frames.

    Supports:
    - legacy wind-only revision files
    - new hybrid RID/RDA files where row 21 contains:
      H: Solar AvC, I: Solar Schedule, J: Wind AvC, K: Wind Schedule
    """
    df_raw = load_schedule_revision_dataframe(bio, filename)

    header_row_idx = find_template_row_index(df_raw, "Block")
    is_new_hybrid_format = False

    if header_row_idx is not None and len(df_raw.columns) >= 11:
        header_tokens = [
            normalize_template_token(df_raw.iloc[header_row_idx, col_idx])
            for col_idx in range(11)
        ]
        expected_tokens = [
            "block",
            "declaredplantschedule",
            "plantavc",
            "declaredhybridschedule",
            "hybridavc",
            "totalhybridsolar",
            "totalhybridwind",
            "avc",
            "schedule",
            "avc",
            "schedule",
        ]
        is_new_hybrid_format = header_tokens == expected_tokens

    if is_new_hybrid_format:
        data_start_idx = header_row_idx + 1
        data_end_idx = data_start_idx + 96
        data_df = df_raw.iloc[data_start_idx:data_end_idx, :11].copy()

        solar_df = build_schedule_component_df(
            target_date,
            data_df.iloc[:, 0],
            data_df.iloc[:, 8],
            data_df.iloc[:, 7],
        )
        solar_df["AvC"] = solar_df["AvC"].apply(
            lambda value: round(
                min(max(to_float(value), 0.0), SCHEDULE_SOLAR_AVC_OVERRIDE_MW),
                3,
            ) if to_float(value) > 0 else 0.0
        )

        wind_df = build_schedule_component_df(
            target_date,
            data_df.iloc[:, 0],
            data_df.iloc[:, 10],
            data_df.iloc[:, 9],
        )

        return {
            "format": "hybrid_r2",
            "solar_df": solar_df,
            "wind_df": wind_df,
        }

    # Legacy wind-only format fallback.
    if header_row_idx is None:
        header_row_idx = 11

    data_start_idx = header_row_idx + 1
    data_end_idx = data_start_idx + 96
    data_df = df_raw.iloc[data_start_idx:data_end_idx, :].copy()

    if data_df.shape[1] < 5:
        raise ValueError(
            f"Expected at least 5 columns in legacy schedule file, found {data_df.shape[1]}"
        )

    wind_avc_series = data_df.iloc[:, 5] if data_df.shape[1] > 5 else pd.Series([0] * len(data_df))
    wind_df = build_schedule_component_df(
        target_date,
        data_df.iloc[:, 0],
        data_df.iloc[:, 4],
        wind_avc_series,
    )

    return {
        "format": "legacy_wind_only",
        "solar_df": None,
        "wind_df": wind_df,
    }

# -------------------------------
# MySQL Connection & Table Creation
# -------------------------------

def get_db_connection(max_retries=3, retry_delay=2):
    """
    Get MySQL database connection with retry logic.

    Args:
        max_retries: Maximum number of connection attempts (default: 3)
        retry_delay: Delay in seconds between retries (default: 2)

    Returns:
        MySQL connection object

    Raises:
        mysql.connector.Error: If connection fails after all retries
    """
    import time

    last_error = None
    for attempt in range(1, max_retries + 1):
        try:
            conn = mysql.connector.connect(
                host='172.16.0.65',
                port=3306,
                user='root',
                password='271220021247@Admin',
                database='FP2',
                use_pure=True,
                connection_timeout=10,
                raise_on_warnings=False
            )
            # Test the connection
            conn.ping(reconnect=True, attempts=1, delay=0)
            return conn

        except mysql.connector.Error as e:
            last_error = e
            error_msg = f"Database connection attempt {attempt}/{max_retries} failed: {str(e)}"

            if attempt < max_retries:
                print(f"⚠️ {error_msg}")
                print(f"   Retrying in {retry_delay} seconds...")
                time.sleep(retry_delay)
            else:
                print(f"❌ {error_msg}")
                print(f"   All {max_retries} connection attempts failed.")

    # If we get here, all retries failed
    raise last_error

# -------------------------------
# Email Functions
# -------------------------------

def create_rtm_file(rtm_values, delivery_date, energy_label="Wind"):
    """
    Create RTM file by editing the template xlsm file.

    Args:
        rtm_values: List of 96 RTM values from bifurcation
        delivery_date: Date object for the delivery date (already adjusted by caller)
        energy_label: Label used for output naming/debugging

    Returns:
        Path to the created file, or None if failed
    """
    try:
        portfolio_id = get_rtm_portfolio_id(energy_label)

        # Get the template file from FTP
        template_data = get_rtm_template_from_ftp()

        if template_data is None:
            st.error(f"❌ Failed to load RTM template from FTP: {FTP_PATH_RTM_TEMPLATE}{RTM_TEMPLATE_FILENAME}")
            return None

        # Load the template file from memory
        wb = openpyxl.load_workbook(template_data, keep_vba=True)
        ws = wb.active

        replacements_made = replace_rtm_portfolio_in_workbook(wb, portfolio_id)
        if replacements_made > 0:
            st.info(
                f"🔁 Updated {energy_label} RTM workbook portfolio to {portfolio_id} "
                f"({replacements_made} cell{'s' if replacements_made != 1 else ''})"
            )

        # Use the delivery date as provided (caller has already handled time-based logic)
        st.info(f"📅 Using delivery date: {delivery_date.strftime('%d/%m/%Y')}")

        # Set delivery date in cell D12 (DD/MM/YYYY format)
        date_str = delivery_date.strftime('%d/%m/%Y')
        ws['D12'] = date_str

        # Set RTM values in cells J21 to J116 (96 cells)
        if len(rtm_values) != 96:
            st.error(f"❌ Expected 96 {energy_label} RTM values, got {len(rtm_values)}")
            return None

        for i, value in enumerate(rtm_values, start=21):
            cell = f'J{i}'
            # Round to 1 decimal point
            ws[cell] = round(float(value), 1)

        # Save the edited file in a unique temp directory so multiple RTM files
        # can coexist in the same run without overwriting each other.
        safe_energy_label = re.sub(r"[^A-Za-z0-9]+", "_", str(energy_label).strip()) or "RTM"
        output_filename = f"RTM_Bid_{safe_energy_label}_{portfolio_id}_{delivery_date.strftime('%d-%m-%Y')}.xlsm"
        unique_temp_dir = tempfile.mkdtemp(prefix=f"rtm_{safe_energy_label.lower()}_")
        output_path = os.path.join(unique_temp_dir, output_filename)
        wb.save(output_path)

        return output_path

    except Exception as e:
        import traceback
        error_msg = f"Error creating {energy_label} RTM file: {str(e)}\n{traceback.format_exc()}"
        st.error(f"❌ {error_msg}")
        # Try to log if logger is available in scope
        try:
            import logging
            logger = logging.getLogger(__name__)
            logger.error(error_msg)
        except:
            pass
        return None


def send_rtm_email(attachment_path, delivery_date, logger):
    """
    Send RTM bid email with attachment.

    Args:
        attachment_path: Path to the RTM file to attach
        delivery_date: Date object for the delivery date
        logger: Logger instance for logging

    Returns:
        True if email sent successfully, False otherwise
    """
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = EMAIL_USER
        msg['To'] = ', '.join(EMAIL_TO)
        msg['Cc'] = ', '.join(EMAIL_CC)
        msg['Subject'] = f"RTM Bid - Zenataris Renewable Energy Pvt Ltd (S1ZE1MAN2483) - {delivery_date.strftime('%d-%m-%Y')}"

        # Email body
        body = f"""Respected Sir,

Please find the attached RTM bid of Zenataris Renewable Energy Private Limited (S1ZE1MAN2483) for date {delivery_date.strftime('%d-%m-%Y')}.

Best Regards,
Bifurcation System
"""

        msg.attach(MIMEText(body, 'plain'))

        # Attach the RTM file
        if not os.path.exists(attachment_path):
            st.error(f"❌ Attachment file not found: {attachment_path}")
            logger.error(f"Attachment file not found: {attachment_path}")
            return False

        with open(attachment_path, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {os.path.basename(attachment_path)}'
            )
            msg.attach(part)

        # Send email using SSL
        logger.info(f"Connecting to SMTP server: {SMTP_SERVER}:{SMTP_PORT}")
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
            server.login(EMAIL_USER, EMAIL_PASSWORD)

            logger.info(f"Sending email to: {', '.join(EMAIL_TO)}")
            logger.info(f"CC: {', '.join(EMAIL_CC)}")

            server.send_message(msg)

        logger.info("✅ RTM email sent successfully")
        return True

    except Exception as e:
        error_msg = f"❌ Error sending email: {str(e)}"
        st.error(error_msg)
        logger.error(error_msg)
        return False


def send_remc_template_email(attachment_path, file_type, effective_date, revision_number, logger):
    """
    Send REMC template email with attachment.

    Args:
        attachment_path: Path to the REMC template file to attach
        file_type: 'Intraday' or 'Day Ahead'
        effective_date: Date object for the effective date
        revision_number: Revision number extracted from filename (e.g., 11 from R-11)
        logger: Logger instance for logging

    Returns:
        True if email sent successfully, False otherwise
    """
    try:
        # Create message
        msg = MIMEMultipart()
        msg['From'] = EMAIL_USER
        msg['To'] = ', '.join(REMC_EMAIL_TO)
        msg['Cc'] = ', '.join(REMC_EMAIL_CC)

        # Subject format: "Intraday R -11  Schedule for Pooling station HIRIYUR_ZREPL_W Dt-06-01-2026"
        msg['Subject'] = f"{file_type} R -{revision_number}  Schedule for Pooling station HIRIYUR_ZREPL_W Dt-{effective_date.strftime('%d-%m-%Y')}"

        # Email body
        body = f"""Respected Sir,

Please find the attached Schedule {file_type}  R-{revision_number}  HIRIYUR_ZREPL_W effective for {effective_date.strftime('%d.%m.%Y')}.

Best Regards,
Shift In-Charge
Manikaran Analytics Limited
"""

        msg.attach(MIMEText(body, 'plain'))

        # Attach the REMC template file
        if not os.path.exists(attachment_path):
            st.error(f"❌ Attachment file not found: {attachment_path}")
            logger.error(f"Attachment file not found: {attachment_path}")
            return False

        with open(attachment_path, 'rb') as attachment:
            part = MIMEBase('application', 'octet-stream')
            part.set_payload(attachment.read())
            encoders.encode_base64(part)
            part.add_header(
                'Content-Disposition',
                f'attachment; filename= {os.path.basename(attachment_path)}'
            )
            msg.attach(part)

        # Send email using SSL
        logger.info(f"Connecting to SMTP server: {SMTP_SERVER}:{SMTP_PORT}")
        with smtplib.SMTP_SSL(SMTP_SERVER, SMTP_PORT) as server:
            server.login(EMAIL_USER, EMAIL_PASSWORD)

            logger.info(f"Sending REMC template email to: {', '.join(REMC_EMAIL_TO)}")
            logger.info(f"CC: {', '.join(REMC_EMAIL_CC)}")

            server.send_message(msg)

        logger.info(f"✅ REMC {file_type} template email sent successfully")
        return True

    except Exception as e:
        error_msg = f"❌ Error sending REMC template email: {str(e)}"
        st.error(error_msg)
        logger.error(error_msg)
        return False



# -------------------------------
# REMC Portal Automation
# -------------------------------

def remc_portal_login(page, logger):
    """
    Performs login to the REMC portal.
    
    Args:
        page: Playwright page object
        logger: Logger instance
        
    Returns:
        bool: True if login successful, False otherwise
    """
    USERNAME = "HIRIYUR_ZREPL_W"
    PASSWORD = "a991bf"
    
    try:
        # 1. Login Logic
        logger.info("Navigating to login page...")
        st.info("Navigating to REMC Login page...")

        # Use 'domcontentloaded' instead of 'networkidle' for faster loading
        page.goto("https://remc.srldc.in/index.php/Login", wait_until="domcontentloaded", timeout=30000)
        logger.info("Login page loaded")

        # Wait for login form
        page.wait_for_selector("#username", timeout=15000)
        logger.info("Login form found")

        # Fill credentials with human-like typing
        logger.info("Filling username...")
        page.click("#username")  # Click to focus
        page.type("#username", USERNAME, delay=100)
        time.sleep(0.5)

        logger.info("Filling password...")
        page.click("#passowrd")  # Note: ID is 'passowrd'
        page.type("#passowrd", PASSWORD, delay=100)
        time.sleep(0.5)
        logger.info("Credentials filled")

        # Solve Captcha
        captcha_text = page.locator("#question").inner_text()
        logger.info(f"Login Captcha: {captcha_text}")
        st.info(f"Solving Login Captcha: {captcha_text}")

        # Extract numbers and calculate
        parts = re.findall(r'\d+', captcha_text)
        if len(parts) >= 2:
            ans = int(parts[0]) + int(parts[1])
            logger.info(f"Calculated answer: {ans}")

            # Fill captcha answer
            logger.info("Filling captcha answer...")
            page.click("#ans")
            time.sleep(0.3)
            page.type("#ans", str(ans), delay=150)
            logger.info(f"Captcha answer typed: {ans}")

            # Trigger blur event to validate
            page.evaluate("document.querySelector('#ans').blur()")
            time.sleep(0.5)

            # Trigger any validation events
            page.evaluate("""
                const ansInput = document.querySelector('#ans');
                if (ansInput) {
                    ansInput.dispatchEvent(new Event('input', { bubbles: true }));
                    ansInput.dispatchEvent(new Event('change', { bubbles: true }));
                    ansInput.dispatchEvent(new Event('blur', { bubbles: true }));
                }
            """)

            logger.info("Waiting for verification...")
            time.sleep(2)  # Wait for validation to complete
        else:
            logger.warning("Could not parse login captcha, trying to proceed manually or expecting failure")

        # Submit Login by clicking the login button
        try:
            # Wait for login button to be enabled
            page.wait_for_selector("#btn_login:not([disabled])", timeout=5000)
            logger.info("Login button is enabled, clicking...")
            page.click("#btn_login")
            logger.info("✅ Login button clicked")
        except Exception as btn_error:
            logger.warning(f"Could not click login button: {btn_error}, trying Enter key")
            page.press("#ans", "Enter")
            logger.info("Login submitted via Enter key")

        time.sleep(3)  # Wait for login

        # Verify login was successful
        current_url = page.url
        logger.info(f"After login, current URL: {current_url}")

        if "Login" in current_url or "login" in current_url:
            logger.warning("Still on login page - login may have failed")
            st.warning("⚠️ Login may have failed - still on login page")
            return False
        else:
            logger.info("✅ Login successful - navigated away from login page")
            st.success("✅ Login successful")
            return True
            
    except Exception as e:
        logger.error(f"Login process failed: {str(e)}")
        st.error(f"❌ Login failed: {str(e)}")
        return False

def download_remc_templates_from_portal(logger):
    """
    Downloads DAY-AHEAD and INTRADAY templates from REMC portal and uploads them to FTP.
    Does NOT use headless mode so user can see it happening if they want.
    """
    logger.info("Starting automated template download from REMC portal")
    st.info("🤖 Starting automated REMC template download...")
    
    # Check if templates already exist on FTP to avoid unnecessary portal login
    try:
        import pytz
        ist = pytz.timezone('Asia/Kolkata')
        current_time_ist = datetime.datetime.now(ist)
        today_date_str = current_time_ist.strftime("%d-%m-%Y")
        tomorrow_date_str = (current_time_ist + datetime.timedelta(days=1)).strftime("%d-%m-%Y")
        
        # Expected filenames
        # Note: These must match what the portal generates or what we expect
        # Usually format is : HIRIYUR_ZREPL_W_Intraday_DD-MM-YYYY.csv
        # We need to check for BOTH Intraday and DayAhead for relevant dates
        
        # For simplicity, if we see ANY relevant file for today, we might skip, 
        # BUT the user said "only fetch if template files not found".
        # Let's check specifically for the files we are about to generate.
        
        # Intraday is for today
        id_filename = f"HIRIYUR_ZREPL_W_Intraday_{today_date_str}.csv"
        # Day Ahead is for tomorrow
        da_filename = f"HIRIYUR_ZREPL_W_DayAhead_{tomorrow_date_str}.csv" 
        # Note: Portal might name it DayAhead or DA or Day-Ahead. 
        # Based on script later: f"HIRIYUR_ZREPL_W_{file_type}_{target_date}.csv" where file_type is 'DA' or 'Intraday'
        # Wait, the script uses 'DA' for file_type in one place but 'DayAhead' might be portal default?
        # Let's check listing.
        
        ftp = connect_ftp()
        try:
            ftp.cwd(FTP_PATH_REMC_TEMPLATE)
            files_on_ftp = ftp.nlst()
            
            # Check for rough matches if exact names are uncertain
            # The script later uses: output_filename = f"HIRIYUR_ZREPL_W_{file_type}_{target_date.strftime('%d-%m-%Y')}.csv"
            # where file_type is "Intraday" or "Day_Ahead" (passed as arg).
            # The user request said: /WIND/SCHEDULE/KA/Kudligi_HYB_FP/REMC_TEMPLATE/
            
            # Simple check: do we have ANY file for today/tomorrow?
            # Or strict check. Let's do a smart check.
            
            # Intraday: contains 'Intraday' or '_ID_' AND today's date
            id_exists = any((today_date_str in f) and ("Intraday" in f or "_ID_" in f) for f in files_on_ftp)
            
            # Day Ahead: contains 'DayAhead' or '_DA_' or 'DA' AND (today's date OR tomorrow's date)
            # The download function seems to timestamp with download time (today), regardless of "valid for" date.
            da_exists = any(((today_date_str in f) or (tomorrow_date_str in f)) and ("DayAhead" in f or "_DA_" in f or " DA " in f) for f in files_on_ftp)
            
            if id_exists and da_exists:
                logger.info("✅ Valid template files found on FTP. Skipping portal download.")
                st.info("✅ Valid template files found on FTP. Skipping portal download.")
                ftp.quit()
                return True
            else:
                logger.info(f"Templates missing on FTP (ID: {id_exists}, DA: {da_exists}). Proceeding to download.")
                st.info("ℹ️ Latest templates not found on FTP. Downloading from portal...")
                
        except Exception as e:
            logger.warning(f"Could not check FTP for existing files: {e}. Proceeding with download.")
            
        ftp.quit()
        
    except Exception as e:
        logger.warning(f"Pre-download FTP check failed: {e}")

    playwright = None
    browser = None
    
    try:
        playwright = sync_playwright().start()
        
        # Use existing env vars for headless config
        headless_mode = os.getenv('PLAYWRIGHT_HEADLESS', 'true').lower() == 'true'
        slow_mo = int(os.getenv('PLAYWRIGHT_SLOWMO', '0'))
        
        browser = playwright.chromium.launch(
            headless=headless_mode,
            slow_mo=slow_mo,
            args=['--no-sandbox', '--disable-setuid-sandbox']
        )
        
        context = browser.new_context(viewport={'width': 1920, 'height': 1080})
        page = context.new_page()
        page.set_default_timeout(60000)
        
        # Login
        if not remc_portal_login(page, logger):
            logger.error("Login failed, aborting template download")
            return False
            
        # Function to download specific template type
        def download_and_upload_template(template_type):
            """
            template_type: "DA" or "INTRADAY"
            """
            try:
                type_name = "Day Ahead" if template_type == "DA" else "Intraday"
                logger.info(f"Downloading {type_name} template...")
                st.info(f"📥 Fetching {type_name} template...")
                
                # Navigate to Prepare Schedule using the 'Prepare Schedule' pill/tab
                # We assume we are on the 'Schedule' -> 'By Template' page or need to get there
                # Direct navigation is safest
                page.goto("https://remc.srldc.in/index.php/submitSchedule/scheduleByTemplate", wait_until="domcontentloaded")
                
                # Click "Prepare Schedule" tab: <a data-toggle="pill" href="#util">Prepare Schedule</a>
                logger.info("Clicking Prepare Schedule tab...")
                page.click("a[data-toggle='pill'][href='#util']")
                page.wait_for_selector("#util", state="visible", timeout=5000)
                
                # Select Intraday or Day Ahead
                # <select class="form-control" id="revision_no" name="revision_no" required="">
                logger.info(f"Selecting {template_type} from dropdown...")
                page.select_option("#revision_no", template_type)
                time.sleep(1) # Wait for UI update
                
                # Check "Select All" posfact
                # <input id="posfact" type="checkbox" class="selectallpos_names_list" value="selall">
                logger.info("Checking 'Select All' checkbox...")
                # Ensure it's checked. If not, click it.
                # Use force=True because a span might be intercepting the click
                if not page.is_checked("#posfact"):
                    page.click("#posfact", force=True)
                    logger.info("Checked 'Select All' (forced)")
                
                # Select Download Type: MTOA
                # <input type="radio" class="contracttype" value="wod" id="MTOA" name="download_type" required="">
                logger.info("Selecting 'MTOA' download type...")
                page.click("#MTOA", force=True)
                
                # Select AVC Validation: Yes
                # <input type="radio" class="validatetype" id="Yes" name="avc_validation" value="yes" required="" checked="">
                logger.info("Selecting 'Yes' for AVC validation...")
                page.click("#Yes", force=True) # It might be checked by default, but clicking ensures it
                
                # Click Submit and wait for download
                # <button class="btn btn-primary">Submit</button>
                logger.info("Clicking Submit to download...")
                
                with page.expect_download(timeout=60000) as download_info:
                    # Find the submit button inside the Prepare Schedule tab content check visibility
                    page.click("#util .btn-primary")
                    
                download = download_info.value
                download_path = download.path()
                suggested_filename = download.suggested_filename
                
                logger.info(f"Downloaded file: {suggested_filename}")
                st.success(f"✅ Downloaded {type_name} template: {suggested_filename}")
                
                # Read file content for upload
                with open(download_path, 'rb') as f:
                    file_data = BytesIO(f.read())
                
                # Upload to FTP
                ftp = connect_ftp()
                
                # Determine remote filename based on type to match what our script expects
                # The script looks for _ID_ or _DA_ in filenames in FTP_PATH_REMC_TEMPLATE
                # We should upload with the filename from portal which usually contains these patterns
                
                # Ensure filename indicates type clearly if portal name is ambiguous (unlikely)
                remote_filename = suggested_filename
                
                
                logger.info(f"Uploading to FTP: {FTP_PATH_REMC_TEMPLATE}{remote_filename}")
                
                # Check if we need to change directory first or if upload_ftp_file handles full paths
                # Based on the error "Template files not found", the standard upload might be putting it in root or wrong place.
                # Let's ensure we are in the correct directory.
                
                # Try to clean up path and upload
                target_dir = FTP_PATH_REMC_TEMPLATE
                if target_dir.endswith('/'):
                    target_dir = target_dir[:-1]
                
                # Navigate to directory to be safe
                try:
                    ftp.cwd(target_dir)
                    logger.info(f"Changed FTP directory to {target_dir}")
                    # Now upload with just filename since we are in the dir
                    ftp.storbinary(f"STOR {remote_filename}", file_data)
                    st.success(f"✅ Uploaded to FTP: {remote_filename}")
                    logger.info("Upload successful")
                    success = True
                except Exception as upload_err:
                     # Fallback to full path upload if CWD fails
                    logger.warning(f"CWD failed, trying full path upload: {upload_err}")
                    file_data.seek(0)
                    ftp.storbinary(f"STOR {FTP_PATH_REMC_TEMPLATE}/{remote_filename}".replace('//', '/'), file_data)
                    success = True

                # if upload_ftp_file(ftp, FTP_PATH_REMC_TEMPLATE, remote_filename, file_data):
                #     st.success(f"✅ Uploaded to FTP: {remote_filename}")
                #     logger.info("Upload successful")
                #     success = True
                # else:
                #     st.error(f"❌ FTP Upload failed for {remote_filename}")
                #     success = False
                    
                ftp.quit()
                return success
                
            except Exception as e:
                logger.error(f"Error downloading {template_type} template: {str(e)}")
                st.error(f"❌ Failed to download {type_name} template: {str(e)}")
                return False

        # Execute for both types
        da_success = download_and_upload_template("DA")
        id_success = download_and_upload_template("INTRADAY")
        
        if da_success and id_success:
            st.success("✅ All templates updated from portal successfully!")
            return True
        else:
            st.warning("⚠️ Some templates failed to update.")
            return False
            
            
    except Exception as e:
        logger.error(f"Template download process crashed: {str(e)}")
        st.error(f"❌ Template automation failed: {str(e)}")
        return False
        
    finally:
        if browser:
            browser.close()
        if playwright:
            playwright.stop()

def submit_remc_schedule(file_path, schedule_type, logger):
    """
    Submits the REMC schedule to the REMC portal using Playwright.

    Args:
        file_path: Absolute path to the schedule file
        schedule_type: 'Intraday' or 'Day Ahead'
        logger: Logger instance

    Returns:
        bool: True if submission (or attemp) was successful, False otherwise
    """
    logger.info(f"Starting REMC portal submission for {schedule_type}")
    logger.info(f"File to upload: {file_path}")
    logger.info(f"File exists: {os.path.exists(file_path)}")
    if os.path.exists(file_path):
        logger.info(f"File size: {os.path.getsize(file_path)} bytes")

        # Read and log first few schedule values for verification
        try:
            import pandas as pd
            df = pd.read_csv(file_path, header=None)
            # Row 17 (index 16) is the first schedule row (Block 1)
            if len(df) > 16:
                block1_values = df.iloc[16, 3:].tolist()  # Columns 3+ are buyer schedules
                logger.info(f"Block 1 schedule values: {block1_values[:5]}")  # First 5 buyers
                st.info(f"🔍 Block 1 values: {block1_values[:5]}")
        except Exception as e:
            logger.warning(f"Could not read file for verification: {e}")

    st.info(f"🤖 Starting automated REMC portal submission for {schedule_type}...")
    st.info(f"📁 Uploading file: {os.path.basename(file_path)}")

    # Portal Credentials
    USERNAME = "HIRIYUR_ZREPL_W"
    PASSWORD = "a991bf"

    playwright = None
    browser = None
    page = None
    try:
        st.info("Initializing Playwright browser...")

        # Start Playwright
        playwright = sync_playwright().start()

        # Check if headless mode should be disabled (via environment variable)
        # Set PLAYWRIGHT_HEADLESS=false to see the browser
        # Default is TRUE (headless) for faster performance
        headless_mode = os.getenv('PLAYWRIGHT_HEADLESS', 'true').lower() == 'true'

        # Check if slow motion should be enabled (for debugging)
        # Set PLAYWRIGHT_SLOWMO=1000 to slow down actions by 1000ms
        slow_mo = int(os.getenv('PLAYWRIGHT_SLOWMO', '0'))

        logger.info(f"Launching browser in {'headless' if headless_mode else 'non-headless'} mode")

        # Launch browser with configurable headless mode
        browser = playwright.chromium.launch(
            headless=headless_mode,
            slow_mo=slow_mo,  # Slow down actions for debugging
            args=[
                '--no-sandbox',
                '--disable-setuid-sandbox',
                '--disable-dev-shm-usage',
                '--disable-blink-features=AutomationControlled',
                '--disable-gpu',  # Disable GPU for faster headless performance
                '--disable-software-rasterizer'
            ]
        )

        # Create browser context
        context = browser.new_context(
            viewport={'width': 1920, 'height': 1080},
            user_agent='Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36',
            ignore_https_errors=True
        )

        # Create page
        page = context.new_page()
        page.set_default_timeout(60000)  # 60 seconds timeout

        st.success("✅ Playwright browser created successfully!")
        
        st.success("✅ Playwright browser created successfully!")
        
        # 1. Login using helper
        if not remc_portal_login(page, logger):
             raise Exception("Login failed")
        
        # 2. Navigate to Upload Schedule (Multi-step navigation)
        logger.info("Navigating to Upload Schedule page...")
        st.info("Navigating to Upload Schedule page...")

        # Step 2a: Click "Schedule" menu (the dropdown with arrow)
        try:
            logger.info("Step 1: Looking for 'Schedule' dropdown menu...")
            # Use first occurrence to avoid "strict mode violation" with multiple matches
            schedule_menu = page.locator("a.label-nav:has-text('Schedule')").first
            schedule_menu.wait_for(state="visible", timeout=10000)
            logger.info("Found 'Schedule' menu, clicking...")
            schedule_menu.click(force=True)
            logger.info("✅ Clicked 'Schedule' dropdown menu")
            time.sleep(1)  # Wait for dropdown to expand
        except Exception as menu_error:
            logger.warning(f"Could not click Schedule menu: {menu_error}")
            # Try alternative selector
            try:
                page.locator("a.label-nav").filter(has_text="Schedule").first.click(force=True)
                logger.info("✅ Clicked 'Schedule' menu (alternative selector)")
                time.sleep(1)
            except:
                logger.error("Failed to click Schedule menu with all selectors")

        # Step 2b: Click "By Template" submenu (now visible in dropdown)
        try:
            logger.info("Step 2: Looking for 'By Template' submenu in dropdown...")
            # Wait for the submenu to appear after dropdown opens
            template_link_selector = "a.label-nav[href='https://remc.srldc.in/index.php/submitSchedule/scheduleByTemplate']"
            template_link = page.locator(template_link_selector)
            template_link.wait_for(state="visible", timeout=10000)
            logger.info("Found 'By Template' link, clicking...")
            template_link.click(force=True)
            logger.info("✅ Clicked 'By Template' link")

            # Wait for page to load (use domcontentloaded for faster loading)
            page.wait_for_load_state("domcontentloaded")
            logger.info("Template page loaded")
            time.sleep(0.5)  # Reduced wait time
        except Exception as template_error:
            logger.warning(f"Could not click By Template: {template_error}")
            # Try alternative: direct navigation
            try:
                logger.info("Trying direct navigation to template page...")
                page.goto("https://remc.srldc.in/index.php/submitSchedule/scheduleByTemplate", wait_until="domcontentloaded", timeout=30000)
                logger.info("✅ Navigated directly to template page")
            except:
                logger.error("Failed to reach template page")
                raise

        # Step 2c: Click "Upload Schedule" pill/tab
        try:
            logger.info("Step 3: Looking for 'Upload Schedule' tab...")
            # Use the exact selector from the HTML
            upload_tab_selector = "a[data-toggle='pill'][href='#upload_bulk']"
            page.wait_for_selector(upload_tab_selector, state="visible", timeout=10000)
            logger.info("Found 'Upload Schedule' tab, clicking...")
            page.click(upload_tab_selector)
            logger.info("✅ Clicked 'Upload Schedule' tab")

            # Wait for the tab content to be visible
            page.wait_for_selector("#upload_bulk", state="visible", timeout=5000)
            logger.info("✅ Upload tab content is now visible")
            time.sleep(1)  # Wait for any animations

        except Exception as tab_error:
            logger.error(f"Could not click Upload Schedule tab: {tab_error}")
            # Try alternative selectors
            try:
                alt_selectors = [
                    "a[href='#upload_bulk']",
                    "//a[contains(text(), 'Upload Schedule')]"
                ]
                for sel in alt_selectors:
                    if page.locator(sel).count() > 0:
                        page.click(sel)
                        logger.info(f"✅ Clicked upload tab with alternative: {sel}")
                        time.sleep(1)
                        break
            except:
                logger.error("Failed to click upload tab with all selectors")
                raise

        # 3. Upload File
        logger.info(f"Uploading file: {file_path}")
        st.info(f"Uploading file: {os.path.basename(file_path)}")

        # Debug: Check for all file inputs on the page
        try:
            all_file_inputs = page.locator("input[type='file']").all()
            logger.info(f"Found {len(all_file_inputs)} file input(s) on page")
            for i, inp in enumerate(all_file_inputs):
                try:
                    inp_id = inp.get_attribute("id")
                    inp_name = inp.get_attribute("name")
                    inp_visible = inp.is_visible()
                    logger.info(f"  File input {i+1}: id='{inp_id}', name='{inp_name}', visible={inp_visible}")
                except:
                    pass
        except Exception as debug_error:
            logger.warning(f"Could not debug file inputs: {debug_error}")

        # Wait for file input to be available
        try:
            # Try multiple approaches to find and use the file input
            file_uploaded = False

            # Approach 1: Wait for #prepared_schedule
            try:
                page.wait_for_selector("#prepared_schedule", state="attached", timeout=5000)
                logger.info("File input #prepared_schedule found")

                # Try to make it visible if hidden
                page.evaluate("""
                    const input = document.querySelector('#prepared_schedule');
                    if (input) {
                        input.style.display = 'block';
                        input.style.visibility = 'visible';
                        input.style.opacity = '1';
                    }
                """)

                page.locator("#prepared_schedule").set_input_files(file_path)
                logger.info("✅ File uploaded using #prepared_schedule")
                file_uploaded = True

            except Exception as approach1_error:
                logger.warning(f"Approach 1 failed: {approach1_error}")

            # Approach 2: Use name attribute
            if not file_uploaded:
                try:
                    page.wait_for_selector("input[name='prepared_schedule']", state="attached", timeout=5000)
                    page.locator("input[name='prepared_schedule']").set_input_files(file_path)
                    logger.info("✅ File uploaded using name='prepared_schedule'")
                    file_uploaded = True
                except Exception as approach2_error:
                    logger.warning(f"Approach 2 failed: {approach2_error}")

            # Approach 3: Use any visible file input
            if not file_uploaded:
                try:
                    file_inputs = page.locator("input[type='file']").all()
                    for inp in file_inputs:
                        if inp.is_visible() or True:  # Try even if not visible
                            inp.set_input_files(file_path)
                            logger.info("✅ File uploaded using first available file input")
                            file_uploaded = True
                            break
                except Exception as approach3_error:
                    logger.warning(f"Approach 3 failed: {approach3_error}")

            if not file_uploaded:
                raise Exception("Could not upload file using any method")

            st.success(f"✅ File {os.path.basename(file_path)} uploaded")
            time.sleep(1)

        except Exception as upload_error:
            logger.error(f"Failed to upload file: {upload_error}")
            # Take screenshot for debugging
            try:
                page.screenshot(path="upload_error.png", full_page=True)
                logger.info("Screenshot saved to upload_error.png")
                st.warning("Screenshot saved to upload_error.png")
            except:
                pass
            raise upload_error

        # 4. Solve Captcha 2
        try:
            # Wait for captcha to appear - using the exact selector from the HTML
            logger.info("Waiting for upload captcha...")
            page.wait_for_selector("div#question.question", state="visible", timeout=10000)

            # Get captcha text
            captcha_text_2 = page.locator("div#question.question").inner_text()
            logger.info(f"Upload Captcha: {captcha_text_2}")
            st.info(f"Solving Upload Captcha: {captcha_text_2}")

            # Extract numbers and calculate
            parts_2 = re.findall(r'\d+', captcha_text_2)
            if len(parts_2) >= 2:
                ans_2 = int(parts_2[0]) + int(parts_2[1])
                logger.info(f"Captcha answer calculated: {ans_2}")

                # Wait for answer input to be ready
                page.wait_for_selector("input#ans", state="visible", timeout=5000)

                # Fill the answer with human-like typing
                # Note: There might be multiple #ans inputs (one for login, one for upload)
                # We need to fill the visible one in the upload tab
                ans_input = page.locator("#upload_bulk input#ans")
                if ans_input.count() > 0 and ans_input.is_visible():
                    logger.info("Using upload tab specific answer input")
                    ans_input.click()  # Focus the field
                    time.sleep(0.3)
                    ans_input.type(str(ans_2), delay=150)  # Type with human-like delay
                    logger.info(f"Typed captcha answer: {ans_2}")
                else:
                    # Fallback to any visible #ans input
                    logger.info("Using general answer input")
                    visible_ans = page.locator("input#ans").filter(has_text="").first
                    visible_ans.click()
                    time.sleep(0.3)
                    visible_ans.type(str(ans_2), delay=150)
                    logger.info(f"Typed captcha answer: {ans_2}")

                # Trigger validation events (like clicking background)
                logger.info("Triggering validation events...")
                page.evaluate("""
                    const ansInput = document.querySelector('#upload_bulk input#ans') || document.querySelector('input#ans');
                    if (ansInput) {
                        ansInput.dispatchEvent(new Event('input', { bubbles: true }));
                        ansInput.dispatchEvent(new Event('change', { bubbles: true }));
                        ansInput.dispatchEvent(new Event('blur', { bubbles: true }));
                    }
                """)

                # Click somewhere else to trigger blur (like clicking background)
                try:
                    page.evaluate("document.body.click()")
                    logger.info("Clicked background to trigger validation")
                except:
                    pass

                time.sleep(2)  # Wait for validation to complete
                logger.info("✅ Captcha answer filled and validated")
                st.success(f"✅ Captcha solved: {ans_2}")

            else:
                logger.warning(f"Could not parse captcha: {captcha_text_2}")
                st.warning("⚠️ Could not parse captcha automatically")

        except Exception as captcha_error:
            logger.error(f"Error solving captcha: {captcha_error}")
            # Take screenshot for debugging
            try:
                screenshot_path = "captcha_error.png"
                page.screenshot(path=screenshot_path)
                logger.info(f"Screenshot saved to {screenshot_path}")
                st.warning(f"Screenshot saved to {screenshot_path}")
            except:
                pass
            raise captcha_error

        # 5. Submit
        try:
            logger.info("Waiting for Submit button to be enabled...")
            st.info("Submitting data to portal...")

            # Wait for submit button to be visible first
            page.wait_for_selector("button#Schedule_Ok", state="visible", timeout=10000)
            logger.info("Submit button found")

            # Wait for submit button to be enabled (not disabled)
            logger.info("Waiting for submit button to be enabled...")
            try:
                page.wait_for_selector("button#Schedule_Ok:not([disabled])", timeout=10000)
                logger.info("✅ Submit button is now enabled")
            except Exception as wait_error:
                logger.warning(f"Submit button may still be disabled: {wait_error}")
                # Take screenshot for debugging
                try:
                    page.screenshot(path="submit_button_disabled.png")
                    logger.info("Screenshot saved to submit_button_disabled.png")
                except:
                    pass

            # Click submit and wait for new tab/popup/alert
            logger.info("Clicking submit button...")
            st.info("Clicking submit button...")

            # Variable to capture alert/dialog message
            alert_message = None
            alert_detected = False

            # Listen for JavaScript alerts/dialogs (for error popups)
            def handle_dialog(dialog):
                nonlocal alert_message, alert_detected
                alert_message = dialog.message
                alert_detected = True
                logger.info(f"🔔 Alert/Dialog detected: {dialog.type} - {dialog.message}")
                st.warning(f"🔔 Alert popup detected: {dialog.message}")
                dialog.accept()  # Click OK on the alert

            page.on("dialog", handle_dialog)

            # Click submit button
            page.click("button#Schedule_Ok")
            logger.info("✅ Submit button clicked")

            # Wait longer for alert to appear (alerts can take a few seconds)
            logger.info("Waiting for alert or page navigation...")
            time.sleep(5)

            # Check if alert was triggered (error case)
            if alert_detected and alert_message:
                logger.error(f"❌ REMC Portal Alert: {alert_message}")
                st.error(f"❌ REMC Portal Error: {alert_message}")

                # Check if it's a contract/template error
                if "CONTRACT" in alert_message.upper() or "template" in alert_message.lower():
                    st.error("⚠️ **Template Error Detected!**")
                    st.warning("The template file has an invalid contract ID for this revision.")
                    st.info("**Solution:** Download a new template from REMC portal with the correct contract ID.")
                    logger.error("Template contract ID mismatch - need to download new template from REMC portal")

                logger.error("REMC Portal submission failed due to validation error")
                st.error("❌ Submission failed - please fix the error and try again")
                return False

            logger.info(f"No alert detected after 5 seconds. Alert flag: {alert_detected}")

            # No alert detected, check for toast messages on current page
            try:
                toast_error = False
                toast_title = page.locator("div.toast-title")
                if toast_title.count() > 0:
                    toast_text = toast_title.inner_text()
                    logger.error(f"❌ Toast message detected: {toast_text}")
                    st.error(f"❌ REMC Portal Error: {toast_text}")

                    # Check if it's a contract/template error
                    if "CONTRACT" in toast_text.upper() or "template" in toast_text.lower():
                        st.error("⚠️ **Template Error Detected!**")
                        st.warning("The template file has an invalid contract ID for this revision.")
                        st.info("**Solution:** Download a new template from REMC portal with the correct contract ID.")
                        logger.error("Template contract ID mismatch - need to download new template from REMC portal")

                    toast_error = True

                # Also check for toast-message body
                toast_message = page.locator("div.toast-message")
                if toast_message.count() > 0:
                    message_text = toast_message.inner_text()
                    logger.error(f"Toast error details: {message_text}")
                    st.error(f"Details: {message_text}")
                    toast_error = True

                if toast_error:
                    logger.error("REMC Portal submission failed - toast error detected")
                    st.error("❌ Submission failed - please fix the error and try again")
                    return False

            except Exception as toast_check_error:
                logger.warning(f"Could not check for toast messages: {toast_check_error}")

            # No alert or toast, check for new page/tab (success case)
            try:
                # Get all pages in the context
                pages = context.pages
                logger.info(f"Number of pages after submit: {len(pages)}")

                # If a new page was opened, it should be the last one
                if len(pages) > 1:
                    new_page = pages[-1]
                    logger.info(f"New tab/page detected: {new_page.url}")
                    st.info(f"New tab opened: {new_page.url}")
                else:
                    # No new page, might be same page navigation
                    new_page = page
                    logger.info("No new tab detected, checking current page")

                # Wait for the page to load (use domcontentloaded for faster loading)
                new_page.wait_for_load_state("domcontentloaded", timeout=30000)
                logger.info("Page loaded")

                # Take screenshot of the result page
                if os.getenv('PLAYWRIGHT_SCREENSHOTS', 'false').lower() == 'true':
                    import pytz
                    ist = pytz.timezone('Asia/Kolkata')
                    screenshot_path = f'logs/screenshot_result_{datetime.datetime.now(ist).strftime("%Y%m%d_%H%M%S")}.png'
                    new_page.screenshot(path=screenshot_path)
                    logger.info(f"Result page screenshot saved: {screenshot_path}")

                # Get the content of the page
                result_text = new_page.inner_text("body")

                logger.info(f"Result page content (first 500 chars): {result_text[:500]}")
                st.info(f"📄 Page content: {result_text[:300]}")

                # Check for success or error messages in the new page
                error_detected = False

                # More specific success indicators (avoid matching menu text like "Submission")
                success_keywords = [
                    'submitted successfully',
                    'upload successful',
                    'successfully uploaded',
                    'schedule uploaded successfully',
                    'data saved successfully'
                ]

                # More specific error indicators
                error_keywords = [
                    'error',
                    'failed',
                    'invalid',
                    'not valid',
                    'not a valid',
                    'kindly download',
                    'check cell'
                ]

                # Check for specific success indicators (more strict)
                if any(keyword in result_text.lower() for keyword in success_keywords):
                    logger.info("✅ Success message detected in result page")
                    st.success("✅ Submission successful!")
                    st.info(f"Result: {result_text[:200]}")

                # Check for error indicators
                elif any(keyword in result_text.lower() for keyword in error_keywords):
                    logger.error(f"❌ Error detected in result page: {result_text[:500]}")
                    st.error("❌ Submission failed!")
                    st.error(f"Error message: {result_text[:500]}")
                    error_detected = True

                    # Check for specific contract error
                    if "CONTRACT" in result_text.upper() or "contract" in result_text.lower():
                        st.error("⚠️ **Template Error Detected!**")
                        st.warning("The template file has an invalid contract ID for this revision.")
                        st.info("**Solution:** Download a new template from REMC portal with the correct contract ID.")
                        logger.error("Template contract ID mismatch - need to download new template from REMC portal")

                else:
                    # No clear success or error message - just showing the upload page
                    logger.warning(f"⚠️ Unclear result - page content: {result_text[:200]}")
                    st.warning("⚠️ Could not determine if submission succeeded or failed")
                    st.info("The page is still showing the upload form. This usually means:")
                    st.info("1. An error occurred but wasn't shown clearly")
                    st.info("2. The submission didn't actually go through")
                    st.info("3. Please check the REMC portal manually to verify")
                    st.info(f"Page content: {result_text[:300]}")

                # Keep the new page open for a few seconds so user can see it
                time.sleep(3)

                # Close the new page (only if it's a different page, not the main page)
                if len(pages) > 1 and new_page != page:
                    try:
                        new_page.close()
                        logger.info("Result page closed")
                    except:
                        logger.info("Result page already closed or couldn't close")
                else:
                    logger.info("No separate result page to close")

                if not error_detected:
                    st.success("✅ Automation sequence finished. Please verify portal for confirmation.")
                    logger.info("REMC Portal submission sequence completed successfully.")
                    return True
                else:
                    logger.error("REMC Portal submission failed due to validation error")
                    st.error("❌ Submission failed - please fix the error and try again")
                    return False

            except Exception as new_page_error:
                logger.warning(f"No new page opened or error waiting for new page: {new_page_error}")
                st.warning("⚠️ No new result page detected")

                # Fall back to checking current page for toast messages
                time.sleep(3)

                error_detected = False
                try:
                    # Check for toast error messages on current page
                    toast_title = page.locator("div.toast-title")
                    if toast_title.count() > 0:
                        toast_text = toast_title.inner_text()
                        logger.error(f"❌ REMC Portal Error: {toast_text}")
                        st.error(f"❌ REMC Portal Error: {toast_text}")

                        # Check if it's a contract/template error
                        if "CONTRACT" in toast_text.upper() or "template" in toast_text.lower():
                            st.error("⚠️ **Template Error Detected!**")
                            st.warning("The template file has an invalid contract ID for this revision.")
                            st.info("**Solution:** Download a new template from REMC portal with the correct contract ID.")
                            logger.error("Template contract ID mismatch - need to download new template from REMC portal")

                        error_detected = True

                    # Also check for toast-message body
                    toast_message = page.locator("div.toast-message")
                    if toast_message.count() > 0:
                        message_text = toast_message.inner_text()
                        logger.error(f"Error details: {message_text}")
                        st.error(f"Details: {message_text}")

                except Exception as toast_check_error:
                    logger.warning(f"Could not check for toast messages: {toast_check_error}")

                # If no error detected, check for success messages on current page
                if not error_detected:
                    try:
                        page_content = page.content()
                        if "success" in page_content.lower() or "submitted" in page_content.lower():
                            logger.info("Success message detected on page")
                            st.success("✅ Submission appears successful!")
                        else:
                            logger.info("No clear success message found, but submission completed")
                            st.info("ℹ️ Submission completed - please verify on portal")
                    except:
                        pass

                    st.success("✅ Automation sequence finished. Please verify portal for confirmation.")
                    logger.info("REMC Portal submission sequence completed.")
                    return True
                else:
                    logger.error("REMC Portal submission failed due to validation error")
                    st.error("❌ Submission failed - please fix the error and try again")
                    return False

        except Exception as submit_error:
            logger.error(f"Error during submit: {submit_error}")
            # Take screenshot
            try:
                page.screenshot(path="submit_error.png")
                logger.info("Submit error screenshot saved to submit_error.png")
            except:
                pass
            raise submit_error
        
    except Exception as e:
        error_msg = f"❌ Error during REMC portal submission: {str(e)}"
        st.error(error_msg)
        logger.error(error_msg)
        import traceback
        logger.error(traceback.format_exc())

        # Try to take screenshot for debugging
        try:
            if page:
                screenshot_path = "remc_error.png"
                page.screenshot(path=screenshot_path)
                logger.info(f"Error screenshot saved to {screenshot_path}")
                st.info(f"Screenshot saved to {screenshot_path} for debugging")
        except:
            pass

        return False

    finally:
        # Clean up Playwright resources
        try:
            if page:
                page.close()
            if browser:
                browser.close()
            if playwright:
                playwright.stop()
        except:
            pass

def create_tables():
    conn = get_db_connection()
    cursor = conn.cursor()

    # Table for load data (note: "load_value" avoids reserved word conflict)
    create_load_table = """
    CREATE TABLE IF NOT EXISTS load_data (
        id INT AUTO_INCREMENT PRIMARY KEY,
        date DATE,
        block INT,
        buyer_name VARCHAR(100),
        load_value FLOAT
    );
    """
    cursor.execute(create_load_table)

    # Table for obligation data
    create_obligation_table = """
    CREATE TABLE IF NOT EXISTS obligation (
        id INT AUTO_INCREMENT PRIMARY KEY,
        date DATE,
        block INT,
        market FLOAT DEFAULT 0,
        market_fp FLOAT DEFAULT 0,
        market_linde FLOAT DEFAULT 0,
        UNIQUE KEY date_block_idx (date, block)
    );
    """
    cursor.execute(create_obligation_table)

    for alter_sql in [
        "ALTER TABLE obligation ADD COLUMN market_fp FLOAT DEFAULT 0",
        "ALTER TABLE obligation ADD COLUMN market_linde FLOAT DEFAULT 0",
        "ALTER TABLE obligation MODIFY COLUMN market FLOAT DEFAULT 0",
    ]:
        try:
            cursor.execute(alter_sql)
        except Exception:
            pass

    # Create two schedule tables: one for Solar and one for Wind.
    create_solar_schedule = """
    CREATE TABLE IF NOT EXISTS schedule_solar (
        id INT AUTO_INCREMENT PRIMARY KEY,
        date DATE,
        block INT,
        sch FLOAT,
        avc FLOAT DEFAULT 0,
        UNIQUE KEY date_block_idx (date, block)
    );
    """
    cursor.execute(create_solar_schedule)

    create_wind_schedule = """
    CREATE TABLE IF NOT EXISTS schedule_wind (
        id INT AUTO_INCREMENT PRIMARY KEY,
        date DATE,
        block INT,
        sch FLOAT,
        avc FLOAT DEFAULT 0,
        UNIQUE KEY date_block_idx (date, block)
    );
    """
    cursor.execute(create_wind_schedule)

    for table_name in ["schedule_solar", "schedule_wind"]:
        try:
            cursor.execute(f"ALTER TABLE {table_name} ADD COLUMN avc FLOAT DEFAULT 0")
        except Exception:
            pass

    create_hybrid_avc_cap = """
    CREATE TABLE IF NOT EXISTS hybrid_avc_cap (
        date DATE PRIMARY KEY,
        avc_cap FLOAT NOT NULL DEFAULT 200
    );
    """
    cursor.execute(create_hybrid_avc_cap)

    conn.commit()
    cursor.close()
    conn.close()

# Create GDAM-RTM Ratio table
def create_gdam_rtm_ratio_table():
    conn = get_db_connection()
    cursor = conn.cursor()

    create_gdam_rtm_ratio_table = """
    CREATE TABLE IF NOT EXISTS gdam_rtm_ratio (
        id INT AUTO_INCREMENT PRIMARY KEY,
        block INT,
        gdam FLOAT,
        rtm FLOAT,
        UNIQUE KEY block_idx (block)
    );
    """
    cursor.execute(create_gdam_rtm_ratio_table)

    # Check if table is empty and insert default values
    cursor.execute("SELECT COUNT(*) FROM gdam_rtm_ratio")
    count = cursor.fetchone()[0]

    if count == 0:
        # Insert default values where GDAM = 0.5 and RTM = 0.5 for all 96 blocks
        for block in range(1, 97):
            cursor.execute(
                "INSERT INTO gdam_rtm_ratio (block, gdam, rtm) VALUES (%s, %s, %s)",
                (block, 0.5, 0.5)
            )

    conn.commit()
    cursor.close()
    conn.close()

# Create configuration tables: contract_value, state, tariff_difference, buyer_mapping.
def create_config_tables():
    conn = get_db_connection()
    cursor = conn.cursor()

    create_contract_table = """
    CREATE TABLE IF NOT EXISTS contract_value (
        buyer_name VARCHAR(100) PRIMARY KEY,
        contract_value FLOAT
    );
    """
    cursor.execute(create_contract_table)

    create_state_table = """
    CREATE TABLE IF NOT EXISTS state (
        buyer_name VARCHAR(100) PRIMARY KEY,
        state VARCHAR(100)
    );
    """
    cursor.execute(create_state_table)

    create_tariff_table = """
    CREATE TABLE IF NOT EXISTS tariff_difference (
        block INT,
        state VARCHAR(100),
        tariff_difference FLOAT,
        PRIMARY KEY (block, state)
    );
    """
    cursor.execute(create_tariff_table)

    # Create buyer mapping table for Final Buyer Schedule
    # Final Revision CSV = REMC template submitted for that day
    # file_column_name = Buyer name from Row 16 of REMC template (e.g., LINDEINDIA)
    # capacity_mw = Capacity from Row 19 of REMC template (e.g., 10)
    # Multiple buyers can have same name (e.g., LINDEINDIA) but different capacities
    # So we use file_column_name + capacity_mw as unique identifier
    create_buyer_mapping_table = """
    CREATE TABLE IF NOT EXISTS buyer_mapping (
        id INT AUTO_INCREMENT PRIMARY KEY,
        file_column_name VARCHAR(100),
        actual_buyer_name VARCHAR(100),
        capacity_mw DECIMAL(10,2),
        description VARCHAR(255),
        UNIQUE KEY unique_mapping (file_column_name, capacity_mw)
    );
    """
    cursor.execute(create_buyer_mapping_table)

    create_buyer_access_type_table = """
    CREATE TABLE IF NOT EXISTS buyer_access_type (
        buyer_name VARCHAR(100) PRIMARY KEY,
        access_type VARCHAR(10) NOT NULL DEFAULT 'GNA'
    );
    """
    cursor.execute(create_buyer_access_type_table)

    create_gdam_allocation_config_table = """
    CREATE TABLE IF NOT EXISTS gdam_allocation_config (
        id INT PRIMARY KEY,
        allocation_pct FLOAT NOT NULL DEFAULT 0
    );
    """
    cursor.execute(create_gdam_allocation_config_table)
    cursor.execute(
        """
        INSERT INTO gdam_allocation_config (id, allocation_pct)
        VALUES (1, 0)
        ON DUPLICATE KEY UPDATE allocation_pct = allocation_pct
        """
    )

    # Add capacity_mw column if it doesn't exist (for existing databases)
    try:
        cursor.execute("ALTER TABLE buyer_mapping ADD COLUMN capacity_mw DECIMAL(10,2)")
    except:
        pass  # Column already exists

    # Drop old unique constraint and add new one
    try:
        cursor.execute("ALTER TABLE buyer_mapping DROP INDEX file_column_name")
    except:
        pass  # Constraint doesn't exist

    try:
        cursor.execute("ALTER TABLE buyer_mapping ADD UNIQUE KEY unique_mapping (file_column_name, capacity_mw)")
    except:
        pass  # Constraint already exists

    # Remove remc_template_name column if it exists (it's duplicate of file_column_name)
    try:
        cursor.execute("ALTER TABLE buyer_mapping DROP COLUMN remc_template_name")
    except:
        pass  # Column doesn't exist

    # NOTE: No default mappings inserted here - all mappings should be managed through the UI
    # Users should add buyer mappings via "Edit Final Revision Integration" page
    # This ensures all data comes from the database and nothing is hardcoded
    # Buyer mappings are created dynamically when processing REMC templates or Final Buyer Schedules

    conn.commit()
    cursor.close()
    conn.close()

# Create CUF% table to store daily cumulative CUF% for each buyer.
def create_cuf_pct_table():
    conn = get_db_connection()
    cursor = conn.cursor()
    query = """
    CREATE TABLE IF NOT EXISTS cuf_pct (
        date DATE,
        buyer_name VARCHAR(100),
        capacity_mw FLOAT,
        cuf_pct FLOAT,
        PRIMARY KEY (date, buyer_name, capacity_mw)
    );
    """
    cursor.execute(query)
    conn.commit()
    cursor.close()
    conn.close()


# Create manual_priority table to store persistent manual buyer priorities.
def create_manual_priority_table():
    conn = get_db_connection()
    cursor = conn.cursor()
    query = """
    CREATE TABLE IF NOT EXISTS manual_priority (
        buyer_name VARCHAR(100) PRIMARY KEY,
        priority VARCHAR(10) NOT NULL DEFAULT 'High'
    );
    """
    cursor.execute(query)
    conn.commit()
    cursor.close()
    conn.close()


def ensure_hybrid_avc_cap_defaults(target_dates, default_cap=DEFAULT_HYBRID_AVC_CAP):
    """Ensure the requested dates exist in the hybrid AVC cap table."""
    create_tables()

    conn = get_db_connection()
    cursor = conn.cursor()

    for target_date in target_dates:
        cursor.execute(
            "INSERT IGNORE INTO hybrid_avc_cap (date, avc_cap) VALUES (%s, %s)",
            (target_date, float(default_cap))
        )

    conn.commit()
    cursor.close()
    conn.close()


def load_manual_priorities():
    """Load manual priorities from the manual_priority table. Returns dict {buyer_name: 'High'/'Low'}."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("SELECT buyer_name, priority FROM manual_priority")
        rows = cursor.fetchall()
        cursor.close()
        conn.close()
        return {row[0]: row[1] for row in rows}
    except Exception:
        return {}


def save_manual_priorities(priority_map):
    """Save manual priorities to the manual_priority table. priority_map: {buyer_name: 'High'/'Low'}."""
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        for buyer_name, priority in priority_map.items():
            cursor.execute(
                """
                INSERT INTO manual_priority (buyer_name, priority)
                VALUES (%s, %s)
                ON DUPLICATE KEY UPDATE priority = VALUES(priority)
                """,
                (buyer_name, priority)
            )
        conn.commit()
        cursor.close()
        conn.close()
        return True
    except Exception as e:
        st.error(f"❌ Failed to save manual priorities: {str(e)}")
        return False


def normalize_buyer_access_type(access_type):
    if isinstance(access_type, str) and access_type.strip().upper() == "TGNA":
        return "TGNA"
    return "GNA"


def load_buyer_access_types():
    """Load stored buyer access types. Returns {buyer_name: 'GNA'|'TGNA'}."""
    create_config_tables()

    conn = None
    try:
        conn = get_db_connection()
        access_df = execute_query_to_dataframe(
            "SELECT buyer_name, access_type FROM buyer_access_type",
            conn
        )
    except Exception:
        return {}
    finally:
        if conn is not None:
            conn.close()

    if access_df.empty:
        return {}

    access_map = {}
    for _, row in access_df.iterrows():
        buyer_name = str(row.get("buyer_name", "")).strip()
        if not buyer_name:
            continue
        access_map[buyer_name] = normalize_buyer_access_type(row.get("access_type"))
    return access_map


def split_energy_by_contract(total_energy, primary_weight, secondary_weight):
    """Split a block's available energy between two groups using contract-value weights."""
    total_energy = round(max(to_float(total_energy), 0.0), 1)
    primary_weight = max(to_float(primary_weight), 0.0)
    secondary_weight = max(to_float(secondary_weight), 0.0)
    total_weight = primary_weight + secondary_weight

    if total_energy <= 0 or total_weight <= 0:
        return 0.0, 0.0

    if primary_weight <= 0:
        return 0.0, total_energy

    if secondary_weight <= 0:
        return total_energy, 0.0

    primary_share = round(total_energy * (primary_weight / total_weight), 1)
    secondary_share = round(total_energy - primary_share, 1)

    if secondary_share < 0:
        primary_share = total_energy
        secondary_share = 0.0

    return primary_share, secondary_share


def get_active_contract_totals(contract_values, buyers, demand_by_buyer):
    """Return buyers with positive block demand and their summed contract value."""
    active_buyers = [
        buyer for buyer in buyers
        if round(max(to_float(demand_by_buyer.get(buyer, 0.0)), 0.0), 1) > 0
    ]
    total_contract = round(
        sum(to_float(contract_values.get(buyer, 0.0)) for buyer in active_buyers),
        1,
    )
    return active_buyers, total_contract


def split_energy_with_post_linde_gdam(total_schedule, gdam_fp, gdam_linde, linde_weight, non_linde_weight):
    """
    Apply the manual GDAM flow:
    1. Deduct Fourth Partner GDAM from schedule before the first group split.
    2. Split the remaining energy between Linde and non-Linde groups.
    3. Deduct Linde GDAM from the Linde group allocation.
    4. Clamp the post-GDAM Linde bucket at zero without changing the manually
       entered GDAM values or boosting schedule.
    """
    total_schedule = round(max(to_float(total_schedule), 0.0), 1)
    gdam_fp = round(max(to_float(gdam_fp), 0.0), 1)
    gdam_linde = round(max(to_float(gdam_linde), 0.0), 1)

    adjusted_schedule = total_schedule
    available_after_fp = round(max(adjusted_schedule - gdam_fp, 0.0), 1)

    linde_pre_gdam, non_linde_available = split_energy_by_contract(
        available_after_fp,
        linde_weight,
        non_linde_weight,
    )

    linde_available = round(max(linde_pre_gdam - gdam_linde, 0.0), 1)

    return {
        "adjusted_schedule": adjusted_schedule,
        "available_after_fp": available_after_fp,
        "linde_pre_gdam": round(linde_pre_gdam, 1),
        "linde_available": linde_available,
        "non_linde_available": round(non_linde_available, 1),
        "schedule_boost": 0.0,
        "post_gdam_available_total": round(linde_available + non_linde_available, 1),
    }


def assign_source_eligible_allocations(
    allocated_totals,
    buyer_list,
    buyer_access_types,
    solar_budget,
    wind_budget,
    high_priority_buyers,
    low_priority_buyers,
    tariff_diff_by_buyer,
):
    """
    Convert total buyer allocations into solar/wind allocations while enforcing:
    - GNA buyers consume solar first, then wind
    - TGNA buyers consume wind first, then solar
    - TGNA solar receipts are written into the wind contract column

    Returns:
        tuple[dict, dict, dict, dict]:
        (contract_allocation_row, served_totals, shortfall_by_buyer, source_usage_row)
    """
    solar_budget = round(max(to_float(solar_budget), 0.0), 1)
    wind_budget = round(max(to_float(wind_budget), 0.0), 1)

    planned_totals = {
        buyer: round(max(to_float(allocated_totals.get(buyer, 0.0)), 0.0), 1)
        for buyer in buyer_list
    }
    served_totals = {buyer: 0.0 for buyer in buyer_list}
    shortfall_by_buyer = {buyer: 0.0 for buyer in buyer_list}
    contract_allocation_row = {}
    source_usage_row = {}

    def allocate_budget_by_priority(available_budget, active_buyers, targets):
        available_budget = round(max(to_float(available_budget), 0.0), 1)
        allocations = {buyer: 0.0 for buyer in active_buyers}

        scoped_buyers = [buyer for buyer in active_buyers if round(max(to_float(targets.get(buyer, 0.0)), 0.0), 1) > 0]
        if available_budget <= 0 or not scoped_buyers:
            return allocations, available_budget

        scoped_high_priority = [buyer for buyer in high_priority_buyers if buyer in scoped_buyers]
        scoped_low_priority = [buyer for buyer in low_priority_buyers if buyer in scoped_buyers]

        if available_budget > 0 and scoped_high_priority:
            hp_targets = {buyer: round(max(to_float(targets.get(buyer, 0.0)), 0.0), 1) for buyer in scoped_high_priority}
            hp_alloc, available_budget = allocate_high_priority(
                available_budget,
                scoped_high_priority,
                hp_targets,
            )
            for buyer, amount in hp_alloc.items():
                allocations[buyer] = round(allocations.get(buyer, 0.0) + amount, 1)

        if available_budget > 0 and scoped_low_priority:
            lp_targets = {buyer: round(max(to_float(targets.get(buyer, 0.0)), 0.0), 1) for buyer in scoped_low_priority}
            lp_tariffs = {buyer: tariff_diff_by_buyer.get(buyer, 0.0) for buyer in scoped_low_priority}
            lp_alloc, available_budget = allocate_low_priority(
                available_budget,
                scoped_low_priority,
                lp_targets,
                lp_tariffs,
            )
            for buyer, amount in lp_alloc.items():
                allocations[buyer] = round(allocations.get(buyer, 0.0) + amount, 1)

        return allocations, available_budget

    tgna_buyers = [
        buyer for buyer in buyer_list
        if normalize_buyer_access_type(buyer_access_types.get(buyer)) == "TGNA"
        and planned_totals.get(buyer, 0.0) > 0
    ]
    gna_buyers = [
        buyer for buyer in buyer_list
        if normalize_buyer_access_type(buyer_access_types.get(buyer)) != "TGNA"
        and planned_totals.get(buyer, 0.0) > 0
    ]

    tgna_wind_source, remaining_wind_budget = allocate_budget_by_priority(
        wind_budget,
        tgna_buyers,
        planned_totals,
    )
    gna_solar_source, remaining_solar_budget = allocate_budget_by_priority(
        solar_budget,
        gna_buyers,
        planned_totals,
    )

    tgna_remaining_targets = {
        buyer: round(max(planned_totals[buyer] - tgna_wind_source.get(buyer, 0.0), 0.0), 1)
        for buyer in tgna_buyers
    }
    gna_remaining_targets = {
        buyer: round(max(planned_totals[buyer] - gna_solar_source.get(buyer, 0.0), 0.0), 1)
        for buyer in gna_buyers
    }

    tgna_solar_source, remaining_solar_budget = allocate_budget_by_priority(
        remaining_solar_budget,
        tgna_buyers,
        tgna_remaining_targets,
    )
    gna_wind_source, remaining_wind_budget = allocate_budget_by_priority(
        remaining_wind_budget,
        gna_buyers,
        gna_remaining_targets,
    )

    for buyer in tgna_buyers:
        buyer_wind_source = round(tgna_wind_source.get(buyer, 0.0), 1)
        buyer_solar_source = round(tgna_solar_source.get(buyer, 0.0), 1)
        served_amount = round(buyer_wind_source + buyer_solar_source, 1)
        served_totals[buyer] = served_amount
        shortfall_by_buyer[buyer] = round(max(planned_totals[buyer] - served_amount, 0.0), 1)
        contract_allocation_row[f"{buyer}_Solar"] = 0.0
        contract_allocation_row[f"{buyer}_Wind"] = served_amount
        source_usage_row[f"{buyer}_Solar"] = buyer_solar_source
        source_usage_row[f"{buyer}_Wind"] = buyer_wind_source

    for buyer in gna_buyers:
        buyer_solar_source = round(gna_solar_source.get(buyer, 0.0), 1)
        buyer_wind_source = round(gna_wind_source.get(buyer, 0.0), 1)
        served_amount = round(buyer_solar_source + buyer_wind_source, 1)
        served_totals[buyer] = served_amount
        shortfall_by_buyer[buyer] = round(max(planned_totals[buyer] - served_amount, 0.0), 1)
        contract_allocation_row[f"{buyer}_Solar"] = buyer_solar_source
        contract_allocation_row[f"{buyer}_Wind"] = buyer_wind_source
        source_usage_row[f"{buyer}_Solar"] = buyer_solar_source
        source_usage_row[f"{buyer}_Wind"] = buyer_wind_source

    for buyer in buyer_list:
        contract_allocation_row.setdefault(f"{buyer}_Solar", 0.0)
        contract_allocation_row.setdefault(f"{buyer}_Wind", 0.0)
        source_usage_row.setdefault(f"{buyer}_Solar", 0.0)
        source_usage_row.setdefault(f"{buyer}_Wind", 0.0)

    return contract_allocation_row, served_totals, shortfall_by_buyer, source_usage_row

# -------------------------------
# Data Loading Function (Using FTP)
# -------------------------------

def find_final_buyer_schedule_file(ftp, remote_folder, target_date):
    """
    Find the final buyer schedule file for a specific date in the format dd-mm-yyyy.csv
    Checks both the main and fallback FTP paths. Returns any one file for the date if multiple exist.
    """
    date_str = target_date.strftime("%d-%m-%Y")  # Format as dd-mm-yyyy
    for folder in [FTP_PATH_FINAL_BUYER_SCHEDULE, FTP_PATH_FINAL_BUYER_SCHEDULE_FALLBACK]:
        try:
            ftp.cwd(folder)
            files = ftp.nlst()
            for f in files:
                if date_str in f and f.lower().endswith('.csv'):
                    return f, folder
        except Exception:
            continue
    return None, None

def process_final_buyer_schedule(target_date=None, start_row=17, end_row=115):
    """
    Process the final buyer schedule file for a specific date and calculate CUF%

    Parameters:
    - target_date: The date to process (default: today)
    - start_row: The starting row number in the CSV file (1-based, default: 17)
    - end_row: The ending row number in the CSV file (1-based, default: 115)
    """
    if target_date is None:
        target_date = today_ist()

    st.write(f"Processing final buyer schedule for {target_date}")

    # Convert to 0-based indexing for pandas
    # If header is at row 11 (0-based index 10), then:
    # Row 18 corresponds to index 7 (18-11)
    # Row 113 corresponds to index 102 (113-11)
    start_idx = start_row - 11  # Adjust for header at row 11
    end_idx = end_row - 11 + 1  # +1 because end index is exclusive in pandas

    # Connect to FTP and find the file
    ftp = connect_ftp()
    file_name, folder = find_final_buyer_schedule_file(ftp, None, target_date)
    if not file_name:
        st.warning(f"No final buyer schedule file found for {target_date}")
        ftp.quit()
        return False

    # Download the file
    bio = download_ftp_file(ftp, folder, file_name)
    ftp.quit()

    if not bio:
        st.error(f"Failed to download {file_name}")
        return False

    # Read the CSV file
    try:
        st.write(f"Processing file: {file_name}")

        # Read all lines to get buyer names (Row 11) and capacities (Row 16)
        bio.seek(0)
        all_lines = bio.read().decode('utf-8').splitlines()

        # Row 11 (index 10) = Buyer names
        # Row 16 (index 15) = Maximum Contract Capacity
        buyer_names_row = all_lines[10].split(',') if len(all_lines) > 10 else []
        capacity_row = all_lines[15].split(',') if len(all_lines) > 15 else []

        st.write(f"\n**Row 11 (Buyer Names):** {len(buyer_names_row)} columns")
        st.write(f"**Row 16 (Capacities):** {len(capacity_row)} columns")

        # Reset and read with pandas
        bio.seek(0)

        # Read the file with header at row 10 (11th row in the file)
        df = pd.read_csv(bio, header=10)

        st.write(f"\n**Pandas found {len(df.columns)} columns**")

        # Get the buyer mapping from the database
        conn = get_db_connection()
        buyer_mapping_df = execute_query_to_dataframe("SELECT file_column_name, actual_buyer_name, capacity_mw FROM buyer_mapping", conn)

        # Create a mapping that uses BOTH file_column_name AND capacity_mw
        # Key: (file_column_name, capacity_mw) -> Value: actual_buyer_name
        buyer_mapping_with_capacity = {}
        for _, row in buyer_mapping_df.iterrows():
            key = (row['file_column_name'], float(row['capacity_mw']))
            buyer_mapping_with_capacity[key] = row['actual_buyer_name']

        # Get contract values (capacity_mw) for CUF calculation from buyer_mapping table
        # Convert Decimal to float to avoid type errors in calculations
        contract_values = {buyer: float(capacity) for buyer, capacity in zip(buyer_mapping_df['actual_buyer_name'], buyer_mapping_df['capacity_mw'])}

        # Calculate total energy for each buyer
        buyer_energy = {}

        # Process each column in the CSV file
        # Match using BOTH column name (from Row 11) AND capacity (from Row 16)
        st.write("\n**Column Matching:**")
        for col_idx, file_col in enumerate(df.columns):
            # Get the buyer name from Row 11 and capacity from Row 16
            # Note: df.columns are from Row 11, so we need to match by index
            if col_idx < len(buyer_names_row) and col_idx < len(capacity_row):
                buyer_name_from_file = buyer_names_row[col_idx].strip()
                capacity_from_file_str = capacity_row[col_idx].strip()

                # Try to parse capacity as float
                try:
                    capacity_from_file = float(capacity_from_file_str) if capacity_from_file_str else 0
                except ValueError:
                    capacity_from_file = 0

                # Skip system columns
                if capacity_from_file == 0 or buyer_name_from_file in ['Buyer Name', 'Block No.', 'TimeStamp', 'AVC', '']:
                    continue

                # Match using (file_column_name, capacity_mw)
                key = (buyer_name_from_file, capacity_from_file)

                if key in buyer_mapping_with_capacity:
                    actual_buyer_name = buyer_mapping_with_capacity[key]

                    # Sum the energy values using the specified row range
                    energy_values = df.iloc[start_idx:end_idx][file_col].astype(float)
                    total_energy = energy_values.sum()
                    buyer_energy[actual_buyer_name] = total_energy

                    st.write(f"✅ Matched '{buyer_name_from_file}' ({capacity_from_file} MW) → {actual_buyer_name}: {total_energy:.2f} MWh")
                else:
                    st.warning(f"⚠️ No mapping found for: {buyer_name_from_file} ({capacity_from_file} MW)")

        # Calculate CUF% for each buyer
        cuf_pct = {}
        for buyer, energy in buyer_energy.items():
            if buyer in contract_values and contract_values[buyer] > 0:
                # Calculate CUF%
                contract_val = contract_values[buyer]
                days_in_month = calendar.monthrange(target_date.year, target_date.month)[1]
                denominator = contract_val * 96 * days_in_month
                daily_cuf_pct = (energy / denominator) * 100 if denominator > 0 else 0
                cuf_pct[buyer] = daily_cuf_pct
            else:
                st.warning(f"No contract value found for {buyer}")

        # Get previous day's cumulative CUF% (using buyer_name AND capacity_mw)
        prev_day = target_date - timedelta(days=1)
        prev_cuf_query = f"SELECT buyer_name, capacity_mw, cuf_pct FROM cuf_pct WHERE date = '{prev_day}'"
        prev_cuf_df = execute_query_to_dataframe(prev_cuf_query, conn)

        # If no previous day data, try to get the most recent data
        if prev_cuf_df.empty:
            prev_cuf_query = f"""
                SELECT buyer_name, capacity_mw, cuf_pct
                FROM cuf_pct
                WHERE date = (
                    SELECT MAX(date)
                    FROM cuf_pct
                    WHERE date < '{target_date}'
                )
            """
            prev_cuf_df = execute_query_to_dataframe(prev_cuf_query, conn)

        # Create dict with (buyer_name, capacity) as key
        prev_cuf_pct = {}
        if not prev_cuf_df.empty:
            for _, row in prev_cuf_df.iterrows():
                key = (row['buyer_name'], float(row['capacity_mw']))
                prev_cuf_pct[key] = float(row['cuf_pct'])

        # Calculate cumulative CUF%
        cumulative_cuf_pct = {}
        for buyer, daily_pct in cuf_pct.items():
            capacity = contract_values.get(buyer, 0)
            key = (buyer, capacity)

            # For the first day of the month, start fresh
            if target_date.day == 1:
                cumulative_cuf_pct[key] = daily_pct
            else:
                # Add previous day's CUF% to accumulate over days
                prev_cuf = prev_cuf_pct.get(key, 0)  # Get last day's CUF or 0
                cumulative_cuf_pct[key] = prev_cuf + daily_pct

        # Update the CUF% table
        cursor = conn.cursor()
        for (buyer, capacity), pct in cumulative_cuf_pct.items():
            query = "REPLACE INTO cuf_pct (date, buyer_name, capacity_mw, cuf_pct) VALUES (%s, %s, %s, %s)"
            cursor.execute(query, (target_date, buyer, float(capacity), float(pct)))

        conn.commit()
        cursor.close()
        conn.close()

        st.success(f"Successfully processed final buyer schedule for {target_date}")
        return True

    except Exception as e:
        st.error(f"Error processing final buyer schedule: {str(e)}")
        return False

def load_data_to_mysql():
    create_tables()

    # Define target dates: current date and current date + 1.
    today = today_ist()
    tomorrow = today + timedelta(days=1)
    yesterday = today - timedelta(days=1)
    ensure_hybrid_avc_cap_defaults([today, tomorrow])

    # Delete all data except for yesterday, today, and tomorrow
    conn = get_db_connection()
    cursor = conn.cursor()
    for table in ["load_data", "obligation", "schedule_wind", "schedule_solar", "hybrid_avc_cap"]:
        cursor.execute(f"DELETE FROM {table} WHERE date NOT IN (%s, %s, %s)", (yesterday, today, tomorrow))
    conn.commit()
    cursor.close()
    conn.close()

    conn = get_db_connection()
    cursor = conn.cursor()
    ftp = connect_ftp()
    for target_date in [today, tomorrow]:
        # 1. Process Load folder.
        load_filename = find_file_for_date(ftp, FTP_PATH_LOAD, target_date, file_type='load')
        if load_filename:
            bio = download_ftp_file(ftp, FTP_PATH_LOAD, load_filename)
            if bio:
                if load_filename.lower().endswith(".xlsx"):
                    df = pd.read_excel(bio)
                else:
                    df = pd.read_csv(bio)

                # Clear existing entries for this date before inserting new ones.
                delete_query = "DELETE FROM load_data WHERE date = %s"
                cursor.execute(delete_query, (target_date,))
                conn.commit()

                # Process each row with upsert logic.
                # Note: We now force the date value to the target_date rather than reading from file.
                for _, row in df.iterrows():
                    # Use target_date as the date value.
                    date_val = target_date

                    block = row.get('Time Block', None)
                    if pd.isna(block):
                        block = 0  # Or another appropriate default.
                    try:
                        block = int(block)
                    except Exception:
                        block = 0

                    # Process each buyer column (ignoring Date and Time Block)
                    for col in df.columns:
                        if col not in ['Date', 'Time Block']:
                            # Force buyer_name to a string. If the column name is empty or "nan", replace with "0".
                            buyer_name = str(col).strip()
                            if buyer_name == "" or buyer_name.lower() == "nan":
                                buyer_name = "0"

                            load_value = row[col]
                            if pd.isna(load_value):
                                load_value = 0

                            query = """INSERT INTO load_data (date, block, buyer_name, load_value)
                                       VALUES (%s, %s, %s, %s)
                                       ON DUPLICATE KEY UPDATE load_value = VALUES(load_value)"""
                            cursor.execute(query, (date_val, block, buyer_name, load_value))
                conn.commit()
                st.success(f"Load data from file '{load_filename}' inserted/updated successfully.")

        # 2. Preserve manual obligation/GDAM input.
        # G-DAM values are maintained only through the obligation table UI and
        # should never be overwritten from FTP during Load Data.
        obligation_filename = find_file_for_date(ftp, FTP_PATH_OBLIGATION, target_date, file_type='obligation')
        check_query = "SELECT COUNT(*) FROM obligation WHERE date = %s"
        cursor.execute(check_query, (target_date,))
        count = cursor.fetchone()[0]

        if obligation_filename:
            st.info(
                f"Found obligation file '{obligation_filename}', but skipped importing it because "
                "G-DAM/obligation values are input-only and preserved from the database."
            )

        if count > 0:
            st.success(f"Existing manual obligation data for date '{target_date}' preserved.")
        else:
            st.warning(
                f"No manual obligation data found for date '{target_date}'. Inserting default zero "
                "rows so G-DAM can be filled manually in the app."
            )
            for block in range(1, 97):
                query = """
                    REPLACE INTO obligation (date, block, market, market_fp, market_linde)
                    VALUES (%s, %s, %s, %s, %s)
                """
                cursor.execute(query, (target_date, block, 0.0, 0.0, 0.0))
            conn.commit()
            st.success(f"Default obligation data for date '{target_date}' inserted successfully.")

        # 3. Process Hybrid Schedule Revision File.
        # Determine schedule type: 'DA' (Day Ahead) if target_date is after today; else 'IntraDay'
        schedule_type = "DA" if target_date > today else "IntraDay"
        st.write(f"Processing {target_date} as {schedule_type} schedule")
        schedule_filename, schedule_revision = find_latest_revision_file(ftp, FTP_PATH_SCHEDULE, target_date, schedule_type)

        # Store revision number in session state for later use in REMC templates
        if schedule_revision:
            if schedule_type == "DA":
                st.session_state.da_revision = schedule_revision
                st.info(f"📌 Stored DA revision: R-{schedule_revision} from file: {schedule_filename}")
            else:
                st.session_state.id_revision = schedule_revision
                st.info(f"📌 Stored ID revision: R-{schedule_revision} from file: {schedule_filename}")

        if schedule_filename:
            bio = download_ftp_file(ftp, FTP_PATH_SCHEDULE, schedule_filename)
            if bio:
                try:
                    parsed_components = parse_schedule_revision_components(bio, schedule_filename, target_date)
                    st.info(
                        f"✅ Parsed {parsed_components['format']} schedule file for {schedule_type}: "
                        f"{schedule_filename}"
                    )

                    solar_df = parsed_components.get("solar_df")
                    wind_df = parsed_components.get("wind_df")

                    if solar_df is not None:
                        cursor.execute("DELETE FROM schedule_solar WHERE date = %s", (target_date,))
                        conn.commit()

                        solar_tuples = [
                            (
                                row["Date"],
                                int(to_float(row["Block"])),
                                to_float(row["Sch"]),
                                to_float(row["AvC"]),
                            )
                            for _, row in solar_df.iterrows()
                        ]
                        cursor.executemany(
                            "REPLACE INTO schedule_solar (date, block, sch, avc) VALUES (%s, %s, %s, %s)",
                            solar_tuples,
                        )
                        conn.commit()
                        st.success(f"Solar schedule ({schedule_type}) data from '{schedule_filename}' updated successfully.")
                        if parsed_components["format"] == "hybrid_r2":
                            st.caption(
                                f"Solar AvC values from the new RDA/RID format are capped at "
                                f"{SCHEDULE_SOLAR_AVC_OVERRIDE_MW} MW during import."
                            )
                    else:
                        check_query = "SELECT COUNT(*) FROM schedule_solar WHERE date = %s"
                        cursor.execute(check_query, (target_date,))
                        count = cursor.fetchone()[0]

                        if count > 0:
                            st.success(f"Existing solar schedule for date '{target_date}' preserved.")
                        else:
                            default_solar_data = [(target_date, block, 0.0, 0.0) for block in range(1, 97)]
                            cursor.executemany(
                                "REPLACE INTO schedule_solar (date, block, sch, avc) VALUES (%s, %s, %s, %s)",
                                default_solar_data,
                            )
                            conn.commit()
                            st.success(f"Default solar schedule for date '{target_date}' inserted successfully.")

                    cursor.execute("DELETE FROM schedule_wind WHERE date = %s", (target_date,))
                    conn.commit()

                    wind_tuples = [
                        (
                            row["Date"],
                            int(to_float(row["Block"])),
                            to_float(row["Sch"]),
                            to_float(row["AvC"]),
                        )
                        for _, row in wind_df.iterrows()
                    ]
                    cursor.executemany(
                        "REPLACE INTO schedule_wind (date, block, sch, avc) VALUES (%s, %s, %s, %s)",
                        wind_tuples,
                    )
                    conn.commit()
                    st.success(f"Wind schedule ({schedule_type}) data from '{schedule_filename}' updated successfully.")

                except Exception as e:
                    st.error(f"Error processing {schedule_filename}: {str(e)}")
                    conn.rollback()
        else:
            for table_name, schedule_label in [("schedule_solar", "solar"), ("schedule_wind", "wind")]:
                check_query = f"SELECT COUNT(*) FROM {table_name} WHERE date = %s"
                cursor.execute(check_query, (target_date,))
                count = cursor.fetchone()[0]

                if count > 0:
                    st.warning(f"{schedule_type} {schedule_label} schedule file for date '{target_date}' not found. Using existing data in database.")
                    st.success(f"Existing {schedule_type} {schedule_label} schedule for date '{target_date}' preserved.")
                else:
                    st.warning(f"{schedule_type} {schedule_label} schedule file for date '{target_date}' not found. No existing data found. Inserting default schedule.")
                    default_data = [(target_date, block, 0.0, 0.0) for block in range(1, 97)]
                    cursor.executemany(
                        f"REPLACE INTO {table_name} (date, block, sch, avc) VALUES (%s, %s, %s, %s)",
                        default_data
                    )
                    conn.commit()
                    st.success(f"Default {schedule_type} {schedule_label} schedule for date '{target_date}' inserted successfully.")

    cursor.close()
    conn.close()

    # 5. Load GDAM-RTM Ratio data from FTP
    st.write("### Loading GDAM-RTM Ratio Data")
    try:
        # Find the latest modified file in the GDAM_RTM_Ratio folder
        latest_file = find_latest_modified_file(ftp, FTP_PATH_GDAM_RTM_RATIO, ['.xlsx', '.csv'])

        if latest_file:
            st.info(f"Found GDAM-RTM ratio file: {latest_file}")

            # Download the file
            bio = download_ftp_file(ftp, FTP_PATH_GDAM_RTM_RATIO, latest_file)

            if bio:
                # Read the file based on extension
                if latest_file.lower().endswith('.xlsx'):
                    df = pd.read_excel(bio)
                else:
                    df = pd.read_csv(bio)

                # Validate and process the file
                expected_columns = ['Time Block', 'GDAM', 'RTM']

                # Create a mapping from expected columns to actual columns (case-insensitive)
                column_mapping = {}
                for expected_col in expected_columns:
                    expected_lower = expected_col.lower()
                    found = False
                    for actual_col in df.columns:
                        if actual_col.lower().strip() == expected_lower:
                            column_mapping[expected_col] = actual_col
                            found = True
                            break
                    if not found:
                        st.error(f"Required column '{expected_col}' not found in GDAM-RTM ratio file. Available columns: {list(df.columns)}")
                        ftp.quit()
                        return

                # Rename columns to standard names
                df = df.rename(columns={v: k for k, v in column_mapping.items()})

                # Validate and clean data
                df['Time Block'] = pd.to_numeric(df['Time Block'], errors='coerce')
                df['GDAM'] = pd.to_numeric(df['GDAM'], errors='coerce')
                df['RTM'] = pd.to_numeric(df['RTM'], errors='coerce')

                # Remove invalid rows
                df = df.dropna(subset=['Time Block', 'GDAM', 'RTM'])
                df = df[(df['Time Block'] >= 1) & (df['Time Block'] <= 96)]
                df['Time Block'] = df['Time Block'].astype(int)

                # Remove duplicates, keeping last
                df = df.drop_duplicates(subset=['Time Block'], keep='last')

                # Validate that GDAM + RTM = 1 for each block (with some tolerance)
                df['sum_check'] = df['GDAM'] + df['RTM']
                invalid_sums = df[abs(df['sum_check'] - 1.0) > 0.001]
                if not invalid_sums.empty:
                    st.warning(f"Warning: {len(invalid_sums)} blocks have GDAM + RTM ≠ 1.0")

                # Remove the sum_check column
                df = df.drop(columns=['sum_check'])

                if not df.empty:
                    # Load data into database
                    conn = get_db_connection()
                    cursor = conn.cursor()

                    # Clear existing data
                    cursor.execute("DELETE FROM gdam_rtm_ratio")
                    conn.commit()

                    # Insert new data
                    success_count = 0
                    for _, row in df.iterrows():
                        try:
                            query = "INSERT INTO gdam_rtm_ratio (block, gdam, rtm) VALUES (%s, %s, %s)"
                            cursor.execute(query, (int(row['Time Block']), float(row['GDAM']), float(row['RTM'])))
                            success_count += 1
                        except Exception as e:
                            st.error(f"Error inserting GDAM-RTM ratio for block {row['Time Block']}: {str(e)}")

                    # Ensure all 96 blocks exist with default values if missing
                    cursor.execute("SELECT block FROM gdam_rtm_ratio")
                    existing_blocks = set(row[0] for row in cursor.fetchall())
                    missing_blocks = set(range(1, 97)) - existing_blocks

                    if missing_blocks:
                        for block in missing_blocks:
                            cursor.execute(
                                "INSERT INTO gdam_rtm_ratio (block, gdam, rtm) VALUES (%s, %s, %s)",
                                (block, 0.5, 0.5)
                            )
                        success_count += len(missing_blocks)

                    conn.commit()
                    cursor.close()
                    conn.close()

                    st.success(f"Successfully loaded {success_count} GDAM-RTM ratio records from {latest_file}")
                else:
                    st.error("No valid GDAM-RTM ratio data found in file")
            else:
                st.error(f"Failed to download GDAM-RTM ratio file: {latest_file}")
        else:
            st.warning("No GDAM-RTM ratio files found in FTP folder. Using existing database values.")

    except Exception as e:
        st.error(f"Error loading GDAM-RTM ratio data: {str(e)}")

    ftp.quit()

def load_gdam_rtm_ratio():
    """
    Load GDAM-RTM ratio data from the latest modified file in the FTP folder.
    Expected file format:
    - Columns: Time Block, GDAM, RTM
    - 96 rows for time blocks 1-96
    - File extensions: .xlsx or .csv
    """
    st.subheader("Load GDAM-RTM Ratio Data from FTP")

    try:
        # Connect to FTP
        ftp = connect_ftp()

        # Find the latest modified file in the GDAM_RTM_Ratio folder
        latest_file = find_latest_modified_file(ftp, FTP_PATH_GDAM_RTM_RATIO, ['.xlsx', '.csv'])

        if not latest_file:
            st.error("No GDAM-RTM ratio files found in the FTP folder.")
            ftp.quit()
            return False

        st.info(f"Found latest file: {latest_file}")

        # Download the file
        bio = download_ftp_file(ftp, FTP_PATH_GDAM_RTM_RATIO, latest_file)
        ftp.quit()

        if not bio:
            st.error(f"Failed to download {latest_file}")
            return False

        # Read the file based on extension
        try:
            if latest_file.lower().endswith('.xlsx'):
                df = pd.read_excel(bio)
            else:
                df = pd.read_csv(bio)
        except Exception as e:
            st.error(f"Error reading file {latest_file}: {str(e)}")
            return False

        # Validate the file structure
        expected_columns = ['Time Block', 'GDAM', 'RTM']

        # Create a mapping from expected columns to actual columns
        column_mapping = {}
        for expected_col in expected_columns:
            expected_lower = expected_col.lower()
            found = False
            for actual_col in df.columns:
                if actual_col.lower().strip() == expected_lower:
                    column_mapping[expected_col] = actual_col
                    found = True
                    break
            if not found:
                st.error(f"Required column '{expected_col}' not found in file. Available columns: {list(df.columns)}")
                return False

        # Rename columns to standard names
        df = df.rename(columns={v: k for k, v in column_mapping.items()})

        # Validate data
        if df.empty:
            st.error("File is empty.")
            return False

        # Check for required columns after renaming
        missing_cols = [col for col in expected_columns if col not in df.columns]
        if missing_cols:
            st.error(f"Missing columns after processing: {missing_cols}")
            return False

        # Validate Time Block column
        df['Time Block'] = pd.to_numeric(df['Time Block'], errors='coerce')
        invalid_blocks = df[df['Time Block'].isna() | (df['Time Block'] < 1) | (df['Time Block'] > 96)]
        if not invalid_blocks.empty:
            st.error(f"Invalid time blocks found. Time blocks must be integers between 1 and 96.")
            st.dataframe(invalid_blocks)
            return False

        # Convert Time Block to integer
        df['Time Block'] = df['Time Block'].astype(int)

        # Validate GDAM and RTM columns
        df['GDAM'] = pd.to_numeric(df['GDAM'], errors='coerce')
        df['RTM'] = pd.to_numeric(df['RTM'], errors='coerce')

        invalid_values = df[df['GDAM'].isna() | df['RTM'].isna()]
        if not invalid_values.empty:
            st.error("Invalid GDAM or RTM values found. All values must be numeric.")
            st.dataframe(invalid_values)
            return False

        # Validate that GDAM and RTM are between 0 and 1
        invalid_range = df[(df['GDAM'] < 0) | (df['GDAM'] > 1) | (df['RTM'] < 0) | (df['RTM'] > 1)]
        if not invalid_range.empty:
            st.warning("GDAM and RTM values should be between 0 and 1.")
            st.dataframe(invalid_range)

        # Validate that GDAM + RTM = 1 for each block (with some tolerance)
        df['sum_check'] = df['GDAM'] + df['RTM']
        invalid_sums = df[abs(df['sum_check'] - 1.0) > 0.001]
        if not invalid_sums.empty:
            st.warning("Warning: GDAM + RTM should equal 1.0 for each block. The following blocks have invalid sums:")
            st.dataframe(invalid_sums[['Time Block', 'GDAM', 'RTM', 'sum_check']])

        # Remove the sum_check column
        df = df.drop(columns=['sum_check'])

        # Check for duplicate time blocks
        duplicates = df.duplicated(subset=['Time Block'], keep=False)
        if duplicates.any():
            st.warning("Duplicate time blocks found. Keeping only the last entry for each block.")
            df = df.drop_duplicates(subset=['Time Block'], keep='last')

        # Show preview of data
        st.write("Preview of data to be loaded:")
        st.dataframe(df.head(10))
        st.write(f"Total rows: {len(df)}")

        # Load data into database
        conn = get_db_connection()
        cursor = conn.cursor()

        # Clear existing data
        cursor.execute("DELETE FROM gdam_rtm_ratio")
        conn.commit()

        # Insert new data
        success_count = 0
        for _, row in df.iterrows():
            try:
                query = "INSERT INTO gdam_rtm_ratio (block, gdam, rtm) VALUES (%s, %s, %s)"
                cursor.execute(query, (int(row['Time Block']), float(row['GDAM']), float(row['RTM'])))
                success_count += 1
            except Exception as e:
                st.error(f"Error inserting row for block {row['Time Block']}: {str(e)}")

        conn.commit()
        cursor.close()
        conn.close()

        st.success(f"Successfully loaded {success_count} GDAM-RTM ratio records from {latest_file}")

        # Ensure all 96 blocks exist with default values if missing
        conn = get_db_connection()
        cursor = conn.cursor()

        # Check which blocks are missing
        cursor.execute("SELECT block FROM gdam_rtm_ratio")
        existing_blocks = set(row[0] for row in cursor.fetchall())
        missing_blocks = set(range(1, 97)) - existing_blocks

        if missing_blocks:
            st.info(f"Adding default values for missing blocks: {sorted(missing_blocks)}")
            for block in missing_blocks:
                cursor.execute(
                    "INSERT INTO gdam_rtm_ratio (block, gdam, rtm) VALUES (%s, %s, %s)",
                    (block, 0.5, 0.5)
                )
            conn.commit()
            st.success(f"Added default values (0.5, 0.5) for {len(missing_blocks)} missing blocks.")

        cursor.close()
        conn.close()

        return True

    except Exception as e:
        st.error(f"Error loading GDAM-RTM ratio data: {str(e)}")
        return False

# -------------------------------
# Manual Data Entry Functions
# -------------------------------

def execute_query_to_dataframe(query, conn=None, params=None):
    """
    Execute a SQL query and return results as a pandas DataFrame.
    This replaces execute_query_to_dataframe() to avoid SQLAlchemy dependency issues.
    """
    close_conn = False
    if conn is None:
        conn = get_db_connection()
        close_conn = True

    cursor = conn.cursor()

    try:
        if params is not None:
            cursor.execute(query, params)
        else:
            cursor.execute(query)
        data = cursor.fetchall()

        # Get column names from cursor description
        columns = [desc[0] for desc in cursor.description] if cursor.description else []

        # Create DataFrame
        df = pd.DataFrame(data, columns=columns)

    except Exception as e:
        print(f"Error executing query: {e}")
        df = pd.DataFrame()  # Return empty DataFrame on error
    finally:
        cursor.close()
        if close_conn:
            conn.close()

    return df

def display_table(table_name):
    query = f"SELECT * FROM {table_name}"
    return execute_query_to_dataframe(query)

def ensure_complete_blocks_for_load(df_wide, dates, buyer_columns):
    """Ensure that each date has all 96 time blocks (1-96) for load data.
    If any blocks are missing, they will be added with default values of 0 for all buyers."""
    complete_df = df_wide.copy()

    for date in dates:
        # Get all blocks for this date
        date_blocks = complete_df[complete_df['date'] == date]['block'].unique() if not complete_df.empty else []

        # Find missing blocks
        all_blocks = set(range(1, 97))
        existing_blocks = set(date_blocks)
        missing_blocks = all_blocks - existing_blocks

        # Add missing blocks with default values
        if missing_blocks:
            new_rows = []
            for block in missing_blocks:
                new_row = {'date': date, 'block': block}
                for buyer in buyer_columns:
                    new_row[buyer] = 0.0
                new_rows.append(new_row)

            # Use pd.concat instead of append (which is deprecated)
            if new_rows:
                new_df = pd.DataFrame(new_rows)
                complete_df = pd.concat([complete_df, new_df], ignore_index=True)

    return complete_df

def cleanup_duplicate_load_data():
    """Clean up duplicate date-block-buyer combinations in the load_data table."""
    conn = get_db_connection()
    cursor = conn.cursor()

    # First, identify all date-block-buyer combinations that have duplicates
    query = """
    SELECT date, block, buyer_name, COUNT(*) as count
    FROM load_data
    GROUP BY date, block, buyer_name
    HAVING COUNT(*) > 1
    """
    cursor.execute(query)
    duplicates = cursor.fetchall()

    for date, block, buyer_name, _ in duplicates:  # Using _ to ignore the count variable
        # Get all duplicate rows, ordered by id (to keep the most recent entry)
        query = "SELECT id FROM load_data WHERE date = %s AND block = %s AND buyer_name = %s ORDER BY id DESC"
        cursor.execute(query, (date, block, buyer_name))
        ids = [row[0] for row in cursor.fetchall()]

        # Keep only the first id (most recent due to DESC ordering)
        if len(ids) > 1:
            ids_to_delete = ids[1:]
            placeholders = ", ".join([str(id) for id in ids_to_delete])
            query = f"DELETE FROM load_data WHERE id IN ({placeholders})"
            cursor.execute(query)
            conn.commit()

    cursor.close()
    conn.close()

def manual_fill_load_data():
    st.subheader("Manual Data Entry: Load Data")

    # Add CSV Upload Section
    st.write("---")
    st.write("#### 📁 Upload CSV File to FTP")

    col1, col2 = st.columns([1, 1])

    with col1:
        upload_date = st.date_input(
            "Select Date for Upload",
            value=today_ist(),
            help="Select the date for which you want to upload load data"
        )

    with col2:
        uploaded_file = st.file_uploader(
            "Choose CSV or Excel file",
            type=['csv', 'xlsx'],
            help="Upload a CSV or Excel file with load data. File will be uploaded to FTP with correct naming convention."
        )



    # Show expected file format
    with st.expander("📋 Expected CSV Format"):
        st.write("**Required columns:**")
        st.code("""
Date        | Time Block | [Buyer1]     | [Buyer2]     | [Buyer3]     | ...
2025-10-11  | 1          | 5.0          | 8.0          | 7.5          | ...
2025-10-11  | 2          | 5.2          | 8.1          | 7.6          | ...
...         | ...        | ...          | ...          | ...          | ...
2025-10-11  | 96         | 4.8          | 7.9          | 7.3          | ...
        """)
        st.write("**Requirements:**")
        st.write("- Exactly 96 rows (time blocks 1-96)")
        st.write("- 'Date' column with YYYY-MM-DD format")
        st.write("- 'Time Block' column with values 1-96")
        st.write("- Additional columns for each buyer with load values (MW)")
        st.write("- Missing values will be treated as 0")

    if uploaded_file is not None and st.button("🚀 Upload to FTP", type="primary"):
        try:
            # Determine file extension and generate correct filename for FTP
            file_extension = uploaded_file.name.split('.')[-1].lower()
            if file_extension == 'xlsx':
                ftp_filename = upload_date.strftime("%Y%m%d") + ".xlsx"
            else:
                ftp_filename = upload_date.strftime("%Y%m%d") + ".csv"

            # Read and validate the uploaded file
            file_content = uploaded_file.read()
            uploaded_file.seek(0)  # Reset file pointer

            # Validate file structure based on type
            if file_extension == 'xlsx':
                df_validate = pd.read_excel(uploaded_file)
            else:
                df_validate = pd.read_csv(uploaded_file)
            required_columns = ['Date', 'Time Block']
            missing_columns = [col for col in required_columns if col not in df_validate.columns]

            if missing_columns:
                st.error(f"❌ Missing required columns: {missing_columns}")
                st.info("CSV file must have 'Date' and 'Time Block' columns, plus buyer columns")
            elif len(df_validate) != 96:
                st.error(f"❌ CSV file must have exactly 96 rows (time blocks), found {len(df_validate)} rows")
            else:
                # Check if file already exists on FTP
                ftp = connect_ftp()
                try:
                    ftp.cwd(FTP_PATH_LOAD)
                    existing_files = ftp.nlst()
                    file_exists = ftp_filename in existing_files
                except:
                    file_exists = False

                if file_exists:
                    st.warning(f"⚠️ File {ftp_filename} already exists on FTP server")
                    if not st.checkbox(f"✅ Confirm overwrite {ftp_filename}", key="confirm_overwrite"):
                        st.info("Upload cancelled. Check the box above to confirm overwrite.")
                        ftp.quit()
                        return

                # Upload to FTP
                file_data = BytesIO(file_content)
                success = upload_ftp_file(ftp, FTP_PATH_LOAD, ftp_filename, file_data)
                ftp.quit()

                if success:
                    st.success(f"✅ Successfully uploaded {ftp_filename} to FTP server!")
                    st.info(f"📍 File location: {FTP_PATH_LOAD}{ftp_filename}")

                    # Automatically load the data into the database
                    st.info("🔄 Processing uploaded file into database...")

                    try:
                            # Get database connection
                            conn = get_db_connection()
                            cursor = conn.cursor()

                            st.info(f"🔗 Database connection established")

                            # Clear existing entries for this date before inserting new ones
                            delete_query = "DELETE FROM load_data WHERE date = %s"
                            cursor.execute(delete_query, (upload_date,))
                            conn.commit()

                            st.info(f"🗑️ Cleared existing data for {upload_date}")

                            # Process the uploaded data into the database
                            # Use the same logic as the existing load_data_to_mysql function

                            records_inserted = 0
                            buyer_columns = [col for col in df_validate.columns if col not in ['Date', 'Time Block']]

                            st.info(f"📊 Processing {len(df_validate)} rows with {len(buyer_columns)} buyer columns")

                            for _, row in df_validate.iterrows():
                                # Use upload_date as the date value (same as existing load_data_to_mysql function)
                                date_val = upload_date

                                block = row.get('Time Block', None)
                                if pd.isna(block):
                                    block = 0  # Or another appropriate default.
                                try:
                                    block = int(block)
                                except Exception:
                                    block = 0

                                # Process each buyer column (ignoring Date and Time Block)
                                for col in buyer_columns:
                                    # Force buyer_name to a string. If the column name is empty or "nan", replace with "0".
                                    buyer_name = str(col).strip()
                                    if buyer_name == "" or buyer_name.lower() == "nan":
                                        buyer_name = "0"

                                    load_value = row[col]
                                    if pd.isna(load_value):
                                        load_value = 0

                                    query = """INSERT INTO load_data (date, block, buyer_name, load_value)
                                               VALUES (%s, %s, %s, %s)
                                               ON DUPLICATE KEY UPDATE load_value = VALUES(load_value)"""
                                    cursor.execute(query, (date_val, block, buyer_name, load_value))
                                    records_inserted += 1

                            conn.commit()
                            cursor.close()
                            conn.close()

                            st.info(f"💾 Committed {records_inserted} records to database")

                            st.success(f"✅ Data successfully loaded into database for {upload_date}!")
                            st.info("📊 Load data is now ready for bifurcation analysis")

                            # Show preview of uploaded data
                            with st.expander("📊 Preview of uploaded data"):
                                st.dataframe(df_validate.head(10))
                                st.write(f"**Buyer columns found:** {', '.join(buyer_columns)}")
                                st.write(f"**Date range:** {df_validate['Date'].min()} to {df_validate['Date'].max()}")
                                st.write(f"**Time blocks:** {df_validate['Time Block'].min()} to {df_validate['Time Block'].max()}")
                                st.write(f"**Database records created:** {records_inserted}")

                            # Refresh the display to show updated data
                            st.rerun()

                    except Exception as db_error:
                        st.error(f"❌ Failed to load data into database: {str(db_error)}")
                        st.warning("File was uploaded to FTP successfully, but database loading failed.")
                        st.info("You can try using 'Load Data' function to process the file manually.")

                else:
                    st.error("❌ Failed to upload file to FTP server")

        except Exception as e:
            st.error(f"❌ Error processing file: {str(e)}")

    st.write("---")
    st.write("#### ✏️ Manual Data Entry")

    # Clean up any duplicate entries in the load_data table
    cleanup_duplicate_load_data()

    df = display_table("load_data")
    today = today_ist()
    tomorrow = today + datetime.timedelta(days=1)

    if not df.empty:
        df_wide = df.pivot_table(
            index=['date', 'block'],
            columns='buyer_name',
            values='load_value',
            aggfunc='first'
        ).reset_index()
    else:
        df_wide = pd.DataFrame(columns=['date', 'block'])

    if 'date' in df_wide.columns:
        df_wide['date'] = pd.to_datetime(df_wide['date']).dt.date
        df_wide = df_wide[df_wide['date'].isin([today, tomorrow])]

    # Get all buyer columns
    buyer_columns = [col for col in df_wide.columns if col not in ['date', 'block']]

    # Ensure all 96 blocks exist for today and tomorrow
    if not df_wide.empty:
        df_wide = ensure_complete_blocks_for_load(df_wide, [today, tomorrow], buyer_columns)

    # Add a helper for clipboard paste
    st.write("#### Quick Paste Helper")
    st.write("If you're having trouble pasting multiple rows directly into the table below, use this helper:")

    col1, col2, col3 = st.columns(3)
    with col1:
        paste_start_block = st.number_input("Starting Block Number", min_value=1, max_value=96, value=1, step=1, key="load_start_block")
    with col2:
        paste_date = st.date_input("Date for Pasted Values", value=today, key="load_date")
    with col3:
        buyer_options = [""] + buyer_columns if buyer_columns else [""]
        paste_buyer = st.selectbox("Buyer Name", options=buyer_options, key="load_buyer")

    paste_values = st.text_area("Paste Load Values (one per line)", height=100, key="load_values")

    if st.button("Apply Pasted Values", key="load_apply"):
        if paste_values.strip() and paste_buyer:
            try:
                # Parse the pasted values
                values = [float(v.strip()) for v in paste_values.strip().split('\n') if v.strip()]

                # Create a dataframe with the pasted values
                blocks = range(paste_start_block, paste_start_block + len(values))
                if max(blocks) > 96:
                    st.warning(f"Some values exceed block 96 and will be ignored.")
                    blocks = range(paste_start_block, min(paste_start_block + len(values), 97))
                    values = values[:len(blocks)]

                # Insert the data into the database
                conn = get_db_connection()
                cursor = conn.cursor()

                # For each block, either update existing record or insert new one
                for block, value in zip(blocks, values):
                    # Check if this specific block exists for this date and buyer
                    check_block_query = "SELECT id FROM load_data WHERE date = %s AND block = %s AND buyer_name = %s"
                    cursor.execute(check_block_query, (paste_date, block, paste_buyer))
                    result = cursor.fetchone()

                    if result:
                        # Update existing record
                        update_query = "UPDATE load_data SET load_value = %s WHERE date = %s AND block = %s AND buyer_name = %s"
                        cursor.execute(update_query, (value, paste_date, block, paste_buyer))
                    else:
                        # Insert new record
                        insert_query = "INSERT INTO load_data (date, block, buyer_name, load_value) VALUES (%s, %s, %s, %s)"
                        cursor.execute(insert_query, (paste_date, block, paste_buyer, value))

                conn.commit()
                cursor.close()
                conn.close()

                st.success(f"Successfully applied {len(values)} pasted values for {paste_buyer} starting from block {paste_start_block}.")

                # Force a rerun of the app to refresh the data
                st.rerun()
            except Exception as e:
                st.error(f"Error processing pasted values: {str(e)}")
        elif not paste_buyer:
            st.error("Please select a buyer name.")
        else:
            st.error("Please enter values to paste.")

    st.write("---")
    st.write("#### Edit Data Directly")

    edited_df = st.data_editor(df_wide, num_rows="dynamic", use_container_width=True)

    # Handle duplicate date-block combinations by keeping only the last entry
    if not edited_df.empty:
        # Check for duplicates in the edited dataframe
        duplicates = edited_df.duplicated(subset=['date', 'block'], keep=False)
        if duplicates.any():
            # Keep only the last entry for each date-block combination
            edited_df = edited_df.drop_duplicates(subset=['date', 'block'], keep='last')

        # Ensure blocks are between 1 and 96
        if 'block' in edited_df.columns:
            edited_df['block'] = edited_df['block'].astype(int)
            invalid_blocks = edited_df[(edited_df['block'] < 1) | (edited_df['block'] > 96)]
            if not invalid_blocks.empty:
                st.error(f"Time blocks must be between 1 and 96. Please fix these rows:\n{invalid_blocks}")
                return

        # Ensure all 96 blocks exist for each date
        buyer_columns = [col for col in edited_df.columns if col not in ['date', 'block']]
        unique_dates = edited_df['date'].unique()
        edited_df = ensure_complete_blocks_for_load(edited_df, unique_dates, buyer_columns)

    conn = get_db_connection()
    cursor = conn.cursor()

    # Delete only today's and tomorrow's data
    cursor.execute("DELETE FROM load_data WHERE date IN (%s, %s)", (today, tomorrow))
    conn.commit()

    buyer_columns = [col for col in edited_df.columns if col not in ['date', 'block']]
    df_long = edited_df.melt(
        id_vars=['date', 'block'],
        value_vars=buyer_columns,
        var_name='buyer_name',
        value_name='load_value'
    ).dropna(subset=['load_value'])

    for _, row in df_long.iterrows():
        # Convert numpy.float64 to Python float
        load_value = float(row["load_value"]) if row["load_value"] is not None else 0.0
        query = "REPLACE INTO load_data (date, block, buyer_name, load_value) VALUES (%s, %s, %s, %s)"
        cursor.execute(query, (row["date"], row["block"], row["buyer_name"], load_value))

    conn.commit()
    cursor.close()
    conn.close()
    st.success("Load data updated successfully.")

    # Add a button to delete all data except yesterday, today, and tomorrow
    st.write("---")
    st.write("### Danger Zone: Delete All Data Except Yesterday, Today, and Tomorrow")
    if st.button("Delete All Other Data", key="delete_except_recent"):
        today = today_ist()
        yesterday = today - datetime.timedelta(days=1)
        tomorrow = today + datetime.timedelta(days=1)
        conn = get_db_connection()
        cursor = conn.cursor()
        for table in ["load_data", "obligation", "schedule_wind", "schedule_solar", "hybrid_avc_cap"]:
            cursor.execute(f"DELETE FROM {table} WHERE date NOT IN (%s, %s, %s)", (yesterday, today, tomorrow))
        conn.commit()
        cursor.close()
        conn.close()
        st.success("All data except for yesterday, today, and tomorrow has been deleted from all relevant tables.")

def ensure_complete_blocks(df, dates):
    """Ensure that each date has all 96 time blocks (1-96).
    If any blocks are missing, they will be added with a default value of 0."""
    complete_df = df.copy()

    for date in dates:
        # Get all blocks for this date
        date_blocks = complete_df[complete_df['date'] == date]['block'].unique()

        # Find missing blocks
        all_blocks = set(range(1, 97))
        existing_blocks = set(date_blocks)
        missing_blocks = all_blocks - existing_blocks

        # Add missing blocks with default values
        if missing_blocks:
            new_rows = []
            for block in missing_blocks:
                if 'market' in complete_df.columns:  # For obligation table
                    new_row = {'date': date, 'block': block, 'market': 0.0}
                    if 'market_fp' in complete_df.columns:
                        new_row['market_fp'] = 0.0
                    if 'market_linde' in complete_df.columns:
                        new_row['market_linde'] = 0.0
                    new_rows.append(new_row)
                elif 'sch' in complete_df.columns:  # For schedule tables
                    new_row = {'date': date, 'block': block, 'sch': 0.0}
                    if 'avc' in complete_df.columns:
                        new_row['avc'] = 0.0
                    new_rows.append(new_row)

            # Use pd.concat instead of append (which is deprecated)
            if new_rows:
                new_df = pd.DataFrame(new_rows)
                complete_df = pd.concat([complete_df, new_df], ignore_index=True)

    return complete_df

def cleanup_duplicate_dates(table_name, date_column="date", block_column="block"):
    """Clean up duplicate date ranges in the specified table.
    This function will keep only one set of blocks 1-96 for each date."""
    conn = get_db_connection()
    cursor = conn.cursor()

    # First, identify all dates in the table
    query = f"SELECT DISTINCT {date_column} FROM {table_name}"
    cursor.execute(query)
    dates = [row[0] for row in cursor.fetchall()]

    for date in dates:
        # For each date, check if there are duplicate blocks
        query = f"SELECT COUNT(*) FROM {table_name} WHERE {date_column} = %s"
        cursor.execute(query, (date,))
        count = cursor.fetchone()[0]

        # If there are more than 96 blocks for this date, there are duplicates
        if count > 96:
            st.warning(f"Found {count} blocks for date {date} in {table_name} table. Cleaning up duplicates...")

            # Get all rows for this date, ordered by id (to keep the most recent entries)
            query = f"SELECT id, {block_column} FROM {table_name} WHERE {date_column} = %s ORDER BY id DESC"
            cursor.execute(query, (date,))
            rows = cursor.fetchall()

            # Keep track of blocks we've seen
            seen_blocks = set()
            ids_to_keep = []

            # Keep only the first occurrence of each block (which is the most recent due to DESC ordering)
            for id, block in rows:
                if block not in seen_blocks:
                    seen_blocks.add(block)
                    ids_to_keep.append(id)

            # Delete all rows for this date except the ones we want to keep
            if ids_to_keep:
                placeholders = ", ".join([str(id) for id in ids_to_keep])
                query = f"DELETE FROM {table_name} WHERE {date_column} = %s AND id NOT IN ({placeholders})"
                cursor.execute(query, (date,))
                conn.commit()
                st.success(f"Cleaned up duplicate blocks for date {date} in {table_name} table.")

    cursor.close()
    conn.close()

def manual_fill_obligation_data():
    st.subheader("Obligation Data")

    # Clean up any duplicate date ranges in the obligation table
    cleanup_duplicate_dates("obligation")


    df = display_table("obligation")
    today = today_ist()
    tomorrow = today + datetime.timedelta(days=1)

    if df.empty:
        df = pd.DataFrame(columns=["date", "block", "market_fp", "market"])
    else:
        df = normalize_obligation_dataframe(df)
        df['date'] = pd.to_datetime(df['date']).dt.date
        df = df[df['date'].isin([today, tomorrow])]

    # Ensure all 96 blocks exist for today and tomorrow
    if not df.empty:
        df = ensure_complete_blocks(df, [today, tomorrow])
        # Sort by date and then by block number
        df = df.sort_values(by=['date', 'block'])

    st.subheader("Edit Obligation Data")
    df = df.drop(columns=['id', 'market_linde', 'market'], errors='ignore')

    # Configure the data editor with column configurations to better handle clipboard operations
    column_config = {
        "date": st.column_config.DateColumn(
            "Date",
            help="The date for this obligation entry",
            format="YYYY-MM-DD",
        ),
        "block": st.column_config.NumberColumn(
            "Block",
            help="Time block (1-96)",
            min_value=1,
            max_value=96,
            step=1,
            format="%d"
        ),
        "market_fp": st.column_config.NumberColumn(
            "G-DAM",
            help="Manual G-DAM value for the block",
            format="%.1f"
        ),
    }

    # Add a helper for clipboard paste
    st.write("#### Quick Paste Helper")
    st.write("If you're having trouble pasting multiple rows directly into the table below, use this helper:")

    col1, col2 = st.columns(2)
    with col1:
        paste_start_block = st.number_input("Starting Block Number", min_value=1, max_value=96, value=1, step=1)
    with col2:
        paste_date = st.date_input("Date for Pasted Values", value=today)

    paste_values = st.text_area("Paste G-DAM Values (one per line)", height=100)

    if st.button("Apply Pasted Values"):
        if paste_values.strip():
            try:
                # Parse the pasted values
                values = [float(v.strip()) for v in paste_values.strip().split('\n') if v.strip()]

                # Create a dataframe with the pasted values
                blocks = range(paste_start_block, paste_start_block + len(values))
                if max(blocks) > 96:
                    st.warning(f"Some values exceed block 96 and will be ignored.")
                    blocks = range(paste_start_block, min(paste_start_block + len(values), 97))
                    values = values[:len(blocks)]

                paste_df = pd.DataFrame({
                    "date": [paste_date] * len(blocks),
                    "block": blocks,
                    "target_value": values
                })

                # Insert the data into the database
                conn = get_db_connection()
                cursor = conn.cursor()

                for _, row in paste_df.iterrows():
                    check_block_query = "SELECT id, market_fp FROM obligation WHERE date = %s AND block = %s"
                    cursor.execute(check_block_query, (row["date"], row["block"]))
                    result = cursor.fetchone()

                    market_fp_value = to_float(row["target_value"])
                    market_total = round(market_fp_value, 1)

                    if result:
                        # Update existing record
                        update_query = """
                            UPDATE obligation
                            SET market = %s, market_fp = %s, market_linde = %s
                            WHERE date = %s AND block = %s
                        """
                        cursor.execute(
                            update_query,
                            (market_total, market_fp_value, 0.0, row["date"], row["block"])
                        )
                    else:
                        # Insert new record
                        insert_query = """
                            INSERT INTO obligation (date, block, market, market_fp, market_linde)
                            VALUES (%s, %s, %s, %s, %s)
                        """
                        cursor.execute(
                            insert_query,
                            (row["date"], row["block"], market_total, market_fp_value, 0.0)
                        )

                conn.commit()
                cursor.close()
                conn.close()

                st.success(f"Successfully applied {len(paste_df)} pasted values starting from block {paste_start_block}.")

                # Force a rerun of the app to refresh the data
                st.rerun()
            except Exception as e:
                st.error(f"Error processing pasted values: {str(e)}")

    st.write("---")
    st.write("#### Edit Data Directly")

    edited_df = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        column_config=column_config,
        key="obligation_editor",
    )

    # Handle duplicate date-block combinations by keeping only the last entry
    if not edited_df.empty:
        for col in ["market_fp"]:
            if col not in edited_df.columns:
                edited_df[col] = 0.0
            edited_df[col] = edited_df[col].apply(to_float)

        # Check for duplicates in the edited dataframe
        duplicates = edited_df.duplicated(subset=['date', 'block'], keep=False)
        if duplicates.any():
            # Keep only the last entry for each date-block combination
            edited_df = edited_df.drop_duplicates(subset=['date', 'block'], keep='last')
            st.warning(f"Found duplicate date-block combinations. Keeping only the last entry for each combination.")

        # Ensure blocks are between 1 and 96
        invalid_blocks = edited_df[(edited_df['block'] < 1) | (edited_df['block'] > 96)]
        if not invalid_blocks.empty:
            st.error(f"Time blocks must be between 1 and 96. Please fix these rows:\n{invalid_blocks}")
            return

        # Ensure all 96 blocks exist for each date
        unique_dates = edited_df['date'].unique()
        edited_df = ensure_complete_blocks(edited_df, unique_dates)

        # Sort by date and then by block number
        edited_df = edited_df.sort_values(by=['date', 'block'])

    conn = get_db_connection()
    cursor = conn.cursor()

    # Delete only today's and tomorrow's data
    cursor.execute("DELETE FROM obligation WHERE date IN (%s, %s)", (today, tomorrow))
    conn.commit()

    for _, row in edited_df.iterrows():
        market_fp_value = to_float(row.get("market_fp"), 0.0)
        market_value = round(market_fp_value, 1)
        query = """
            REPLACE INTO obligation (date, block, market, market_fp, market_linde)
            VALUES (%s, %s, %s, %s, %s)
        """
        cursor.execute(
            query,
            (row["date"], row["block"], market_value, market_fp_value, 0.0)
        )

    conn.commit()
    cursor.close()
    conn.close()
    st.success("Obligation data updated successfully.")

def manual_fill_gdam_rtm_ratio():
    st.subheader("GDAM-RTM Ratio Data")

    df = display_table("gdam_rtm_ratio")

    if df.empty:
        df = pd.DataFrame(columns=["block", "gdam", "rtm"])

    # Ensure all 96 blocks exist
    existing_blocks = set(df['block'].tolist()) if not df.empty else set()
    missing_blocks = set(range(1, 97)) - existing_blocks

    if missing_blocks:
        # Add missing blocks with default values
        missing_data = []
        for block in sorted(missing_blocks):
            missing_data.append({"block": block, "gdam": 0.5, "rtm": 0.5})
        missing_df = pd.DataFrame(missing_data)
        df = pd.concat([df, missing_df], ignore_index=True)

    # Sort by block number
    df = df.sort_values(by=['block'])
    df = df.drop(columns=['id'], errors='ignore')

    # Configure the data editor with column configurations
    column_config = {
        "block": st.column_config.NumberColumn(
            "Time Block",
            help="Time block (1-96)",
            min_value=1,
            max_value=96,
            step=1,
            format="%d"
        ),
        "gdam": st.column_config.NumberColumn(
            "GDAM",
            help="GDAM ratio (should sum with RTM to 1.0)",
            min_value=0.0,
            max_value=1.0,
            step=0.01,
            format="%.3f"
        ),
        "rtm": st.column_config.NumberColumn(
            "RTM",
            help="RTM ratio (should sum with GDAM to 1.0)",
            min_value=0.0,
            max_value=1.0,
            step=0.01,
            format="%.3f"
        )
    }

    # Add a helper for clipboard paste
    st.write("#### Quick Paste Helper")
    st.write("If you're having trouble pasting multiple rows directly into the table below, use this helper:")

    col1, col2 = st.columns(2)
    with col1:
        paste_start_block = st.number_input("Starting Block Number", min_value=1, max_value=96, value=1, step=1, key="gdam_rtm_start_block")
    with col2:
        paste_column = st.selectbox("Column to Paste Into", options=["gdam", "rtm"], key="gdam_rtm_column")

    paste_values = st.text_area("Paste Values (one per line)", height=100, key="gdam_rtm_values")

    if st.button("Apply Pasted Values", key="gdam_rtm_apply"):
        if paste_values.strip():
            try:
                # Parse the pasted values
                values = [float(v.strip()) for v in paste_values.strip().split('\n') if v.strip()]

                # Create a list of blocks to update
                blocks = range(paste_start_block, paste_start_block + len(values))
                if max(blocks) > 96:
                    st.warning(f"Some values exceed block 96 and will be ignored.")
                    blocks = range(paste_start_block, min(paste_start_block + len(values), 97))
                    values = values[:len(blocks)]

                # Update the database
                conn = get_db_connection()
                cursor = conn.cursor()

                for i, block in enumerate(blocks):
                    # Use INSERT ... ON DUPLICATE KEY UPDATE to handle both insert and update cases
                    if paste_column == "gdam":
                        # When updating GDAM, calculate RTM as 1 - GDAM for new records, but preserve existing RTM for updates
                        query = """
                        INSERT INTO gdam_rtm_ratio (block, gdam, rtm)
                        VALUES (%s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                        gdam = VALUES(gdam)
                        """
                        # For new records, set RTM to complement GDAM; for existing records, only update GDAM
                        other_value = 1.0 - values[i]
                        cursor.execute(query, (block, values[i], other_value))
                    else:  # paste_column == "rtm"
                        # When updating RTM, calculate GDAM as 1 - RTM for new records, but preserve existing GDAM for updates
                        query = """
                        INSERT INTO gdam_rtm_ratio (block, gdam, rtm)
                        VALUES (%s, %s, %s)
                        ON DUPLICATE KEY UPDATE
                        rtm = VALUES(rtm)
                        """
                        # For new records, set GDAM to complement RTM; for existing records, only update RTM
                        other_value = 1.0 - values[i]
                        cursor.execute(query, (block, other_value, values[i]))

                conn.commit()
                cursor.close()
                conn.close()

                st.success(f"Successfully applied {len(values)} pasted values to {paste_column} starting from block {paste_start_block}.")

                # Force a rerun of the app to refresh the data
                st.rerun()
            except Exception as e:
                st.error(f"Error processing pasted values: {str(e)}")

    st.write("---")
    st.write("#### Edit Data Directly")

    edited_df = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        column_config=column_config,
        key="gdam_rtm_editor"
    )

    # Validate the data
    if not edited_df.empty:
        # Ensure blocks are between 1 and 96
        invalid_blocks = edited_df[(edited_df['block'] < 1) | (edited_df['block'] > 96)]
        if not invalid_blocks.empty:
            st.error(f"Time blocks must be between 1 and 96. Please fix these rows:\n{invalid_blocks}")
            return

        # Check for duplicate blocks
        duplicates = edited_df.duplicated(subset=['block'], keep=False)
        if duplicates.any():
            edited_df = edited_df.drop_duplicates(subset=['block'], keep='last')
            st.warning(f"Found duplicate blocks. Keeping only the last entry for each block.")

        # Validate that GDAM + RTM = 1 for each block (with some tolerance)
        edited_df['sum_check'] = edited_df['gdam'] + edited_df['rtm']
        invalid_sums = edited_df[abs(edited_df['sum_check'] - 1.0) > 0.001]
        if not invalid_sums.empty:
            st.warning("Warning: GDAM + RTM should equal 1.0 for each block. The following blocks have invalid sums:")
            st.dataframe(invalid_sums[['block', 'gdam', 'rtm', 'sum_check']])

        # Remove the sum_check column before saving
        edited_df = edited_df.drop(columns=['sum_check'])

        # Ensure all 96 blocks exist
        existing_blocks = set(edited_df['block'].tolist())
        missing_blocks = set(range(1, 97)) - existing_blocks

        if missing_blocks:
            # Add missing blocks with default values
            missing_data = []
            for block in sorted(missing_blocks):
                missing_data.append({"block": block, "gdam": 0.5, "rtm": 0.5})
            missing_df = pd.DataFrame(missing_data)
            edited_df = pd.concat([edited_df, missing_df], ignore_index=True)

        # Sort by block number
        edited_df = edited_df.sort_values(by=['block'])

    conn = get_db_connection()
    cursor = conn.cursor()

    # Clear existing data
    cursor.execute("DELETE FROM gdam_rtm_ratio")
    conn.commit()

    # Insert updated data
    for _, row in edited_df.iterrows():
        gdam_value = float(row["gdam"]) if row["gdam"] is not None else 0.5
        rtm_value = float(row["rtm"]) if row["rtm"] is not None else 0.5
        query = "INSERT INTO gdam_rtm_ratio (block, gdam, rtm) VALUES (%s, %s, %s)"
        cursor.execute(query, (int(row["block"]), gdam_value, rtm_value))

    conn.commit()
    cursor.close()
    conn.close()
    st.success("GDAM-RTM Ratio data updated successfully.")

def manual_fill_schedule_combined():
    create_tables()

    st.subheader("Manual Data Entry: Solar and Wind Schedules")

    # Clean up any duplicate date ranges in both tables
    cleanup_duplicate_dates("schedule_solar", block_column="block", date_column="date")
    cleanup_duplicate_dates("schedule_wind", block_column="block", date_column="date")

    # Get data for both solar and wind
    df_solar = display_table("schedule_solar")
    df_wind = display_table("schedule_wind")
    today = today_ist()
    tomorrow = today + datetime.timedelta(days=1)

    # Process solar data
    if df_solar.empty:
        df_solar = pd.DataFrame(columns=["date", "block", "sch", "avc"])
    else:
        df_solar['date'] = pd.to_datetime(df_solar['date']).dt.date
        df_solar = df_solar[df_solar['date'].isin([today, tomorrow])]
        if 'avc' not in df_solar.columns:
            df_solar['avc'] = 0.0

    # Process wind data
    if df_wind.empty:
        df_wind = pd.DataFrame(columns=["date", "block", "sch", "avc"])
    else:
        df_wind['date'] = pd.to_datetime(df_wind['date']).dt.date
        df_wind = df_wind[df_wind['date'].isin([today, tomorrow])]
        if 'avc' not in df_wind.columns:
            df_wind['avc'] = 0.0

    # Ensure all 96 blocks exist for today and tomorrow
    if not df_solar.empty:
        df_solar = ensure_complete_blocks(df_solar, [today, tomorrow])
        # Sort by date and then by block number
        df_solar = df_solar.sort_values(by=['date', 'block'])
    if not df_wind.empty:
        df_wind = ensure_complete_blocks(df_wind, [today, tomorrow])
        # Sort by date and then by block number
        df_wind = df_wind.sort_values(by=['date', 'block'])

    # Add a helper for clipboard paste
    st.write("#### Quick Paste Helper")
    st.write("If you're having trouble pasting multiple rows directly into the tables below, use this helper:")

    col1, col2, col3, col4 = st.columns(4)
    with col1:
        paste_start_block = st.number_input("Starting Block Number", min_value=1, max_value=96, value=1, step=1, key="schedule_start_block")
    with col2:
        paste_date = st.date_input("Date for Pasted Values", value=today, key="schedule_date")
    with col3:
        schedule_type = st.selectbox("Schedule Type", options=["Solar", "Wind"], key="schedule_type")
    with col4:
        schedule_value_type = st.selectbox("Value Type", options=["Schedule", "AvC"], key="schedule_value_type")

    paste_values = st.text_area("Paste Schedule Values (one per line)", height=100, key="schedule_values")

    if st.button("Apply Pasted Values", key="schedule_apply"):
        if paste_values.strip():
            try:
                # Parse the pasted values
                values = [float(v.strip()) for v in paste_values.strip().split('\n') if v.strip()]

                # Create a dataframe with the pasted values
                blocks = range(paste_start_block, paste_start_block + len(values))
                if max(blocks) > 96:
                    st.warning(f"Some values exceed block 96 and will be ignored.")
                    blocks = range(paste_start_block, min(paste_start_block + len(values), 97))
                    values = values[:len(blocks)]

                # Insert the data into the database
                conn = get_db_connection()
                cursor = conn.cursor()

                table_name = "schedule_solar" if schedule_type == "Solar" else "schedule_wind"
                value_column = "sch" if schedule_value_type == "Schedule" else "avc"

                # We'll use UPDATE or INSERT for each block individually

                # For each block, either update existing record or insert new one
                for block, value in zip(blocks, values):
                    # Check if this specific block exists for this date
                    check_block_query = f"SELECT id FROM {table_name} WHERE date = %s AND block = %s"
                    cursor.execute(check_block_query, (paste_date, block))
                    result = cursor.fetchone()

                    if result:
                        # Update existing record
                        update_query = f"UPDATE {table_name} SET {value_column} = %s WHERE date = %s AND block = %s"
                        cursor.execute(update_query, (value, paste_date, block))
                    else:
                        # Insert new record
                        insert_query = f"INSERT INTO {table_name} (date, block, sch, avc) VALUES (%s, %s, %s, %s)"
                        if value_column == "sch":
                            cursor.execute(insert_query, (paste_date, block, value, 0.0))
                        else:
                            cursor.execute(insert_query, (paste_date, block, 0.0, value))

                conn.commit()

                # No need to execute a sort query here as we'll sort after reloading the data

                cursor.close()
                conn.close()

                st.success(f"Successfully applied {len(values)} pasted values for {schedule_type} schedule starting from block {paste_start_block}.")

                # Force a rerun of the app to refresh the data
                st.rerun()
            except Exception as e:
                st.error(f"Error processing pasted values: {str(e)}")

    st.write("---")

    # Display solar and wind schedules side by side
    col1, col2 = st.columns(2)

    with col1:
        st.subheader("Edit Solar Schedule")
        df_solar = df_solar.drop(columns=['id'], errors='ignore')
        edited_solar = st.data_editor(df_solar, num_rows="dynamic", use_container_width=True, key="solar_editor")

        edited_solar = edited_solar.dropna(subset=['date', 'block'])
        edited_solar['block'] = edited_solar['block'].astype(int)
        edited_solar['sch'] = edited_solar['sch'].where(pd.notnull(edited_solar['sch']), None)
        if 'avc' not in edited_solar.columns:
            edited_solar['avc'] = 0.0
        edited_solar['avc'] = edited_solar['avc'].where(pd.notnull(edited_solar['avc']), 0.0)

        # Handle duplicate date-block combinations by keeping only the last entry
        if not edited_solar.empty:
            # Check for duplicates in the edited dataframe
            duplicates = edited_solar.duplicated(subset=['date', 'block'], keep=False)
            if duplicates.any():
                # Keep only the last entry for each date-block combination
                edited_solar = edited_solar.drop_duplicates(subset=['date', 'block'], keep='last')
                st.warning(f"Found duplicate date-block combinations in Solar schedule. Keeping only the last entry for each combination.")

            # Ensure blocks are between 1 and 96
            invalid_blocks = edited_solar[(edited_solar['block'] < 1) | (edited_solar['block'] > 96)]
            if not invalid_blocks.empty:
                st.error(f"Time blocks must be between 1 and 96 in Solar schedule. Please fix these rows:\n{invalid_blocks}")
                return

            # Ensure all 96 blocks exist for each date
            unique_dates = edited_solar['date'].unique()
            edited_solar = ensure_complete_blocks(edited_solar, unique_dates)
            # Sort by date and then by block number
            edited_solar = edited_solar.sort_values(by=['date', 'block'])

    with col2:
        st.subheader("Edit Wind Schedule")
        df_wind = df_wind.drop(columns=['id'], errors='ignore')
        edited_wind = st.data_editor(df_wind, num_rows="dynamic", use_container_width=True, key="wind_editor")

        edited_wind = edited_wind.dropna(subset=['date', 'block'])
        edited_wind['block'] = edited_wind['block'].astype(int)
        edited_wind['sch'] = edited_wind['sch'].where(pd.notnull(edited_wind['sch']), None)
        if 'avc' not in edited_wind.columns:
            edited_wind['avc'] = 0.0
        edited_wind['avc'] = edited_wind['avc'].where(pd.notnull(edited_wind['avc']), 0.0)

        # Handle duplicate date-block combinations by keeping only the last entry
        if not edited_wind.empty:
            # Check for duplicates in the edited dataframe
            duplicates = edited_wind.duplicated(subset=['date', 'block'], keep=False)
            if duplicates.any():
                # Keep only the last entry for each date-block combination
                edited_wind = edited_wind.drop_duplicates(subset=['date', 'block'], keep='last')
                st.warning(f"Found duplicate date-block combinations in Wind schedule. Keeping only the last entry for each combination.")

            # Ensure blocks are between 1 and 96
            invalid_blocks = edited_wind[(edited_wind['block'] < 1) | (edited_wind['block'] > 96)]
            if not invalid_blocks.empty:
                st.error(f"Time blocks must be between 1 and 96 in Wind schedule. Please fix these rows:\n{invalid_blocks}")
                return

            # Ensure all 96 blocks exist for each date
            unique_dates = edited_wind['date'].unique()
            edited_wind = ensure_complete_blocks(edited_wind, unique_dates)
            # Sort by date and then by block number
            edited_wind = edited_wind.sort_values(by=['date', 'block'])

    # Save button for both schedules
    if st.button("Save Both Schedules"):
        conn = get_db_connection()
        cursor = conn.cursor()

        # Update Solar Schedule
        cursor.execute("DELETE FROM schedule_solar WHERE date IN (%s, %s)", (today, tomorrow))
        conn.commit()

        # Convert numpy.float64 to Python float for each row
        solar_data_tuples = []
        for _, row in edited_solar.iterrows():
            sch_value = float(row["sch"]) if row["sch"] is not None else 0.0
            avc_value = to_float(row.get("avc"), 0.0)
            solar_data_tuples.append((row["date"], row["block"], sch_value, avc_value))

        if solar_data_tuples:
            insert_query = "REPLACE INTO schedule_solar (date, block, sch, avc) VALUES (%s, %s, %s, %s)"
            cursor.executemany(insert_query, solar_data_tuples)
            conn.commit()

        # Update Wind Schedule
        cursor.execute("DELETE FROM schedule_wind WHERE date IN (%s, %s)", (today, tomorrow))
        conn.commit()

        # Convert numpy.float64 to Python float for each row
        wind_data_tuples = []
        for _, row in edited_wind.iterrows():
            sch_value = float(row["sch"]) if row["sch"] is not None else 0.0
            avc_value = to_float(row.get("avc"), 0.0)
            wind_data_tuples.append((row["date"], row["block"], sch_value, avc_value))

        if wind_data_tuples:
            insert_query = "REPLACE INTO schedule_wind (date, block, sch, avc) VALUES (%s, %s, %s, %s)"
            cursor.executemany(insert_query, wind_data_tuples)
            conn.commit()

        cursor.close()
        conn.close()
        st.success("Solar and Wind schedules updated successfully.")

        # Force a rerun of the app to refresh the data
        st.rerun()

# Keep the original functions for backward compatibility
def manual_fill_schedule_solar():
    st.warning("This page has been replaced by the combined Solar and Wind schedule page. Please use the 'Solar and Wind Schedules' option instead.")
    manual_fill_schedule_combined()

def manual_fill_schedule_wind():
    st.warning("This page has been replaced by the combined Solar and Wind schedule page. Please use the 'Solar and Wind Schedules' option instead.")
    manual_fill_schedule_combined()

# -------------------------------
# Configuration Editors (Contract Value, State, Tariff Difference, Buyer Mapping)
# -------------------------------

def update_hybrid_avc_cap():
    create_tables()

    st.subheader("Update Hybrid AVC Cap")
    st.write(
        "Maintain the date-wise hybrid validation limit used to round down the combined "
        "solar + wind schedule and AVC for hybrid schedule creation."
    )

    today = today_ist()
    tomorrow = today + datetime.timedelta(days=1)
    ensure_hybrid_avc_cap_defaults([today, tomorrow])

    conn = get_db_connection()
    existing_df = execute_query_to_dataframe(
        "SELECT date, avc_cap FROM hybrid_avc_cap ORDER BY date",
        conn
    )
    conn.close()

    if existing_df.empty:
        df = pd.DataFrame(
            {
                "date": [today, tomorrow],
                "avc_cap": [DEFAULT_HYBRID_AVC_CAP, DEFAULT_HYBRID_AVC_CAP],
            }
        )
    else:
        existing_df["date"] = pd.to_datetime(existing_df["date"]).dt.date
        df = existing_df.drop_duplicates(subset=["date"], keep="last").sort_values("date")

        missing_rows = []
        existing_dates = set(df["date"].tolist())
        for required_date in [today, tomorrow]:
            if required_date not in existing_dates:
                missing_rows.append({"date": required_date, "avc_cap": DEFAULT_HYBRID_AVC_CAP})

        if missing_rows:
            df = pd.concat([df, pd.DataFrame(missing_rows)], ignore_index=True)
            df = df.sort_values("date")

    edited_df = st.data_editor(
        df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "date": st.column_config.DateColumn("Date", format="YYYY-MM-DD"),
            "avc_cap": st.column_config.NumberColumn(
                "Hybrid Validation Limit / AVC Cap (MW)",
                min_value=0.0,
                step=0.1,
                format="%.1f",
            ),
        },
        key="hybrid_avc_cap_editor",
    )

    if st.button("Save Hybrid AVC Caps", type="primary"):
        if edited_df.empty:
            st.error("Please keep at least one hybrid AVC cap row.")
            return

        cleaned_df = edited_df.dropna(subset=["date"]).copy()
        if cleaned_df.empty:
            st.error("Please provide at least one valid date.")
            return

        cleaned_df["date"] = pd.to_datetime(cleaned_df["date"]).dt.date
        cleaned_df["avc_cap"] = cleaned_df["avc_cap"].apply(lambda value: max(to_float(value), 0.0))

        duplicates = cleaned_df.duplicated(subset=["date"], keep=False)
        if duplicates.any():
            cleaned_df = cleaned_df.drop_duplicates(subset=["date"], keep="last")
            st.warning("Duplicate dates were found. Kept the last entry for each date.")

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM hybrid_avc_cap")
        cursor.executemany(
            "REPLACE INTO hybrid_avc_cap (date, avc_cap) VALUES (%s, %s)",
            [(row["date"], float(row["avc_cap"])) for _, row in cleaned_df.iterrows()],
        )
        conn.commit()
        cursor.close()
        conn.close()

        st.success("Hybrid AVC cap values updated successfully.")
        st.rerun()

def update_contract_value():
    conn = get_db_connection()
    query_buyers = "SELECT DISTINCT buyer_name FROM load_data"
    buyers_df = execute_query_to_dataframe(query_buyers, conn)
    existing_df = execute_query_to_dataframe("SELECT * FROM contract_value", conn)
    merged = buyers_df.merge(existing_df, on="buyer_name", how="left")
    merged['contract_value'] = merged['contract_value'].fillna(0)

    st.write("Edit Contract Values")
    edited = st.data_editor(merged, num_rows="dynamic", use_container_width=True)

    cursor = conn.cursor()
    for _, row in edited.iterrows():
        buyer = row['buyer_name']
        # Convert numpy.float64 to Python float
        value = float(row['contract_value']) if row['contract_value'] is not None else 0.0
        query = "REPLACE INTO contract_value (buyer_name, contract_value) VALUES (%s, %s)"
        cursor.execute(query, (buyer, value))
    conn.commit()
    cursor.close()
    conn.close()
    st.success("Contract values updated.")


def update_buyer_access_type():
    create_config_tables()

    st.subheader("TGNA/ GNA buyers")
    st.write("Classify each buyer as `GNA` or `TGNA`. GNA buyers are allocated solar first and then wind. TGNA buyers are allocated wind first and then solar, and any solar given to TGNA is written into the wind contract column.")

    conn = get_db_connection()
    buyers_df = execute_query_to_dataframe(
        """
        SELECT buyer_name FROM contract_value
        UNION
        SELECT DISTINCT buyer_name FROM load_data
        """,
        conn,
    )
    existing_df = execute_query_to_dataframe(
        "SELECT buyer_name, access_type FROM buyer_access_type",
        conn,
    )
    conn.close()

    if buyers_df.empty and existing_df.empty:
        st.info("No buyers found yet. Add load data or contract values first, then return here.")
        return

    merged = buyers_df.merge(existing_df, on="buyer_name", how="outer")
    merged = merged.dropna(subset=["buyer_name"]).copy()
    merged["buyer_name"] = merged["buyer_name"].astype(str).str.strip()
    merged = merged[merged["buyer_name"] != ""]
    merged["access_type"] = merged["access_type"].apply(normalize_buyer_access_type)
    merged = merged.sort_values("buyer_name").reset_index(drop=True)

    edited_df = st.data_editor(
        merged,
        num_rows="dynamic",
        use_container_width=True,
        hide_index=True,
        column_config={
            "buyer_name": st.column_config.TextColumn("Buyer", required=True),
            "access_type": st.column_config.SelectboxColumn(
                "Access Type",
                options=["GNA", "TGNA"],
                required=True,
            ),
        },
        key="buyer_access_type_editor",
    )

    if st.button("Save Buyer Access Types", type="primary"):
        cleaned_df = edited_df.dropna(subset=["buyer_name"]).copy()
        cleaned_df["buyer_name"] = cleaned_df["buyer_name"].astype(str).str.strip()
        cleaned_df = cleaned_df[cleaned_df["buyer_name"] != ""]
        cleaned_df["access_type"] = cleaned_df["access_type"].apply(normalize_buyer_access_type)

        if cleaned_df.empty:
            st.error("Please keep at least one buyer row.")
            return

        cleaned_df = cleaned_df.drop_duplicates(subset=["buyer_name"], keep="last")

        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute("DELETE FROM buyer_access_type")
        cursor.executemany(
            "REPLACE INTO buyer_access_type (buyer_name, access_type) VALUES (%s, %s)",
            [
                (row["buyer_name"], row["access_type"])
                for _, row in cleaned_df.iterrows()
            ],
        )
        conn.commit()
        cursor.close()
        conn.close()

        st.success("Buyer access types updated successfully.")
        st.rerun()


def get_gdam_allocation_pct(conn=None):
    own_connection = conn is None
    if own_connection:
        conn = get_db_connection()

    try:
        config_df = execute_query_to_dataframe(
            "SELECT allocation_pct FROM gdam_allocation_config WHERE id = 1",
            conn,
        )
        if config_df.empty:
            return 0.0
        return round(max(to_float(config_df.iloc[0]["allocation_pct"]), 0.0), 0)
    finally:
        if own_connection and conn is not None:
            conn.close()


def update_gdam_allocation():
    create_config_tables()

    st.subheader("Edit G-DAM allocation")
    st.write(
        "Set the percentage of Day Ahead RTM to show in the derived "
        "`G-DAM_Solar_Output` and `G-DAM_Wind_Output` columns."
    )

    conn = get_db_connection()
    current_pct = get_gdam_allocation_pct(conn)
    conn.close()

    allocation_pct = st.number_input(
        "G-DAM allocation (%)",
        min_value=0.0,
        max_value=100.0,
        value=float(current_pct),
        step=1.0,
        format="%.0f",
        help="Example: 40 means 40% of Day Ahead RTM will be mirrored into the G-DAM output columns.",
    )

    if st.button("Save G-DAM allocation", type="primary"):
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute(
            """
            REPLACE INTO gdam_allocation_config (id, allocation_pct)
            VALUES (1, %s)
            """,
            (float(allocation_pct),),
        )
        conn.commit()
        cursor.close()
        conn.close()
        st.success("G-DAM allocation percentage updated successfully.")
        st.rerun()

def update_state():
    conn = get_db_connection()
    query_buyers = "SELECT DISTINCT buyer_name FROM load_data"
    buyers_df = execute_query_to_dataframe(query_buyers, conn)
    existing_df = execute_query_to_dataframe("SELECT * FROM state", conn)
    merged = buyers_df.merge(existing_df, on="buyer_name", how="left")
    merged['state'] = merged['state'].fillna("")

    st.write("Edit States")
    edited = st.data_editor(merged, num_rows="dynamic", use_container_width=True)

    cursor = conn.cursor()
    for _, row in edited.iterrows():
        buyer = row['buyer_name']
        state_val = row['state']
        query = "REPLACE INTO state (buyer_name, state) VALUES (%s, %s)"
        cursor.execute(query, (buyer, state_val))
    conn.commit()
    cursor.close()
    conn.close()
    st.success("States updated.")

def update_tariff_difference():
    conn = get_db_connection()

    # ——— 1) PASTE HELPER UI ———
    st.write("#### Quick Paste Helper")
    st.write("Paste one tariff-difference value per line, starting at the block you choose, for a single state.")

    col1, col2 = st.columns(2)
    with col1:
        paste_start_block = st.number_input(
            "Starting Block Number", min_value=1, max_value=96, value=1, step=1, key="td_start_block"
        )
    with col2:
        # fetch your states list dynamically
        states_df = execute_query_to_dataframe("SELECT DISTINCT state FROM state WHERE state <> ''", conn)
        state_options = [""] + sorted(states_df['state'].tolist())
        paste_state = st.selectbox("State", options=state_options, key="td_state")

    paste_values = st.text_area(
        "Paste Tariff-Difference Values (one per line)", height=120, key="td_values"
    )

    if st.button("Apply Tariff Differences", key="td_apply"):
        if not paste_state:
            st.warning("Please select a state.")
        elif not paste_values.strip():
            st.warning("Please paste at least one value.")
        else:
            # parse floats
            try:
                values = [float(v.strip()) for v in paste_values.splitlines() if v.strip()]
            except ValueError:
                st.error("All pasted values must be numeric.")
            else:
                # map to blocks
                blocks = list(range(paste_start_block, paste_start_block + len(values)))
                if max(blocks) > 96:
                    st.warning("Some blocks >96, trimming to 96.")
                    valid = list(range(paste_start_block, 97))
                    values = values[: len(valid)]
                    blocks = valid

                # upsert to tariff_difference
                cur = conn.cursor()
                for blk, td in zip(blocks, values):
                    cur.execute(
                        "SELECT 1 FROM tariff_difference WHERE block=%s AND state=%s",
                        (blk, paste_state),
                    )
                    if cur.fetchone():
                        cur.execute(
                            "UPDATE tariff_difference "
                            "SET tariff_difference=%s WHERE block=%s AND state=%s",
                            (td, blk, paste_state),
                        )
                    else:
                        cur.execute(
                            "INSERT INTO tariff_difference (block, state, tariff_difference) "
                            "VALUES (%s, %s, %s)",
                            (blk, paste_state, td),
                        )
                conn.commit()
                cur.close()
                st.success(f"Applied {len(values)} entries for '{paste_state}'.")

    # ——— 2) FULL GRID EDITOR ———
    # re-fetch distinct states (in case paste helper added new state entries)
    states_df = execute_query_to_dataframe("SELECT DISTINCT state FROM state WHERE state <> ''", conn)
    states = sorted(states_df['state'].tolist())

    # build base dataframe of blocks 1–96
    blocks = list(range(1, 97))
    tariff_df = pd.DataFrame({'Block': blocks})

    # pull existing tariff_difference
    existing = execute_query_to_dataframe("SELECT * FROM tariff_difference", conn)
    if not existing.empty:
        pivot = (
            existing
            .pivot(index='block', columns='state', values='tariff_difference')
            .reset_index()
            .rename(columns={'block': 'Block'})
        )
        tariff_df = tariff_df.merge(pivot, on='Block', how='left')

    # ensure every state column exists
    for s in states:
        if s not in tariff_df.columns:
            tariff_df[s] = 0.0
    tariff_df = tariff_df[['Block'] + states]

    st.write("#### Edit Tariff Differences (Full Table)")
    edited = st.data_editor(
        tariff_df,
        num_rows="dynamic",
        use_container_width=True,
        key="td_grid"
    )

    # melt back to long form and overwrite the table
    long_df = pd.melt(
        edited,
        id_vars=['Block'],
        var_name='state',
        value_name='tariff_difference'
    )

    cur = conn.cursor()
    cur.execute("DELETE FROM tariff_difference")
    for _, row in long_df.iterrows():
        blk = int(row['Block'])
        st_name = row['state']
        td_val = float(row['tariff_difference']) if pd.notna(row['tariff_difference']) else 0.0
        cur.execute(
            "INSERT INTO tariff_difference (block, state, tariff_difference) VALUES (%s, %s, %s)",
            (blk, st_name, td_val),
        )

    conn.commit()
    cur.close()
    conn.close()

    st.success("Tariff differences updated.")

def test_remc_template_mapping():
    """
    Test REMC template mapping by downloading template and showing buyer names and capacities.
    This helps verify that Row 16 (buyer names) and Row 19 (capacities) are correctly aligned.
    """
    st.subheader("🧪 Test REMC Template Mapping")
    st.write("Download REMC template from FTP and verify buyer name → capacity mapping")

    if st.button("🔍 Test REMC Template Mapping"):
        try:
            # Connect to FTP
            st.info("📡 Connecting to FTP...")
            ftp = connect_ftp()
            ftp.cwd(FTP_PATH_REMC_TEMPLATE)
            template_files = ftp.nlst()

            # Find ID template (or DA template if ID not found)
            template_file = pick_latest_template_file(ftp, template_files, is_intraday_template_filename)

            if not template_file:
                template_file = pick_latest_template_file(ftp, template_files, is_dayahead_template_filename)

            if not template_file:
                st.error("❌ No REMC template found on FTP")
                ftp.quit()
                return

            st.success(f"✅ Found template: {template_file}")

            # Download template
            bio = download_ftp_file(ftp, FTP_PATH_REMC_TEMPLATE, template_file)
            ftp.quit()

            if not bio:
                st.error("❌ Failed to download template")
                return

            # Parse template using same logic as process_remc_template
            bio.seek(0)
            all_lines = bio.read().decode('utf-8', errors='replace').split('\n')

            # Find max columns
            max_columns = 0
            for line in all_lines:
                if line.strip():
                    comma_count = line.count(',')
                    max_columns = max(max_columns, comma_count + 1)

            st.info(f"📊 Template has {len(all_lines)} lines, max {max_columns} columns")

            # Pad all lines
            padded_lines = []
            for line in all_lines:
                if not line.strip():
                    blank_line = ',' * (max_columns - 1)
                    padded_lines.append(blank_line)
                    continue

                current_columns = line.count(',') + 1
                if current_columns < max_columns:
                    padding_needed = max_columns - current_columns
                    line = line + ',' * padding_needed

                padded_lines.append(line)

            # Parse with pandas
            bio = BytesIO('\n'.join(padded_lines).encode('utf-8'))
            template_df = pd.read_csv(bio, header=None, keep_default_na=False, engine='python')

            st.success(f"✅ Parsed template: {len(template_df)} rows × {len(template_df.columns)} columns")

            # Show Row 16 (Buyer Names) and Row 19 (Capacities)
            st.write("---")
            st.write("### 📋 Row 16: Buyer Names (index 15)")
            if len(template_df) > 15:
                row_16 = template_df.iloc[15].tolist()
                st.write(row_16)

                # Create a table showing column index → buyer name
                buyer_data = []
                for col_idx, value in enumerate(row_16):
                    if value and str(value).strip():
                        buyer_data.append({
                            'Column Index': col_idx,
                            'Buyer Name in Template': str(value).strip()
                        })

                if buyer_data:
                    st.table(pd.DataFrame(buyer_data))
            else:
                st.warning("Template has fewer than 16 rows")

            st.write("---")
            st.write("### 📊 Row 19: Capacity Values (index 18)")
            if len(template_df) > 18:
                row_19 = template_df.iloc[18].tolist()
                st.write(row_19)

                # Create a table showing column index → capacity
                capacity_data = []
                for col_idx, value in enumerate(row_19):
                    if value and str(value).strip() and str(value).strip().lower() != 'capacity':
                        try:
                            capacity_val = float(value)
                            capacity_data.append({
                                'Column Index': col_idx,
                                'Capacity (MW)': capacity_val
                            })
                        except:
                            pass

                if capacity_data:
                    st.table(pd.DataFrame(capacity_data))
            else:
                st.warning("Template has fewer than 19 rows")

            # Show combined mapping
            st.write("---")
            st.write("### 🔗 Combined Mapping: Buyer Name → Capacity")

            if len(template_df) > 15 and len(template_df) > 18:
                mapping_data = []
                row_16 = template_df.iloc[15].tolist()
                row_19 = template_df.iloc[18].tolist()

                for col_idx in range(len(template_df.columns)):
                    buyer_name = row_16[col_idx] if col_idx < len(row_16) else ''
                    capacity = row_19[col_idx] if col_idx < len(row_19) else ''

                    # Only show columns with buyer names
                    if buyer_name and str(buyer_name).strip() and str(buyer_name).strip().lower() != 'buyer name':
                        try:
                            capacity_val = float(capacity) if capacity else 0
                            mapping_data.append({
                                'Column Index': col_idx,
                                'Buyer Name (Row 16)': str(buyer_name).strip(),
                                'Capacity MW (Row 19)': capacity_val
                            })
                        except:
                            mapping_data.append({
                                'Column Index': col_idx,
                                'Buyer Name (Row 16)': str(buyer_name).strip(),
                                'Capacity MW (Row 19)': str(capacity)
                            })

                if mapping_data:
                    mapping_df = pd.DataFrame(mapping_data)
                    st.table(mapping_df)

                    st.success(f"✅ Found {len(mapping_data)} buyer columns in REMC template")

                    # Compare with database buyer_mapping
                    st.write("---")
                    st.write("### 🔍 Compare with Database Buyer Mapping")
                    st.info("💡 **New Logic:** Buyers are matched using BOTH name AND capacity (allows multiple buyers with same name)")

                    conn = get_db_connection()
                    db_mapping = execute_query_to_dataframe("SELECT * FROM buyer_mapping", conn)
                    conn.close()

                    st.write("**Current buyer_mapping table:**")
                    st.dataframe(db_mapping)

                    # Check which template buyers (name + capacity) are NOT in database
                    missing_mappings = []
                    for m in mapping_data:
                        template_name = m['Buyer Name (Row 16)']
                        template_capacity = m['Capacity MW (Row 19)']

                        # Check if this name + capacity combination exists in database
                        found = False
                        for _, db_row in db_mapping.iterrows():
                            db_file_col_name = db_row.get('file_column_name', '')
                            db_capacity = db_row.get('capacity_mw', 0)

                            if str(template_name) == str(db_file_col_name) and abs(float(template_capacity) - float(db_capacity)) < 0.01:
                                found = True
                                break

                        if not found:
                            missing_mappings.append(f"{template_name} ({template_capacity} MW)")

                    if missing_mappings:
                        st.warning(f"⚠️ Found {len(missing_mappings)} buyer(s) in template that are NOT in database:")
                        for buyer in missing_mappings:
                            st.write(f"- **{buyer}**")
                        st.info("💡 You need to add these to the buyer_mapping table in 'Edit Final Revision Integration'")
                        st.info("**Remember:** Use the exact REMC template name from Row 16 and capacity from Row 19")
                    else:
                        st.success("✅ All template buyers (name + capacity) are mapped in database!")
                else:
                    st.warning("No buyer columns found in template")

        except Exception as e:
            st.error(f"❌ Error testing template mapping: {str(e)}")
            import traceback
            st.code(traceback.format_exc())


def update_buyer_mapping():
    st.subheader("Edit Final Revision Integration")
    st.write("Map column names in the final buyer schedule files to actual buyer names")

    conn = get_db_connection()

    # Get the list of buyer names from load_data table
    query_buyers = "SELECT DISTINCT buyer_name FROM load_data"
    buyers_df = execute_query_to_dataframe(query_buyers, conn)
    buyer_names = buyers_df['buyer_name'].tolist()

    # Get existing mappings
    mapping_df = execute_query_to_dataframe("SELECT * FROM buyer_mapping", conn)

    # Add a form to add a new mapping
    st.write("### Add New Mapping")
    st.info("💡 **Important:** File Column Name is the buyer name from Row 16 of the REMC template (Final Revision CSV). Multiple buyers can have the same name (e.g., LINDEINDIA), so use name + capacity to uniquely identify them.")

    col1, col2 = st.columns(2)
    with col1:
        new_file_column = st.text_input("File Column Name", key="new_file_column", help="The buyer name from Row 16 of REMC template (e.g., 'LINDEINDIA')")
        new_capacity = st.number_input("Capacity (MW)", min_value=0.0, step=0.1, key="new_capacity", help="Contract capacity from Row 19 of REMC template (e.g., 10)")
    with col2:
        new_buyer_name = st.selectbox("Actual Buyer Name", options=[""] + buyer_names, key="new_buyer_name", help="The buyer name used in the system")
        new_description = st.text_input("Description", key="new_description", help="Optional description (e.g., 'Linde Rourkela 2 - 10 MW contract')")

    if st.button("Add Mapping"):
        if new_file_column and new_buyer_name and new_capacity > 0:
            cursor = conn.cursor()
            try:
                query = """
                INSERT INTO buyer_mapping (file_column_name, actual_buyer_name, capacity_mw, description)
                VALUES (%s, %s, %s, %s)
                """
                cursor.execute(query, (new_file_column, new_buyer_name, new_capacity, new_description))
                conn.commit()
                st.success(f"✅ Added mapping: '{new_file_column}' ({new_capacity} MW) → '{new_buyer_name}'")
                # Refresh the page to show the new mapping
                st.rerun()
            except Exception as e:
                st.error(f"❌ Error adding mapping: {str(e)}")
            finally:
                cursor.close()
        else:
            if not new_file_column:
                st.warning("Please enter a File Column Name")
            elif not new_buyer_name:
                st.warning("Please select an Actual Buyer Name")
            elif new_capacity <= 0:
                st.warning("Please enter a valid Capacity (MW)")

    st.write("---")
    st.write("### Current Buyer Mappings")
    st.info("💡 **Unique Key:** File Column Name + Capacity MW (allows multiple buyers with same name)")
    st.info("📝 **Note:** File Column Name is the buyer name from Row 16 of REMC template (Final Revision CSV)")

    edited = st.data_editor(
        mapping_df,
        num_rows="dynamic",
        use_container_width=True,
        column_config={
            "id": st.column_config.NumberColumn("ID", disabled=True),
            "file_column_name": st.column_config.TextColumn("File Column Name", help="Buyer name from Row 16 of REMC template (e.g., 'LINDEINDIA')"),
            "actual_buyer_name": st.column_config.SelectboxColumn(
                "Actual Buyer Name",
                help="The buyer name used in the system",
                options=buyer_names,
                required=True
            ),
            "capacity_mw": st.column_config.NumberColumn("Capacity (MW)", help="Contract capacity from Row 19 of REMC template", format="%.2f"),
            "description": st.column_config.TextColumn("Description", help="Optional description")
        }
    )

    # Handle updates and inserts
    cursor = conn.cursor()

    # First, get existing IDs
    existing_ids = set(mapping_df['id'].tolist())

    for _, row in edited.iterrows():
        if pd.isna(row['id']) or row['id'] == 0:  # New row
            if not pd.isna(row['file_column_name']) and not pd.isna(row['actual_buyer_name']):
                query = """
                INSERT INTO buyer_mapping (file_column_name, actual_buyer_name, capacity_mw, description)
                VALUES (%s, %s, %s, %s)
                """
                description = row['description'] if not pd.isna(row['description']) else ""
                capacity = float(row['capacity_mw']) if not pd.isna(row['capacity_mw']) else 0.0
                cursor.execute(query, (row['file_column_name'], row['actual_buyer_name'], capacity, description))
        else:  # Existing row
            existing_ids.discard(row['id'])  # Remove from set of existing IDs
            if not pd.isna(row['file_column_name']) and not pd.isna(row['actual_buyer_name']):
                query = """
                UPDATE buyer_mapping
                SET file_column_name = %s, actual_buyer_name = %s, capacity_mw = %s, description = %s
                WHERE id = %s
                """
                description = row['description'] if not pd.isna(row['description']) else ""
                capacity = float(row['capacity_mw']) if not pd.isna(row['capacity_mw']) else 0.0
                cursor.execute(query, (row['file_column_name'], row['actual_buyer_name'], capacity, description, row['id']))

    # Delete rows that were removed in the UI
    for id_to_delete in existing_ids:
        query = "DELETE FROM buyer_mapping WHERE id = %s"
        cursor.execute(query, (id_to_delete,))

    conn.commit()
    cursor.close()
    conn.close()
    st.success("Buyer mappings updated.")

# -------------------------------
# REMC Template Functions
# -------------------------------

def create_remc_template_files(bifurcation_results, id_date, da_date, buyer_list, conn, logger):
    """
    Create REMC template files (ID and DA) based on bifurcation results.
    Downloads template from FTP, populates with data, and returns file paths.

    Args:
        bifurcation_results: DataFrame with bifurcation results
        id_date: Intraday date
        da_date: Day-Ahead date
        buyer_list: List of buyer names
        conn: Database connection
        logger: Logger instance for logging

    Returns:
        tuple: (id_file_path, id_revision, da_file_path, da_revision) or (None, None, None, None) if failed
    """
    try:
        # Download latest templates from portal to FTP before processing
        # This ensures we are using the most up-to-date format with correct contract IDs
        # NOTE: This downloads the TEMPLATE STRUCTURE (headers, contract IDs, etc.)
        # We will then OVERWRITE the schedule values with our bifurcated allocation
        download_remc_templates_from_portal(logger)

        # Get buyer mapping from database
        # file_column_name = buyer name from Row 16 of REMC template (e.g., "LINDEINDIA")
        # capacity_mw = capacity from Row 19 of REMC template (e.g., 10)
        buyer_mapping_df = execute_query_to_dataframe(
            "SELECT file_column_name, actual_buyer_name, capacity_mw FROM buyer_mapping", conn
        )

        # Create mapping: actual_buyer_name -> (file_column_name, capacity_mw)
        # This allows us to match buyers by BOTH name AND capacity in REMC template
        buyer_to_remc_mapping = {}
        for _, row in buyer_mapping_df.iterrows():
            buyer_to_remc_mapping[row['actual_buyer_name']] = {
                'file_column_name': row['file_column_name'],
                'capacity_mw': float(row['capacity_mw']) if pd.notna(row['capacity_mw']) else 0.0
            }

        logger.info(f"Loaded {len(buyer_to_remc_mapping)} buyer mappings with names and capacities")

        # Connect to FTP and find template files
        ftp = connect_ftp()
        ftp.cwd(FTP_PATH_REMC_TEMPLATE)
        template_files = ftp.nlst()

        # Find ID and DA template files
        id_template = pick_latest_template_file(ftp, template_files, is_intraday_template_filename, logger)
        da_template = pick_latest_template_file(ftp, template_files, is_dayahead_template_filename, logger)

        if not id_template or not da_template:
            st.error(
                f"❌ Template files not found in {FTP_PATH_REMC_TEMPLATE}. "
                f"Matched ID={id_template}, DA={da_template}"
            )
            ftp.quit()
            return None, None, None, None

        logger.info(f"Found templates: ID={id_template}, DA={da_template}")

        # Download and process templates
        id_file_path, id_revision = process_remc_template(
            ftp, id_template, bifurcation_results, id_date,
            buyer_list, buyer_to_remc_mapping, 'ID', logger
        )

        da_file_path, da_revision = process_remc_template(
            ftp, da_template, bifurcation_results, da_date,
            buyer_list, buyer_to_remc_mapping, 'DA', logger
        )

        ftp.quit()
        return id_file_path, id_revision, da_file_path, da_revision

    except Exception as e:
        st.error(f"❌ Error creating REMC template files: {str(e)}")
        logger.error(f"REMC template creation failed: {str(e)}")
        return None, None, None, None


def process_remc_template(ftp, template_filename, bifurcation_results, target_date,
                          buyer_list, buyer_to_remc_mapping, file_type, logger):
    """
    Process a single REMC template file (ID or DA).

    Uses BOTH buyer name AND capacity to match buyers in REMC template.
    This allows multiple buyers to have the same REMC template name (e.g., LINDEINDIA).

    Args:
        ftp: FTP connection
        template_filename: Name of template file on FTP
        bifurcation_results: DataFrame with bifurcation results
        target_date: Date for this file
        buyer_list: List of buyer names
        buyer_to_remc_mapping: Dict mapping buyer_name -> {remc_template_name, capacity_mw, file_column_name}
        file_type: 'ID' or 'DA'
        logger: Logger instance for logging

    Returns:
        tuple: (file_path, revision_number) or (None, None) if failed
    """
    try:
        # Get revision number from session state (stored when loading schedule files)
        if file_type == 'DA':
            revision_number = st.session_state.get('da_revision', '1')
            logger.info(f"Retrieved DA revision from session state: {revision_number}")
        else:
            revision_number = st.session_state.get('id_revision', '1')
            logger.info(f"Retrieved ID revision from session state: {revision_number}")

        logger.info(f"Using revision number: {revision_number} for {file_type} template")

        # Download template
        bio = download_ftp_file(ftp, FTP_PATH_REMC_TEMPLATE, template_filename)
        if not bio:
            st.error(f"❌ Failed to download template: {template_filename}")
            return None, None

        # Check file size
        file_size = bio.getbuffer().nbytes
        logger.info(f"Downloaded template file size: {file_size} bytes")

        if file_size == 0:
            st.error(f"❌ Template file is empty: {template_filename}")
            return None, None

        # Read first few bytes to check if it's actually a CSV
        bio.seek(0)
        first_bytes = bio.read(100)
        bio.seek(0)

        logger.info(f"First 100 bytes of template: {first_bytes[:100]}")

        # Check if file looks like HTML (common issue when downloading from web portals)
        if b'<html' in first_bytes.lower() or b'<!doctype' in first_bytes.lower():
            st.error(f"❌ Template file appears to be HTML, not CSV: {template_filename}")
            st.error("This usually means the file on FTP is a download page, not the actual template.")
            st.info("**Solution:** Download the actual CSV template from REMC portal and upload to FTP.")
            logger.error(f"Template file is HTML, not CSV. First bytes: {first_bytes}")
            return None, None

        # Try to read template CSV without treating first row as header
        # This preserves the original template structure including the title row
        template_df = None
        parsing_error = None

        # REMC templates have INCONSISTENT column counts per row
        # First line has 1 field, but other lines have 3, 7, etc.
        # Pandas will skip lines that don't match the first line's column count
        # Solution: Read all lines, find max columns, pad all lines to same length

        bio.seek(0)
        all_lines = bio.read().decode('utf-8', errors='replace').split('\n')

        # Find maximum number of columns by counting commas in each line
        max_columns = 0
        for line in all_lines:
            if line.strip():
                comma_count = line.count(',')
                max_columns = max(max_columns, comma_count + 1)

        logger.info(f"Template has maximum {max_columns} columns across all rows")
        logger.info(f"First line: {all_lines[0][:200]}")
        logger.info(f"Total lines: {len(all_lines)}")
        st.info(f"📊 Template structure: {len(all_lines)} lines, max {max_columns} columns")

        # PAD all lines to have the same number of columns
        # This is critical - pandas requires consistent column counts
        # IMPORTANT: Blank lines must also be padded to preserve row structure
        logger.info(f"Padding all lines to {max_columns} columns...")
        padded_lines = []
        for line in all_lines:
            if not line.strip():
                # Blank lines need to be padded too! Otherwise pandas drops them
                # Create a blank line with correct number of columns
                blank_line = ',' * (max_columns - 1)  # e.g., for 7 columns: ",,,,,"
                padded_lines.append(blank_line)
                continue

            # Count current columns
            current_columns = line.count(',') + 1

            # Add commas to pad to max_columns
            if current_columns < max_columns:
                padding_needed = max_columns - current_columns
                line = line + ',' * padding_needed

            padded_lines.append(line)

        logger.info(f"Padded {len(padded_lines)} lines to {max_columns} columns (including blank lines)")

        # Convert back to BytesIO for pandas
        bio = BytesIO('\n'.join(padded_lines).encode('utf-8'))

        # Now parse with pandas - should work since all lines have same column count
        # IMPORTANT: Use Python engine, not C engine!
        # C engine is strict and fails on inconsistent columns even after padding
        # Python engine is more lenient and handles padded lines correctly
        template_df = None
        parsing_error = None

        try:
            bio.seek(0)
            template_df = pd.read_csv(bio, header=None, keep_default_na=False, engine='python')
            logger.info(f"✅ Successfully parsed template! Rows: {len(template_df)}, Columns: {len(template_df.columns)}")
            st.info(f"✅ Successfully parsed template")
        except Exception as e:
            parsing_error = e
            logger.error(f"Failed to parse template even after padding: {str(e)}")

        # Fallback: try with different strategies if padding didn't work
        if template_df is None:
            logger.warning("Padding strategy failed, trying alternative parsing methods...")
            parsing_strategies = [
                {'header': None, 'engine': 'python', 'keep_default_na': False},
                {'header': None, 'delimiter': ',', 'keep_default_na': False},
            ]

            for i, strategy in enumerate(parsing_strategies, 1):
                try:
                    bio.seek(0)
                    logger.info(f"Trying parsing strategy {i}: {strategy}")
                    template_df = pd.read_csv(bio, **strategy)
                    logger.info(f"✅ Strategy {i} succeeded! Template structure: {len(template_df)} rows x {len(template_df.columns)} columns")
                    st.info(f"✅ Successfully parsed template using fallback strategy {i}")
                    break
                except Exception as e:
                    logger.warning(f"Strategy {i} failed: {str(e)}")
                    parsing_error = e
                    continue

        # Show template structure if successfully parsed
        if template_df is not None:
            template_df = template_df.astype(object)
            logger.info(f"First 5 rows of parsed template:\n{template_df.head()}")
            st.info(f"📊 Template: {len(template_df)} rows × {len(template_df.columns)} columns")

            # Show sample of key rows for verification
            with st.expander("🔍 Template Structure Preview"):
                if len(template_df) > 2:
                    st.write(f"**Row 3 (Date field, index 2):** {template_df.iloc[2].tolist()}")
                if len(template_df) > 5:
                    st.write(f"**Row 6 (Blank row, index 5):** {template_df.iloc[5].tolist()}")
                if len(template_df) > 15:
                    st.write(f"**Row 16 (Buyer names, index 15):** {template_df.iloc[15].tolist()}")
                if len(template_df) > 18:
                    st.write(f"**Row 19 (Capacity row, index 18):** {template_df.iloc[18].tolist()}")
                if len(template_df) > 19:
                    st.write(f"**Row 20 (Column headers, index 19):** {template_df.iloc[19].tolist()}")
                if len(template_df) > 20:
                    st.write(f"**Row 21 (First data row, index 20):** {template_df.iloc[20].tolist()}")

        if template_df is None:
            st.error(f"❌ Error reading template CSV: {str(parsing_error)}")
            st.error(f"Template file: {template_filename}")
            st.error(f"File size: {file_size} bytes")
            st.info("**Tried multiple parsing strategies, all failed.**")
            st.info("**Possible causes:**")
            st.info("1. Template file has inconsistent number of columns per row")
            st.info("2. Template file has special characters or quotes")
            st.info("3. Template file has mixed delimiters (commas and tabs)")
            st.info("**Solution:** Download fresh template from REMC portal and upload to FTP")
            logger.error(f"CSV parsing error after all strategies: {parsing_error}")
            logger.error(f"First 500 bytes: {first_bytes}")

            # Show the actual content for debugging
            bio.seek(0)
            content = bio.read(1000).decode('utf-8', errors='replace')
            st.text_area("First 1000 characters of file:", content, height=200)

            return None, None

        buyer_row_idx = find_template_row_index(template_df, "Buyer Name")
        capacity_row_idx = find_template_row_index(template_df, "Capacity")
        energy_type_row_idx = find_template_row_index(template_df, "Energy Type")
        header_row_idx = find_template_row_index(template_df, "Block")
        date_row_idx = find_template_row_index(template_df, "Date")

        # Determine date to use based on file type
        # ID template: current date, DA template: tomorrow's date (using IST)
        if file_type == 'DA':
            date_to_use = target_date  # DA uses tomorrow's date
        else:
            import pytz
            ist = pytz.timezone('Asia/Kolkata')
            date_to_use = datetime.datetime.now(ist)  # ID uses current date in IST

        existing_date_value = ""
        if date_row_idx is not None and len(template_df.columns) > 2:
            existing_date_value = template_df.iloc[date_row_idx, 2]
        elif len(template_df.columns) > 2 and len(template_df) > 2:
            existing_date_value = template_df.iloc[2, 2]

        if "-" in str(existing_date_value):
            date_str = date_to_use.strftime("%Y-%m-%d")
        else:
            date_str = date_to_use.strftime("%d/%m/%y")

        if not set_template_value_next_to_label(template_df, "Date", date_str, logger):
            if len(template_df.columns) > 2 and len(template_df) > 2:
                template_df.iloc[2, 2] = format_template_cell_value(date_str)
                logger.info(f"Updated fallback template date cell with value: {date_str}")
            else:
                logger.warning(
                    f"Template has only {len(template_df)} rows and {len(template_df.columns)} columns, cannot update date"
                )

        # Find buyer columns in row 16 (index 15) - buyer names row
        # Row 16 has: ['Buyer Name', '', '', 'LINDE INDIA LIMITED A.P.', 'LINDEGJ62955', 'LINDEINDIA12', 'LINDEINDIA']
        # Row 19 has: ['Capacity', '66', '66', '8', '5', '8', '17', '10']
        #
        # IMPORTANT: Multiple buyers can have the same name in Row 16 (e.g., LINDEINDIA)
        # We use BOTH buyer name (Row 16) AND capacity (Row 19) to uniquely identify them

        buyer_col_mappings = {}
        if buyer_row_idx is not None and capacity_row_idx is not None:
            # Match buyer columns using BOTH name (Row 16) AND capacity (Row 19)
            for col_idx in range(len(template_df.columns)):
                # Get buyer name from Row 16
                buyer_name_in_template = template_df.iloc[buyer_row_idx, col_idx]
                if not pd.notna(buyer_name_in_template) or not buyer_name_in_template:
                    continue

                buyer_name_str = str(buyer_name_in_template).strip()
                if buyer_name_str.lower() == 'buyer name':
                    continue  # Skip header

                # Get capacity from Row 19
                capacity_in_template = template_df.iloc[capacity_row_idx, col_idx]
                if not pd.notna(capacity_in_template) or not capacity_in_template:
                    continue

                try:
                    capacity_val = float(capacity_in_template)
                except:
                    continue  # Skip if capacity is not a number

                energy_type_in_template = ""
                if energy_type_row_idx is not None:
                    energy_type_in_template = str(
                        template_df.iloc[energy_type_row_idx, col_idx]
                    ).strip().upper()

                # Match with buyer_to_remc_mapping using BOTH name AND capacity
                for buyer_name, mapping_info in buyer_to_remc_mapping.items():
                    file_col_name = mapping_info.get('file_column_name', '')
                    mapped_capacity = mapping_info.get('capacity_mw', 0.0)

                    # Match if BOTH name AND capacity match
                    if buyer_name_str == str(file_col_name).strip() and abs(capacity_val - mapped_capacity) < 0.01:
                        buyer_col_mappings.setdefault(buyer_name, []).append(
                            {
                                "col_idx": col_idx,
                                "energy_type": energy_type_in_template,
                            }
                        )
                        logger.info(
                            f"✅ Matched column {col_idx}: '{buyer_name_str}' "
                            f"({capacity_val} MW, energy_type={energy_type_in_template or 'UNKNOWN'}) "
                            f"→ buyer '{buyer_name}'"
                        )
                        st.info(
                            f"✅ Matched column {col_idx}: '{buyer_name_str}' "
                            f"({capacity_val} MW, energy_type={energy_type_in_template or 'UNKNOWN'}) "
                            f"→ buyer '{buyer_name}'"
                        )
                        break

        total_buyer_columns = sum(len(column_list) for column_list in buyer_col_mappings.values())
        logger.info(f"Found {total_buyer_columns} buyer columns in template across {len(buyer_col_mappings)} buyers")

        if len(buyer_col_mappings) == 0:
            logger.warning("No buyer columns matched! Check buyer_mapping table.")
            logger.warning(f"Buyer row values: {template_df.iloc[buyer_row_idx].tolist() if buyer_row_idx is not None else 'N/A'}")
            logger.warning(f"Capacity row values: {template_df.iloc[capacity_row_idx].tolist() if capacity_row_idx is not None else 'N/A'}")
            logger.warning(f"Expected mappings: {[(k, v['file_column_name'], v['capacity_mw']) for k, v in buyer_to_remc_mapping.items()]}")

        declared_forecast_col = None
        declared_plant_schedule_col = None
        plant_avc_col = None
        declared_hybrid_schedule_col = None
        hybrid_avc_col = None
        total_hybrid_solar_col = None
        total_hybrid_wind_col = None
        avc_columns = []

        if header_row_idx is not None:
            for col_idx in range(len(template_df.columns)):
                header_value = template_df.iloc[header_row_idx, col_idx]
                header_token = normalize_template_token(header_value)
                if not header_token:
                    continue

                if header_token == "declaredforecast":
                    declared_forecast_col = col_idx
                elif header_token == "declaredplantschedule":
                    declared_plant_schedule_col = col_idx
                elif header_token == "plantavc":
                    plant_avc_col = col_idx
                elif header_token == "declaredhybridschedule":
                    declared_hybrid_schedule_col = col_idx
                elif header_token == "hybridavc":
                    hybrid_avc_col = col_idx
                elif header_token == "totalhybridsolar":
                    total_hybrid_solar_col = col_idx
                elif header_token == "totalhybridwind":
                    total_hybrid_wind_col = col_idx
        avc_columns = classify_template_avc_columns(
            template_df,
            header_row_idx,
            energy_type_row_idx,
            capacity_row_idx,
        )

        if header_row_idx is None:
            logger.warning("Could not find template header row using 'Block'")
        elif declared_plant_schedule_col is None and declared_forecast_col is None:
            logger.warning("No declared schedule column found in template header row")
        if not avc_columns:
            logger.warning("Generic AVC column not found in template header row")
        else:
            logger.info(
                "Detected AVC columns: %s",
                [
                    (
                        avc_column.get("col_idx"),
                        avc_column.get("energy_type") or "GENERIC",
                        avc_column.get("capacity_value"),
                    )
                    for avc_column in avc_columns
                ],
            )

        # Get AVC capacity value from the first AVC column in the capacity row.
        avc_capacity = None
        primary_avc_col = avc_columns[0]["col_idx"] if avc_columns else None
        if primary_avc_col is not None and capacity_row_idx is not None:
            avc_capacity_value = template_df.iloc[capacity_row_idx, primary_avc_col]
            if pd.notna(avc_capacity_value) and avc_capacity_value:
                try:
                    avc_capacity = float(avc_capacity_value)
                    logger.info(f"AVC capacity value from capacity row: {avc_capacity}")
                except:
                    logger.warning(f"Could not convert AVC capacity to number: {avc_capacity_value}")

        if avc_capacity is None:
            logger.warning("AVC capacity value not found in capacity row")

        # Filter bifurcation results for this date
        bifurcation_results['Date_only'] = pd.to_datetime(bifurcation_results['Date']).dt.date
        date_results = bifurcation_results[bifurcation_results['Date_only'] == target_date]

        if date_results.empty:
            st.warning(f"⚠️ No bifurcation results for {target_date} ({file_type})")
            return None

        validation_limit_value = round(
            to_float(date_results["Hybrid_AvC_Cap"].dropna().iloc[0], DEFAULT_HYBRID_AVC_CAP),
            1,
        )
        set_template_value_next_to_label(
            template_df,
            "Hybrid Validation Limit",
            validation_limit_value,
            logger,
        )

        # Sort by block to ensure correct order
        date_results = date_results.sort_values('Block')

        start_row = (header_row_idx + 1) if header_row_idx is not None else 20

        # Populate buyer schedules and declared forecast
        for idx, row in date_results.iterrows():
            block = int(row['Block'])
            row_idx = start_row + block - 1  # Block 1 goes to row 21 (index 20)

            if row_idx >= len(template_df):
                # Extend dataframe if needed
                while len(template_df) <= row_idx:
                    template_df.loc[len(template_df)] = [''] * len(template_df.columns)

            # Fill buyer schedule columns (sum of Solar + Wind) and calculate total
            total_buyer_schedules = 0
            for buyer_name, buyer_columns in buyer_col_mappings.items():
                solar_col = f"{buyer_name}_Solar"
                wind_col = f"{buyer_name}_Wind"

                solar_val = row.get(solar_col, 0) if solar_col in row else 0
                wind_val = row.get(wind_col, 0) if wind_col in row else 0

                for buyer_column in buyer_columns:
                    col_idx = buyer_column["col_idx"]
                    energy_type = buyer_column.get("energy_type", "")

                    if energy_type == "SOLAR":
                        schedule_value = solar_val
                    elif energy_type == "WIND":
                        schedule_value = wind_val
                    else:
                        schedule_value = solar_val + wind_val

                    template_df.iloc[row_idx, col_idx] = format_template_cell_value(schedule_value)
                    total_buyer_schedules += schedule_value

                    # Log Block 1 values for verification
                    if block == 1:
                        logger.info(
                            f"Block 1 - {buyer_name} [{energy_type or 'TOTAL'}]: "
                            f"Solar={solar_val}, Wind={wind_val}, Written={schedule_value}"
                        )

            declared_plant_schedule = round(
                to_float(row.get("Declared_Plant_Schedule", row.get("Hyb_Sch", total_buyer_schedules))),
                1,
            )
            plant_avc_value = round(
                to_float(row.get("Plant_AvC", row.get("Hybrid_AvC_Input", 0))),
                1,
            )
            declared_hybrid_schedule = round(to_float(row.get("Hyb_Sch", total_buyer_schedules)), 1)
            hybrid_avc_value = round(
                to_float(row.get("Hybrid_AvC", row.get("Hybrid_AvC_Input", avc_capacity or 0))),
                1,
            )
            wind_avc_value = round(
                to_float(
                    row.get(
                        "Capped_Wind_AvC",
                        row.get("Wind_AvC", row.get("avc_wind", 0)),
                    )
                ),
                1,
            )
            solar_avc_value = round(
                to_float(
                    row.get(
                        "Capped_Solar_AvC",
                        row.get("Solar_AvC", row.get("avc_solar", 0)),
                    )
                ),
                1,
            )
            total_hybrid_solar = round(to_float(row.get("Total_Hybrid_Solar", row.get("Solar_Sch", 0))), 1)
            total_hybrid_wind = round(to_float(row.get("Total_Hybrid_Wind", row.get("Wind_Sch", 0))), 1)

            if declared_plant_schedule_col is not None:
                template_df.iloc[row_idx, declared_plant_schedule_col] = format_template_cell_value(declared_plant_schedule)
            if plant_avc_col is not None:
                template_df.iloc[row_idx, plant_avc_col] = format_template_cell_value(plant_avc_value)
            if declared_hybrid_schedule_col is not None:
                template_df.iloc[row_idx, declared_hybrid_schedule_col] = format_template_cell_value(declared_hybrid_schedule)
            if hybrid_avc_col is not None:
                template_df.iloc[row_idx, hybrid_avc_col] = format_template_cell_value(hybrid_avc_value)
            if total_hybrid_solar_col is not None:
                template_df.iloc[row_idx, total_hybrid_solar_col] = format_template_cell_value(total_hybrid_solar)
            if total_hybrid_wind_col is not None:
                template_df.iloc[row_idx, total_hybrid_wind_col] = format_template_cell_value(total_hybrid_wind)

            for avc_column in avc_columns:
                avc_col_idx = avc_column["col_idx"]
                avc_energy_type = avc_column.get("energy_type", "")
                avc_capacity_value = to_float(avc_column.get("capacity_value", ""), 0.0)

                if avc_energy_type == "WIND":
                    avc_value_to_write = wind_avc_value if wind_avc_value > 0 else avc_capacity_value
                elif avc_energy_type == "SOLAR":
                    avc_value_to_write = solar_avc_value if solar_avc_value > 0 else avc_capacity_value
                elif len(avc_columns) == 1:
                    avc_value_to_write = hybrid_avc_value
                else:
                    avc_value_to_write = ""

                template_df.iloc[row_idx, avc_col_idx] = format_template_cell_value(avc_value_to_write)

            # Legacy template support: fill Declared Forecast with buyer schedules total.
            if declared_forecast_col is not None:
                template_df.iloc[row_idx, declared_forecast_col] = format_template_cell_value(total_buyer_schedules)

        # Generate output filename
        output_filename = f"HIRIYUR_ZREPL_W_{file_type}_{target_date.strftime('%d-%m-%Y')}.csv"
        
        # Create a unique temporary directory for this specific file generation
        # This prevents race conditions where multiple runs overwrite the same file in the shared temp folder
        unique_temp_dir = tempfile.mkdtemp(prefix=f"remc_{file_type}_")
        output_path = os.path.join(unique_temp_dir, output_filename)

        # Save to file without header (template already has its own structure)
        template_df.to_csv(output_path, index=False, header=False)

        # Verify saved file by reading Block 1 values
        try:
            verify_df = pd.read_csv(output_path, header=None)
            verify_row_idx = start_row if len(verify_df) > start_row else None
            if verify_row_idx is not None:
                block1_values = verify_df.iloc[verify_row_idx, 1:].tolist()
                logger.info(f"✅ Saved file Block 1 values: {block1_values[:5]}")
                st.info(f"✅ Saved file Block 1 values: {block1_values[:5]}")
        except Exception as e:
            logger.warning(f"Could not verify saved file: {e}")

        st.success(f"✅ Created {file_type} template: {output_filename}")
        logger.info(f"Created REMC template file: {output_path}")

        return output_path, revision_number

    except Exception as e:
        st.error(f"❌ Error processing {file_type} template: {str(e)}")
        logger.error(f"REMC template processing failed for {file_type}: {str(e)}")
        return None, None


# -------------------------------
# Bifurcation Logic & CUF% Calculation
# -------------------------------

def compute_cuf(contract_val, current_date):
    days_in_month = calendar.monthrange(current_date.year, current_date.month)[1]
    return contract_val * 96 * days_in_month

def allocate_high_priority(remaining_energy, buyers, req_remaining):
    """
    Allocates energy proportionally to requisitions in one pass while ensuring
    no buyer receives more than its requisition. All allocations are rounded
    to 1 decimal place.
    If any energy remains (because some buyers hit their maximum), the remaining
    energy is redistributed proportionally among those who haven't reached their limit.
    In case a rounding discrepancy of exactly ±0.1 remains, a final adjustment is made.
    For a -0.1 remainder, the adjustment subtracts 0.1 from the buyer with the highest requisition.

    The function ensures that the total allocation never exceeds the available energy.
    """
    # Round the remaining energy to 1 decimal place to ensure consistency
    remaining_energy = round(remaining_energy, 1)

    # Calculate total requested energy.
    total_req = sum(req_remaining[buyer] for buyer in buyers)
    if total_req <= 1e-9:
        return {buyer: 0.0 for buyer in buyers}, remaining_energy

    # First pass: compute each buyer's ideal allocation, capped by their requisition.
    allocations = {}
    for buyer in buyers:
        ideal_allocation = remaining_energy * (req_remaining[buyer] / total_req)
        allocations[buyer] = round(min(ideal_allocation, req_remaining[buyer]), 1)

    allocated_total = round(sum(allocations.values()), 1)
    remainder = round(remaining_energy - allocated_total, 1)

    # Redistribute any remaining energy proportionally among buyers
    # who have not yet reached their requisition limit.
    while remainder > 0:
        capacity_buyers = [buyer for buyer in buyers if allocations[buyer] < req_remaining[buyer]]
        if not capacity_buyers:
            break

        extra_capacity = sum(req_remaining[buyer] - allocations[buyer] for buyer in capacity_buyers)
        if extra_capacity <= 1e-9:
            break

        distributed = 0.0
        for buyer in capacity_buyers:
            cap = req_remaining[buyer] - allocations[buyer]
            additional = remainder * (cap / extra_capacity)
            additional_rounded = round(additional, 1)
            additional_rounded = min(additional_rounded, cap)
            allocations[buyer] = round(allocations[buyer] + additional_rounded, 1)
            distributed += additional_rounded

        distributed = round(distributed, 1)
        new_remainder = round(remainder - distributed, 1)
        if new_remainder == remainder:
            break
        remainder = new_remainder

    # Final adjustment for small rounding errors:
    if remainder == 0.1:
        # Allocate the extra 0.1 to the first buyer with at least 0.1 available capacity.
        for buyer in buyers:
            available = round(req_remaining[buyer] - allocations[buyer], 1)
            if available >= 0.1:
                allocations[buyer] = round(allocations[buyer] + 0.1, 1)
                remainder = 0.0
                break
    elif remainder == -0.1:
        # Subtract 0.1 from the buyer with the highest requisition.
        buyer_to_reduce = max(buyers, key=lambda b: req_remaining[b])
        if allocations[buyer_to_reduce] >= 0.1:
            allocations[buyer_to_reduce] = round(allocations[buyer_to_reduce] - 0.1, 1)
            remainder = 0.0
        else:
            # Fallback: reduce from the first buyer that has at least 0.1 allocated.
            for buyer in buyers:
                if allocations[buyer] >= 0.1:
                    allocations[buyer] = round(allocations[buyer] - 0.1, 1)
                    remainder = 0.0
                    break
    elif remainder < 0:
        # If we somehow allocated more than available (negative remainder),
        # we need to reduce someone's allocation
        while remainder < 0:
            # Find the buyer with the highest allocation
            buyer_to_reduce = max(buyers, key=lambda b: allocations[b])
            if allocations[buyer_to_reduce] >= 0.1:
                allocations[buyer_to_reduce] = round(allocations[buyer_to_reduce] - 0.1, 1)
                remainder = round(remainder + 0.1, 1)
            else:
                # If no buyer has allocation >= 0.1, we can't reduce further
                break

    # Final check to ensure total allocation doesn't exceed remaining_energy
    final_total = round(sum(allocations.values()), 1)
    if final_total > remaining_energy:
        # Find the buyer with the highest allocation to reduce
        buyer_to_reduce = max(buyers, key=lambda b: allocations[b])
        if allocations[buyer_to_reduce] >= 0.1:
            allocations[buyer_to_reduce] = round(allocations[buyer_to_reduce] - 0.1, 1)
            remainder = 0.0

    return allocations, remainder


def allocate_low_priority(remaining_energy, buyers, req_remaining, tariff_diff):
    # Round the remaining energy to 1 decimal place to ensure consistency
    remaining_energy = round(remaining_energy, 1)

    allocations = {buyer: 0.0 for buyer in buyers}
    sorted_buyers = sorted(buyers, key=lambda b: tariff_diff.get(b, 0), reverse=True)

    for buyer in sorted_buyers:
        allocation = min(remaining_energy, req_remaining[buyer])
        allocations[buyer] = round(allocation, 1)  # Round each allocation to 1 decimal place
        req_remaining[buyer] -= allocation
        remaining_energy -= allocation
        remaining_energy = round(remaining_energy, 1)  # Keep rounding to avoid floating point errors

    # Final check to ensure total allocation doesn't exceed original remaining_energy
    total_allocated = round(sum(allocations.values()), 1)
    original_energy = round(sum(allocations.values()) + remaining_energy, 1)

    if total_allocated > original_energy:
        # Find the buyer with the highest allocation to reduce
        if allocations:
            buyer_to_reduce = max(buyers, key=lambda b: allocations.get(b, 0))
            if allocations[buyer_to_reduce] >= 0.1:
                allocations[buyer_to_reduce] = round(allocations[buyer_to_reduce] - 0.1, 1)
                remaining_energy = round(remaining_energy + 0.1, 1)

    return allocations, max(remaining_energy, 0.0)

def check_load_fulfillment(req_remaining, buyer_list, logger):
    """
    Check if all buyers' loads have been fulfilled before proceeding to RTM allocation
    Returns True if all loads are fulfilled (req_remaining <= 0.1 for all buyers)
    """
    unfulfilled_buyers = []
    total_unfulfilled = 0.0

    for buyer in buyer_list:
        remaining_load = req_remaining.get(buyer, 0.0)
        if remaining_load > 0.1:  # Allow small tolerance for rounding
            unfulfilled_buyers.append(buyer)
            total_unfulfilled += remaining_load

    if unfulfilled_buyers:
        logger.info(f"Load fulfillment check: {len(unfulfilled_buyers)} buyers have unfulfilled load")
        logger.info(f"Unfulfilled buyers: {unfulfilled_buyers}")
        logger.info(f"Total unfulfilled load: {total_unfulfilled:.2f} MW")
        return False, unfulfilled_buyers, total_unfulfilled
    else:
        logger.info("Load fulfillment check: All buyers' loads have been fulfilled")
        return True, [], 0.0

def update_cuf_pct_table(curr_date, cuf_pct_dict):
    conn = get_db_connection()
    cursor = conn.cursor()
    for buyer, pct in cuf_pct_dict.items():
        # Convert numpy.float64 to Python float
        pct_float = float(pct) if pct is not None else 0.0
        query = "REPLACE INTO cuf_pct (date, buyer_name, cuf_pct) VALUES (%s, %s, %s)"
        cursor.execute(query, (curr_date, buyer, pct_float))
    conn.commit()
    cursor.close()
    conn.close()

def fetch_buyer_list_for_bifurcation():
    today = today_ist()
    tomorrow = today + datetime.timedelta(days=1)
    try:
        conn = get_db_connection()
        buyer_df = execute_query_to_dataframe(
            "SELECT DISTINCT buyer_name FROM load_data WHERE date IN (%s, %s)",
            conn,
            params=(today, tomorrow)
        )
    except Exception as e:
        st.error(f"❌ Failed to load buyers for manual bifurcation: {str(e)}")
        return []
    finally:
        try:
            conn.close()
        except Exception:
            pass

    if buyer_df.empty or "buyer_name" not in buyer_df.columns:
        return []

    buyers = []
    for buyer in buyer_df["buyer_name"].tolist():
        if buyer and str(buyer).strip():
            buyers.append(str(buyer).strip())

    buyers = sorted(set(buyers))
    if buyers:
        st.caption(f"Showing buyers with load data for {today} and {tomorrow}.")
    return buyers

def render_manual_priority_selector(buyer_list):
    if not buyer_list:
        return {}, False

    # Ensure the manual_priority table exists
    create_manual_priority_table()

    # Load saved priorities from the database
    db_priorities = load_manual_priorities()

    # Initialize session state from DB on first load
    if "manual_bifurcation_priorities" not in st.session_state:
        st.session_state.manual_bifurcation_priorities = {
            buyer: db_priorities.get(buyer, "High") for buyer in buyer_list
        }
        # If there are saved priorities in the DB, mark as ready
        st.session_state.manual_bifurcation_priorities_ready = any(
            buyer in db_priorities for buyer in buyer_list
        )

    # Keep session state aligned with current buyer list
    current_priorities = dict(st.session_state.manual_bifurcation_priorities)
    for buyer in buyer_list:
        if buyer not in current_priorities:
            current_priorities[buyer] = db_priorities.get(buyer, "High")
    for buyer in list(current_priorities.keys()):
        if buyer not in buyer_list:
            current_priorities.pop(buyer)
    st.session_state.manual_bifurcation_priorities = current_priorities

    rows = [
        {"Buyer": buyer, "Low Priority": current_priorities.get(buyer, "High") == "Low"}
        for buyer in buyer_list
    ]
    priority_df = pd.DataFrame(rows)

    with st.form("manual_priority_form", clear_on_submit=False):
        edited_df = st.data_editor(
            priority_df,
            use_container_width=True,
            hide_index=True,
            num_rows="fixed",
            column_config={
                "Buyer": st.column_config.TextColumn("Buyer", disabled=True),
                "Low Priority": st.column_config.CheckboxColumn(
                    "Low Priority",
                    help="Toggle on for Low Priority, off for High Priority."
                )
            },
            key="manual_bifurcation_priority_editor"
        )
        applied = st.form_submit_button("Apply & Save Priorities")

    if applied:
        manual_priority_map = {}
        stored_labels = {}

        for _, row in edited_df.iterrows():
            buyer = row.get("Buyer")
            if not isinstance(buyer, str) or not buyer.strip():
                continue

            is_low = bool(row.get("Low Priority"))
            stored_labels[buyer] = "Low" if is_low else "High"
            manual_priority_map[buyer] = stored_labels[buyer]

        # Save to database for persistence
        if save_manual_priorities(stored_labels):
            st.success("✅ Manual priorities applied and saved to database.")
        else:
            st.warning("⚠️ Priorities applied for this session but failed to save to database.")

        st.session_state.manual_bifurcation_priorities = stored_labels
        st.session_state.manual_bifurcation_priorities_ready = True
    else:
        manual_priority_map = {
            buyer: current_priorities.get(buyer, "High") for buyer in buyer_list
        }
        if st.session_state.get("manual_bifurcation_priorities_ready", False):
            st.caption("✅ Using saved priorities from database. Edit and click 'Apply & Save Priorities' to update.")
        else:
            st.info("Edit priorities and click 'Apply & Save Priorities' to lock them in.")

    high_count = sum(1 for value in manual_priority_map.values() if value == "High")
    low_count = sum(1 for value in manual_priority_map.values() if value == "Low")
    st.caption(f"Manual priorities selected: {high_count} High, {low_count} Low")

    return manual_priority_map, st.session_state.get("manual_bifurcation_priorities_ready", False)

def run_bifurcation():
    st.subheader("Running Bifurcation Process")
    create_tables()
    create_config_tables()

    # Initialize debug logs for this run
    st.session_state.debug_logs = []

    # Setup logging
    logger, log_file = setup_logging()
    logger.info("=== BIFURCATION PROCESS STARTED ===")
    st.info(f"📝 Logging to: {log_file}")

    create_cuf_pct_table()
    logger.info("CUF percentage table created/verified")

    # Check if the final revision file for yesterday exists (using IST)
    import pytz
    ist = pytz.timezone('Asia/Kolkata')
    today = datetime.datetime.now(ist).date()
    tomorrow = today + datetime.timedelta(days=1)
    yesterday = today - timedelta(days=1)

    # Define ID (Intraday) and DA (Day-Ahead) dates.
    id_date = today
    da_date = tomorrow
    ensure_hybrid_avc_cap_defaults([id_date, da_date])

    logger.info(f"Date configuration: Today={today}, Tomorrow={tomorrow}, Yesterday={yesterday}")
    logger.info(f"Analysis dates: ID (Intraday)={id_date}, DA (Day-Ahead)={da_date}")

    # Connect to FTP and check if the file exists
    logger.info("Connecting to FTP server...")
    ftp = connect_ftp()
    logger.info("Searching for final buyer schedule file...")
    file_name = find_final_buyer_schedule_file(ftp, FTP_PATH_FINAL_BUYER_SCHEDULE, yesterday)

    if not file_name:
        ftp.quit()
        error_msg = f"Final Revision Schedule for {yesterday} not found in the folder {FTP_PATH_FINAL_BUYER_SCHEDULE}"
        logger.error(error_msg)
        st.error(error_msg)
        st.error("Bifurcation will only run when Final Revision Schedule is available for the previous day.")
        st.error("Please process the Final Buyer Schedule first.")
        logger.error("Bifurcation process terminated - missing final revision schedule")
        return

    logger.info(f"Found Final Revision Schedule for {yesterday}: {file_name}")
    st.success(f"Found Final Revision Schedule for {yesterday}: {file_name}")

    # Check if CUF% has already been calculated for yesterday
    logger.info("Checking CUF% calculation status...")
    conn = get_db_connection()
    cuf_check_query = f"SELECT COUNT(*) FROM cuf_pct WHERE date = '{yesterday}'"
    cursor = conn.cursor()
    cursor.execute(cuf_check_query)
    cuf_count = cursor.fetchone()[0]
    logger.info(f"CUF% records found for {yesterday}: {cuf_count}")

    # If CUF% hasn't been calculated yet, calculate it now
    if cuf_count == 0:
        logger.warning(f"CUF% for {yesterday} has not been calculated yet. Calculating now...")
        st.warning(f"CUF% for {yesterday} has not been calculated yet. Calculating now...")

        # Download the file
        bio = download_ftp_file(ftp, FTP_PATH_FINAL_BUYER_SCHEDULE, file_name)
        ftp.quit()

        if not bio:
            st.error(f"Failed to download {file_name}")
            return

        # Process the file to calculate CUF%
        try:
            # Hard-coded row range (17-115) as specified earlier
            start_row = 17
            end_row = 115
            start_idx = start_row - 11  # Adjust for header at row 11
            end_idx = end_row - 11 + 1  # +1 because end index is exclusive in pandas

            st.write(f"Processing file: {file_name}")

            # Read all lines to get buyer names (Row 11) and capacities (Row 16)
            bio.seek(0)
            all_lines = bio.read().decode('utf-8').splitlines()

            # Row 11 (index 10) = Buyer names
            # Row 16 (index 15) = Maximum Contract Capacity
            buyer_names_row = all_lines[10].split(',') if len(all_lines) > 10 else []
            capacity_row = all_lines[15].split(',') if len(all_lines) > 15 else []

            st.write(f"\n**Row 11 (Buyer Names):** {len(buyer_names_row)} columns")
            st.write(f"**Row 16 (Capacities):** {len(capacity_row)} columns")

            # Reset and read with pandas
            bio.seek(0)

            # Read the file with header at row 10 (11th row in the file)
            df = pd.read_csv(bio, header=10)

            st.write(f"\n**Pandas found {len(df.columns)} columns**")

            # Get the buyer mapping from the database
            buyer_mapping_df = execute_query_to_dataframe("SELECT file_column_name, actual_buyer_name, capacity_mw FROM buyer_mapping", conn)

            # Create a mapping that uses BOTH file_column_name AND capacity_mw
            # Key: (file_column_name, capacity_mw) -> Value: actual_buyer_name
            buyer_mapping_with_capacity = {}
            for _, row in buyer_mapping_df.iterrows():
                key = (row['file_column_name'], float(row['capacity_mw']))
                buyer_mapping_with_capacity[key] = row['actual_buyer_name']

            # Get contract values (capacity_mw) for CUF calculation from buyer_mapping table
            # Convert Decimal to float to avoid type errors in calculations
            contract_values = {buyer: float(capacity) for buyer, capacity in zip(buyer_mapping_df['actual_buyer_name'], buyer_mapping_df['capacity_mw'])}

            # Calculate total energy for each buyer
            buyer_energy = {}

            # Process each column in the CSV file
            # Match using BOTH column name (from Row 11) AND capacity (from Row 16)
            st.write("\n**Column Matching:**")
            for col_idx, file_col in enumerate(df.columns):
                # Get the buyer name from Row 11 and capacity from Row 16
                # Note: df.columns are from Row 11, so we need to match by index
                if col_idx < len(buyer_names_row) and col_idx < len(capacity_row):
                    buyer_name_from_file = buyer_names_row[col_idx].strip()
                    capacity_from_file_str = capacity_row[col_idx].strip()

                    # Try to parse capacity as float
                    try:
                        capacity_from_file = float(capacity_from_file_str) if capacity_from_file_str else 0
                    except ValueError:
                        capacity_from_file = 0

                    # Skip system columns
                    if capacity_from_file == 0 or buyer_name_from_file in ['Buyer Name', 'Block No.', 'TimeStamp', 'AVC', '']:
                        continue

                    # Match using (file_column_name, capacity_mw)
                    key = (buyer_name_from_file, capacity_from_file)

                    if key in buyer_mapping_with_capacity:
                        actual_buyer_name = buyer_mapping_with_capacity[key]

                        # Sum the energy values using the specified row range
                        energy_values = df.iloc[start_idx:end_idx][file_col].astype(float)
                        total_energy = energy_values.sum()
                        buyer_energy[actual_buyer_name] = total_energy

                        st.write(f"✅ Matched '{buyer_name_from_file}' ({capacity_from_file} MW) → {actual_buyer_name}: {total_energy:.2f} MWh")
                    else:
                        st.warning(f"⚠️ No mapping found for: {buyer_name_from_file} ({capacity_from_file} MW)")

            # Calculate CUF% for each buyer
            cuf_pct = {}
            for buyer, energy in buyer_energy.items():
                if buyer in contract_values and contract_values[buyer] > 0:
                    # Calculate CUF%
                    contract_val = contract_values[buyer]
                    days_in_month = calendar.monthrange(yesterday.year, yesterday.month)[1]
                    denominator = contract_val * 96 * days_in_month
                    daily_cuf_pct = (energy / denominator) * 100 if denominator > 0 else 0
                    cuf_pct[buyer] = daily_cuf_pct
                    st.write(f"Daily CUF% for {buyer}: {daily_cuf_pct:.2f}%")
                else:
                    st.warning(f"No contract value found for {buyer}")

            # Get previous day's cumulative CUF% (using buyer_name AND capacity_mw)
            prev_day = yesterday - timedelta(days=1)
            prev_cuf_query = f"SELECT buyer_name, capacity_mw, cuf_pct FROM cuf_pct WHERE date = '{prev_day}'"
            prev_cuf_df = execute_query_to_dataframe(prev_cuf_query, conn)

            # If no previous day data, try to get the most recent data
            if prev_cuf_df.empty:
                prev_cuf_query = f"""
                    SELECT buyer_name, capacity_mw, cuf_pct
                    FROM cuf_pct
                    WHERE date = (
                        SELECT MAX(date)
                        FROM cuf_pct
                        WHERE date < '{yesterday}'
                    )
                """
                prev_cuf_df = execute_query_to_dataframe(prev_cuf_query, conn)

            # Create dict with (buyer_name, capacity) as key
            prev_cuf_pct = {}
            if not prev_cuf_df.empty:
                for _, row in prev_cuf_df.iterrows():
                    key = (row['buyer_name'], float(row['capacity_mw']))
                    prev_cuf_pct[key] = float(row['cuf_pct'])

            # Calculate cumulative CUF%
            cumulative_cuf_pct = {}
            for buyer, daily_pct in cuf_pct.items():
                capacity = contract_values.get(buyer, 0)
                key = (buyer, capacity)

                # For the first day of the month, start fresh
                if yesterday.day == 1:
                    cumulative_cuf_pct[key] = daily_pct
                else:
                    # Add previous day's CUF% to accumulate over days
                    prev_cuf = prev_cuf_pct.get(key, 0)  # Get last day's CUF or 0
                    cumulative_cuf_pct[key] = prev_cuf + daily_pct

            # Update the CUF% table
            for (buyer, capacity), pct in cumulative_cuf_pct.items():
                query = "REPLACE INTO cuf_pct (date, buyer_name, capacity_mw, cuf_pct) VALUES (%s, %s, %s, %s)"
                cursor.execute(query, (yesterday, buyer, float(capacity), float(pct)))

            conn.commit()
            st.success(f"Successfully calculated and stored CUF% for {yesterday}")

        except Exception as e:
            st.error(f"Error calculating CUF% for {yesterday}: {str(e)}")
            cursor.close()
            conn.close()
            return
    else:
        ftp.quit()
        st.success(f"CUF% for {yesterday} already calculated. Using existing values.")

    cursor.close()
    logger.info("Starting data fetching from MySQL database...")
    st.write("Fetching source data from MySQL...")
    conn = get_db_connection()

    # --- Load Schedules ---
    logger.info("Fetching solar schedule data...")
    solar_df = execute_query_to_dataframe(
        "SELECT date, block, sch as sch_solar, avc as avc_solar FROM schedule_solar",
        conn
    )
    solar_df['date'] = pd.to_datetime(solar_df['date'])
    logger.info(f"Solar schedule records: {len(solar_df)}")

    logger.info("Fetching wind schedule data...")
    wind_df = execute_query_to_dataframe(
        "SELECT date, block, sch as sch_wind, avc as avc_wind FROM schedule_wind",
        conn
    )
    wind_df['date'] = pd.to_datetime(wind_df['date'])
    logger.info(f"Wind schedule records: {len(wind_df)}")

    schedule_df = pd.merge(solar_df, wind_df, on=["date", "block"], how="outer")
    schedule_df.fillna(0, inplace=True)
    schedule_df["hyb_sch"] = schedule_df["sch_solar"] + schedule_df["sch_wind"]
    logger.info(f"Combined schedule records: {len(schedule_df)}")

    # --- Load Obligation Data ---
    logger.info("Fetching obligation data...")
    obligation_df = execute_query_to_dataframe(
        "SELECT date, block, market, market_fp, market_linde FROM obligation",
        conn
    )
    obligation_df = normalize_obligation_dataframe(obligation_df)
    obligation_df['date'] = pd.to_datetime(obligation_df['date'])
    obligation_df.rename(
        columns={
            "market": "GDAM",
        },
        inplace=True,
    )
    logger.info(f"Obligation records: {len(obligation_df)}")

    logger.info("Using manual GDAM values from obligation table only; no GDAM residue split will be calculated.")

    schedule_merged = pd.merge(schedule_df, obligation_df, on=["date", "block"], how="left")
    hybrid_cap_df = execute_query_to_dataframe("SELECT date, avc_cap FROM hybrid_avc_cap", conn)
    gdam_allocation_pct = get_gdam_allocation_pct(conn)
    gdam_allocation_ratio = round(max(to_float(gdam_allocation_pct), 0.0) / 100.0, 4)
    if hybrid_cap_df.empty:
        hybrid_cap_df = pd.DataFrame(
            {
                "date": pd.to_datetime([id_date, da_date]),
                "avc_cap": [DEFAULT_HYBRID_AVC_CAP, DEFAULT_HYBRID_AVC_CAP],
            }
        )
    else:
        hybrid_cap_df["date"] = pd.to_datetime(hybrid_cap_df["date"])

    schedule_merged = pd.merge(schedule_merged, hybrid_cap_df, on=["date"], how="left")
    schedule_merged["avc_cap"] = schedule_merged["avc_cap"].fillna(DEFAULT_HYBRID_AVC_CAP)
    logger.info(
        f"G-DAM allocation percentage for Day Ahead RTM outputs: {gdam_allocation_pct:.0f}%"
    )

    schedule_merged["input_solar_schedule"] = schedule_merged["sch_solar"].apply(to_float)
    schedule_merged["input_wind_schedule"] = schedule_merged["sch_wind"].apply(to_float)
    hybrid_schedule_df = schedule_merged.apply(
        lambda row: pd.Series(
            compute_hybrid_schedule_from_validation_limit(
                solar_input=row.get("sch_solar", 0),
                wind_input=row.get("sch_wind", 0),
                solar_avc=row.get("avc_solar", 0),
                wind_avc=row.get("avc_wind", 0),
                hybrid_cap=row.get("avc_cap", DEFAULT_HYBRID_AVC_CAP),
            )
        ),
        axis=1,
    )
    for col in hybrid_schedule_df.columns:
        schedule_merged[col] = hybrid_schedule_df[col]

    schedule_merged["sch_solar"] = schedule_merged["solar_schedule_final"]
    schedule_merged["sch_wind"] = schedule_merged["wind_schedule_final"]
    schedule_merged["hyb_sch"] = schedule_merged["hybrid_schedule_final"]

    avc_distribution_df = schedule_merged.apply(
        lambda row: pd.Series(
            compute_hybrid_avc_distribution(
                row.get("sch_wind", 0),
                row.get("sch_solar", 0),
                row.get("avc_wind", 0),
                row.get("avc_solar", 0),
                row.get("avc_cap", DEFAULT_HYBRID_AVC_CAP),
            )
        ),
        axis=1,
    )
    for col in avc_distribution_df.columns:
        schedule_merged[col] = avc_distribution_df[col]

    # Compute Adjusted Schedule
    schedule_merged["Adjusted_Sch"] = schedule_merged["hyb_sch"]

    # Round schedule and market values to 1 decimal place (except Date and Block)
    cols_to_round = [
        "sch_solar",
        "sch_wind",
        "hyb_sch",
        "declared_plant_schedule",
        "plant_avc_input",
        "effective_hybrid_cap",
        "validation_limit",
        "GDAM",
        "Adjusted_Sch",
        "avc_solar",
        "avc_wind",
        "hybrid_avc_input",
        "avc_cap",
        "hybrid_avc",
        "solar_avc_capped",
        "wind_avc_capped",
    ]
    for col in cols_to_round:
        schedule_merged[col] = schedule_merged[col].round(1)

    # Process only the two analysis dates for this run.
    # Without this filter, a prior month-start date (e.g. 1st) can reset CUF
    # before we classify buyers for today's ID/DA runs.
    schedule_merged = schedule_merged[
        schedule_merged["date"].dt.date.isin([id_date, da_date])
    ].copy()
    schedule_merged.sort_values(["date", "block"], inplace=True)
    schedule_merged.reset_index(drop=True, inplace=True)

    # Fetch contract values.
    logger.info("Fetching contract values...")
    contract_value_df = execute_query_to_dataframe("SELECT buyer_name, contract_value FROM contract_value", conn)
    contract_values = {row['buyer_name']: row['contract_value'] for _, row in contract_value_df.iterrows()}
    buyer_list = list(contract_values.keys())
    logger.info(f"Contract values loaded for {len(buyer_list)} buyers: {buyer_list}")

    # Fetch Requisition data.
    load_data_df = execute_query_to_dataframe("SELECT date, block, buyer_name, load_value FROM load_data", conn)
    load_data_df['date'] = pd.to_datetime(load_data_df['date'])
    if load_data_df.empty:
        st.error("No requisition data found in load_data table.")
        conn.close()
        return

    # Check if we have load data for both ID and DA dates
    id_load_count = len(load_data_df[load_data_df['date'].dt.date == id_date])
    da_load_count = len(load_data_df[load_data_df['date'].dt.date == da_date])

    st.info(f"Load data availability: ID ({id_date}): {id_load_count} records, DA ({da_date}): {da_load_count} records")

    # Check for missing load data - ID is mandatory, DA is optional
    if id_load_count == 0:
        st.error(f"❌ **No load data found for Intraday date ({id_date})**")
        st.error("**Bifurcation cannot proceed without Intraday load data.**")

        st.write("### 🔍 **Troubleshooting Steps:**")

        col1, col2 = st.columns([1, 1])

        with col1:
            st.write("#### 📁 **Check FTP Files:**")
            expected_filename = id_date.strftime("%Y%m%d")
            st.code(f"""
FTP Path: {FTP_PATH_LOAD}
Expected files:
- {expected_filename}.xlsx
- {expected_filename}.csv
            """)

            st.write("#### 🔧 **Manual Data Entry:**")
            st.info("Go to **Manual Data Entry** → **load_data** to:")
            st.write("- Upload CSV/Excel file directly to FTP")
            st.write("- Edit load data manually")
            st.write("- Ensure all 96 time blocks are present")

        with col2:
            st.write("#### 📊 **Load Data Requirements:**")
            st.write("**Required columns:**")
            st.code(f"""
Date        | Time Block | [Buyer1] | [Buyer2] | ...
{id_date}  | 1          | 5.0      | 8.0      | ...
{id_date}  | 2          | 5.2      | 8.1      | ...
...         | ...        | ...      | ...      | ...
{id_date}  | 96         | 4.8      | 7.9      | ...
            """)

            st.write("**Validation:**")
            st.write("- ✅ Exactly 96 time blocks (1-96)")
            st.write("- ✅ Date column with correct date")
            st.write("- ✅ Buyer columns with load values")

        st.write("---")
        st.error("**🚫 Bifurcation stopped. Please fix load data and try again.**")
        logger.error("Bifurcation process terminated - missing Intraday load data")
        return

    # DA load data is optional - show warning but continue
    if da_load_count == 0:
        st.warning(f"⚠️ **No load data found for Day Ahead date ({da_date})**")
        st.warning("**Bifurcation will proceed with Intraday data only.**")

        st.write("### ℹ️ **Day Ahead Load Data Missing:**")

        col1, col2 = st.columns([1, 1])

        with col1:
            st.write("#### 📁 **Expected FTP Files:**")
            expected_filename = da_date.strftime("%Y%m%d")
            st.code(f"""
FTP Path: {FTP_PATH_LOAD}
Expected files:
- {expected_filename}.xlsx
- {expected_filename}.csv
            """)

            st.write("#### 📝 **Note:**")
            st.info("Day Ahead data can be added later via:")
            st.write("- **Manual Data Entry** → **load_data**")
            st.write("- Upload file to FTP and reload data")

        with col2:
            st.write("#### ⚙️ **Current Processing:**")
            st.write("✅ Intraday (ID) bifurcation will run")
            st.write("⏭️ Day Ahead (DA) bifurcation will be skipped")
            st.write("🔄 Re-run bifurcation after adding DA data")

        st.write("---")
        logger.warning(f"Day Ahead load data missing for {da_date} - proceeding with Intraday only")

    requisition = load_data_df.pivot_table(index=["date", "block"],
                                           columns="buyer_name",
                                           values="load_value",
                                           aggfunc='first').reset_index()

    # Fetch tariff differences.
    tariff_diff_df = execute_query_to_dataframe("SELECT block, state, tariff_difference FROM tariff_difference", conn)
    state_df = execute_query_to_dataframe("SELECT buyer_name, state FROM state", conn)
    cuf_query = f"SELECT buyer_name, cuf_pct FROM cuf_pct WHERE date = '{yesterday}'"
    cuf_data_df = execute_query_to_dataframe(cuf_query, conn)
    conn.close()
    buyer_state = {row['buyer_name']: row['state'] for _, row in state_df.iterrows()}
    tariff_lookup = {
        (int(row["block"]), row["state"]): to_float(row["tariff_difference"])
        for _, row in tariff_diff_df.iterrows()
    }

    daily_active_df = load_data_df.copy()
    daily_active_df["date_only"] = daily_active_df["date"].dt.date
    daily_active_df["load_value"] = daily_active_df["load_value"].apply(to_float)
    daily_active_df = (
        daily_active_df
        .groupby(["date_only", "buyer_name"], as_index=False)["load_value"]
        .sum()
    )
    daily_active_df = daily_active_df[daily_active_df["load_value"] > 0]
    daily_active_buyers_by_date = {
        date_value: set(group["buyer_name"].tolist())
        for date_value, group in daily_active_df.groupby("date_only")
    }
    buyer_access_types = load_buyer_access_types()
    buyer_access_types = {
        buyer: normalize_buyer_access_type(buyer_access_types.get(buyer))
        for buyer in buyer_list
    }

    # -------------------------------
    # Bifurcation Process
    # -------------------------------
    cumulative_allocations = {buyer: {'solar': 0, 'wind': 0} for buyer in buyer_list}
    allocation_results = []
    energy_mismatch_records = []

    unique_dates = schedule_merged["date"].dt.date.unique()
    prev_month = None

    # Fetch low priority buyers from Excel on FTP
    logger.info("Connecting to FTP to fetch low priority buyers dictionary...")
    try:
        ftp = connect_ftp()
        excel_low_priority_buyers = get_low_priority_buyers_from_excel(ftp)
        ftp.quit()
        logger.info(f"Low priority buyers listed in Excel: {excel_low_priority_buyers}")
    except Exception as e:
        logger.error(f"Error fetching low priority buyers from Excel: {str(e)}")
        excel_low_priority_buyers = []
        st.warning("⚠️ Could not fetch low priority buyers list from FTP. Defaulting all to High Priority.")

    # Classify buyers globally for this run
    high_priority_buyers = []
    low_priority_buyers = []

    for buyer in buyer_list:
        if buyer in excel_low_priority_buyers:
            low_priority_buyers.append(buyer)
            priority = "Low"
        else:
            high_priority_buyers.append(buyer)
            priority = "High"

    logger.info(f"Buyer classification results:")
    logger.info(f"  High Priority (Default): {high_priority_buyers}")
    logger.info(f"  Low Priority (Excel): {low_priority_buyers}")

    # Store classification info for debug logs page
    classification_debug_info = {
        'type': 'priority_classification',
        'high_priority_count': len(high_priority_buyers),
        'low_priority_count': len(low_priority_buyers),
        'high_priority_buyers': high_priority_buyers,
        'low_priority_buyers': low_priority_buyers,
        'source': 'Excel file from FTP'
    }
    st.session_state.debug_logs.append(classification_debug_info)

    for curr_date in sorted(unique_dates):
        if prev_month is None or curr_date.month != prev_month:
            # Reset cumulative allocations at the beginning of each month
            cumulative_allocations = {buyer: {'solar': 0, 'wind': 0} for buyer in buyer_list}

            # Reset CUF% at the beginning of each month
            if curr_date.day == 1:
                prev_day_cuf_pct = {buyer: 0 for buyer in buyer_list}

            prev_month = curr_date.month

        # Reset cumulative allocations at the beginning of each day to track only today's allocations
        cumulative_allocations = {buyer: {'solar': 0, 'wind': 0} for buyer in buyer_list}

        day_schedule = schedule_merged[schedule_merged["date"].dt.date == curr_date].drop_duplicates(subset=["block"])

        active_day_buyers = daily_active_buyers_by_date.get(curr_date, set())

        # Log buyer classification for this date
        logger.info(f"Date {curr_date} - Using Excel-based buyer classification:")
        logger.info(f"  High Priority: {high_priority_buyers}")
        logger.info(f"  Low Priority: {low_priority_buyers}")

        # Store debug info for debug logs page
        if 'debug_logs' not in st.session_state:
            st.session_state.debug_logs = []

        debug_info = {
            'type': 'date_processing',
            'date': curr_date,
            'high_priority_buyers': high_priority_buyers,
            'low_priority_buyers': low_priority_buyers,
            'schedule_records': len(day_schedule)
        }
        st.session_state.debug_logs.append(debug_info)

        # Identify Linde and non-Linde buyers
        linde_buyers = [b for b in buyer_list if "linde" in b.lower()]
        non_linde_buyers = [b for b in buyer_list if "linde" not in b.lower()]
        active_linde_day_buyers = [b for b in linde_buyers if b in active_day_buyers]
        active_non_linde_day_buyers = [b for b in non_linde_buyers if b in active_day_buyers]
        day_linde_contract_total = sum(to_float(contract_values.get(b, 0.0)) for b in active_linde_day_buyers)
        day_non_linde_contract_total = sum(to_float(contract_values.get(b, 0.0)) for b in active_non_linde_day_buyers)
        
        logger.info(f"Date {curr_date} - Groups identified:")
        logger.info(f"  Linde Buyers: {linde_buyers}")
        logger.info(f"  Non-Linde Buyers: {non_linde_buyers}")
        logger.info(
            f"  Daily Active Contract Split - Linde: {day_linde_contract_total:.1f} ({active_linde_day_buyers}), "
            f"Non-Linde: {day_non_linde_contract_total:.1f} ({active_non_linde_day_buyers})"
        )

        # Debug: Track first few blocks for detailed analysis
        debug_blocks = [1, 2, 3]

        for _, row in day_schedule.iterrows():
            block = row["block"]
            solar_val = row["sch_solar"]
            wind_val = row["sch_wind"]
            total_sch = solar_val + wind_val

            req_row = requisition[
                (pd.to_datetime(requisition["date"]).dt.date == curr_date) &
                (requisition["block"] == block)
            ]
            req_remaining = {buyer: 0 for buyer in buyer_list}
            if not req_row.empty:
                req_row = req_row.iloc[0]
                for buyer in buyer_list:
                    req_remaining[buyer] = round(max(to_float(req_row.get(buyer, 0)), 0.0), 1)

            req_original = req_remaining.copy()
            active_linde_block_buyers, linde_contract_total = get_active_contract_totals(
                contract_values,
                linde_buyers,
                req_original,
            )
            active_non_linde_block_buyers, non_linde_contract_total = get_active_contract_totals(
                contract_values,
                non_linde_buyers,
                req_original,
            )
            group_split = split_energy_with_post_linde_gdam(
                total_schedule=row["hyb_sch"],
                gdam_fp=row.get("GDAM", 0),
                gdam_linde=0,
                linde_weight=linde_contract_total,
                non_linde_weight=non_linde_contract_total,
            )
            adjusted_sch = group_split["adjusted_schedule"]
            remaining_energy = group_split["post_gdam_available_total"]
            remaining_solar_budget = round(max(to_float(solar_val), 0.0), 1)
            remaining_wind_budget = round(max(to_float(wind_val), 0.0), 1)

            linde_actual_demand = round(sum(req_original.get(buyer, 0.0) for buyer in active_linde_block_buyers), 1)
            linde_additional_energy = round(max(linde_contract_total - linde_actual_demand, 0.0), 1)

            allocation_row_total = {buyer: 0.0 for buyer in buyer_list}
            allocation_row = {f"{buyer}_Solar": 0.0 for buyer in buyer_list}
            allocation_row.update({f"{buyer}_Wind": 0.0 for buyer in buyer_list})
            source_usage_row = {f"{buyer}_Solar": 0.0 for buyer in buyer_list}
            source_usage_row.update({f"{buyer}_Wind": 0.0 for buyer in buyer_list})
            block_tariff_by_buyer = {
                buyer: tariff_lookup.get((int(block), buyer_state.get(buyer)), 0.0)
                for buyer in buyer_list
            }

            # Log block start summary
            logger.info(f"")
            logger.info(
                f"📊 BLOCK {block} START - Schedule: Solar={solar_val:.1f}, Wind={wind_val:.1f}, "
                f"Total={total_sch:.1f}, GDAM={to_float(row.get('GDAM', 0)):.1f}, Adjusted={adjusted_sch:.1f}, "
                f"Available={remaining_energy:.1f}"
            )
            buyer_loads = [req_remaining.get(buyer, 0.0) for buyer in buyer_list]
            logger.info(f"📊 BLOCK {block} LOADS - Buyers: {buyer_list}, Requisitions: {[round(load, 1) for load in buyer_loads]}")

            # Store debug info for first few blocks
            if block in debug_blocks:
                block_debug = {
                    'type': 'block_processing',
                    'date': curr_date,
                    'block': block,
                    'solar_val': solar_val,
                    'wind_val': wind_val,
                    'total_sch': total_sch,
                    'gdam': row['GDAM'],
                    'adjusted_sch': adjusted_sch,
                    'remaining_energy': remaining_energy,
                    'block_linde_contract_total': linde_contract_total,
                    'block_non_linde_contract_total': non_linde_contract_total,
                    'group_split': group_split.copy(),
                    'requisition': req_remaining.copy(),
                    'req_row_empty': req_row.empty,
                    'raw_req_row': dict(req_row) if not req_row.empty else {}
                }
                st.session_state.debug_logs.append(block_debug)

            logger.info(
                f"Block {block}: Contract Split After GDAM - Available: {group_split['available_after_fp']:.1f}, "
                f"Linde Pre-GDAM: {group_split['linde_pre_gdam']:.1f}, "
                f"Linde For Buyers: {group_split['linde_available']:.1f}, Non-Linde Group: {group_split['non_linde_available']:.1f}, "
                f"Schedule Boost: {group_split['schedule_boost']:.1f}"
            )
            round_counter = 0
            while remaining_energy > 0.1:
                active_linde_now, round_linde_weight = get_active_contract_totals(
                    contract_values,
                    linde_buyers,
                    req_remaining,
                )
                active_non_linde_now, round_non_linde_weight = get_active_contract_totals(
                    contract_values,
                    non_linde_buyers,
                    req_remaining,
                )

                if not active_linde_now and not active_non_linde_now:
                    break

                if round_counter == 0:
                    round_linde_budget = group_split["linde_available"]
                    round_non_linde_budget = group_split["non_linde_available"]
                else:
                    round_linde_budget, round_non_linde_budget = split_energy_by_contract(
                        remaining_energy,
                        round_linde_weight,
                        round_non_linde_weight,
                    )

                round_planned_totals = {buyer: 0.0 for buyer in buyer_list}

                if round_linde_budget > 0 and active_linde_now:
                    linde_hp = [b for b in high_priority_buyers if b in active_linde_now]
                    linde_lp = [b for b in low_priority_buyers if b in active_linde_now]

                    if linde_hp:
                        hp_targets = {b: req_remaining[b] for b in linde_hp}
                        hp_alloc, round_linde_budget = allocate_high_priority(round_linde_budget, linde_hp, hp_targets)
                        for buyer, amount in hp_alloc.items():
                            round_planned_totals[buyer] = round(round_planned_totals.get(buyer, 0.0) + amount, 1)

                    if round_linde_budget > 0 and linde_lp:
                        lp_targets = {b: req_remaining[b] for b in linde_lp}
                        lp_tariffs = {b: block_tariff_by_buyer.get(b, 0.0) for b in linde_lp}
                        lp_alloc, round_linde_budget = allocate_low_priority(round_linde_budget, linde_lp, lp_targets, lp_tariffs)
                        for buyer, amount in lp_alloc.items():
                            round_planned_totals[buyer] = round(round_planned_totals.get(buyer, 0.0) + amount, 1)

                if round_non_linde_budget > 0 and active_non_linde_now:
                    non_linde_hp = [b for b in high_priority_buyers if b in active_non_linde_now]
                    non_linde_lp = [b for b in low_priority_buyers if b in active_non_linde_now]

                    if non_linde_hp:
                        hp_targets = {b: req_remaining[b] for b in non_linde_hp}
                        hp_alloc, round_non_linde_budget = allocate_high_priority(round_non_linde_budget, non_linde_hp, hp_targets)
                        for buyer, amount in hp_alloc.items():
                            round_planned_totals[buyer] = round(round_planned_totals.get(buyer, 0.0) + amount, 1)

                    if round_non_linde_budget > 0 and non_linde_lp:
                        lp_targets = {b: req_remaining[b] for b in non_linde_lp}
                        lp_tariffs = {b: block_tariff_by_buyer.get(b, 0.0) for b in non_linde_lp}
                        lp_alloc, round_non_linde_budget = allocate_low_priority(round_non_linde_budget, non_linde_lp, lp_targets, lp_tariffs)
                        for buyer, amount in lp_alloc.items():
                            round_planned_totals[buyer] = round(round_planned_totals.get(buyer, 0.0) + amount, 1)

                round_planned_total = round(sum(round_planned_totals.values()), 1)
                if round_planned_total <= 0.1:
                    break

                round_allocation_row, round_served_totals, _, round_source_usage = assign_source_eligible_allocations(
                    round_planned_totals,
                    buyer_list,
                    buyer_access_types,
                    remaining_solar_budget,
                    remaining_wind_budget,
                    high_priority_buyers,
                    low_priority_buyers,
                    block_tariff_by_buyer,
                )

                served_round_total = round(sum(round_served_totals.values()), 1)
                used_solar_round = round(
                    sum(round_source_usage.get(f"{buyer}_Solar", 0.0) for buyer in buyer_list),
                    1,
                )
                used_wind_round = round(
                    sum(round_source_usage.get(f"{buyer}_Wind", 0.0) for buyer in buyer_list),
                    1,
                )

                if served_round_total <= 0.1 and used_solar_round <= 0.1 and used_wind_round <= 0.1:
                    break

                for buyer in buyer_list:
                    served_amount = round(round_served_totals.get(buyer, 0.0), 1)
                    allocation_row_total[buyer] = round(allocation_row_total.get(buyer, 0.0) + served_amount, 1)
                    req_remaining[buyer] = round(max(req_remaining.get(buyer, 0.0) - served_amount, 0.0), 1)
                    allocation_row[f"{buyer}_Solar"] = round(
                        allocation_row.get(f"{buyer}_Solar", 0.0) + round_allocation_row.get(f"{buyer}_Solar", 0.0),
                        1,
                    )
                    allocation_row[f"{buyer}_Wind"] = round(
                        allocation_row.get(f"{buyer}_Wind", 0.0) + round_allocation_row.get(f"{buyer}_Wind", 0.0),
                        1,
                    )
                    source_usage_row[f"{buyer}_Solar"] = round(
                        source_usage_row.get(f"{buyer}_Solar", 0.0) + round_source_usage.get(f"{buyer}_Solar", 0.0),
                        1,
                    )
                    source_usage_row[f"{buyer}_Wind"] = round(
                        source_usage_row.get(f"{buyer}_Wind", 0.0) + round_source_usage.get(f"{buyer}_Wind", 0.0),
                        1,
                    )

                remaining_solar_budget = round(max(remaining_solar_budget - used_solar_round, 0.0), 1)
                remaining_wind_budget = round(max(remaining_wind_budget - used_wind_round, 0.0), 1)
                remaining_energy = round(max(remaining_energy - served_round_total, 0.0), 1)
                round_counter += 1

            if block in debug_blocks:
                group_debug = {
                    'type': 'group_allocation',
                    'date': curr_date,
                    'block': block,
                    'rounds': round_counter,
                    'block_linde_contract_total': linde_contract_total,
                    'block_non_linde_contract_total': non_linde_contract_total,
                    'final_allocations': allocation_row_total.copy(),
                    'remaining_energy_after_loops': remaining_energy,
                    'remaining_solar_budget': remaining_solar_budget,
                    'remaining_wind_budget': remaining_wind_budget,
                }
                st.session_state.debug_logs.append(group_debug)

            for buyer in buyer_list:
                alloc_solar = round(source_usage_row.get(f"{buyer}_Solar", 0.0), 1)
                alloc_wind = round(source_usage_row.get(f"{buyer}_Wind", 0.0), 1)
                cumulative_allocations[buyer]['solar'] += alloc_solar
                cumulative_allocations[buyer]['wind'] += alloc_wind

            # Store final allocation debug info
            if block in debug_blocks:
                solar_wind_split = {}
                for buyer in buyer_list:
                    solar_key = f"{buyer}_Solar"
                    wind_key = f"{buyer}_Wind"
                    solar_wind_split[buyer] = {
                        'solar': allocation_row.get(solar_key, 'MISSING'),
                        'wind': allocation_row.get(wind_key, 'MISSING')
                    }

                # Check for None values
                none_values = []
                for key, value in allocation_row.items():
                    if is_missing_scalar(value):
                        none_values.append(key)

                split_debug = {
                    'type': 'solar_wind_split',
                    'date': curr_date,
                    'block': block,
                    'solar_wind_split': solar_wind_split,
                    'source_usage_split': {
                        buyer: {
                            'solar': source_usage_row.get(f"{buyer}_Solar", 'MISSING'),
                            'wind': source_usage_row.get(f"{buyer}_Wind", 'MISSING')
                        }
                        for buyer in buyer_list
                    },
                    'allocation_row_keys': list(allocation_row.keys()),
                    'none_values': none_values,
                    'has_none_values': len(none_values) > 0,
                    'remaining_energy_after_loops': remaining_energy,
                    'remaining_solar_budget': remaining_solar_budget,
                    'remaining_wind_budget': remaining_wind_budget,
                }
                st.session_state.debug_logs.append(split_debug)

            # Log detailed block allocation information
            buyer_loads = [req_remaining.get(buyer, 0.0) for buyer in buyer_list]
            buyer_allocations = [allocation_row_total.get(buyer, 0.0) for buyer in buyer_list]

            logger.info(f"Block {block}: Load of Buyers {buyer_list} is {[round(load, 1) for load in buyer_loads]} and schedule is {adjusted_sch:.1f}. Allocated energy is {[round(alloc, 1) for alloc in buyer_allocations]} Remaining energy for RTM is {remaining_energy:.1f}")

            energy_balance_total = round(
                sum(allocation_row_total.values())
                + remaining_energy
                + to_float(row.get("GDAM", 0)),
                1,
            )
            if abs(energy_balance_total - adjusted_sch) > 0.11:
                energy_mismatch_records.append(
                    {
                        "date": curr_date,
                        "block": block,
                        "allocated": round(sum(allocation_row_total.values()), 1),
                        "remaining": remaining_energy,
                        "gdam": round(to_float(row.get("GDAM", 0)), 1),
                        "adjusted": round(adjusted_sch, 1),
                        "total": energy_balance_total,
                    }
                )
                logger.warning(
                    f"Energy mismatch in block {block}: allocated={round(sum(allocation_row_total.values()), 1)}, "
                    f"remaining={remaining_energy:.1f}, GDAM={to_float(row.get('GDAM', 0)):.1f}, total={energy_balance_total:.1f}, "
                    f"adjusted={adjusted_sch:.1f}"
                )

            # Check load fulfillment before proceeding to RTM allocation
            load_fulfilled, unfulfilled_buyers, total_unfulfilled = check_load_fulfillment(req_remaining, buyer_list, logger)

            # GDAM is input-only from the obligation table. Any residue after
            # all buyer loads are fulfilled becomes RTM-eligible energy.
            gdam_solar = 0.0
            gdam_wind = 0.0
            residue_solar = 0.0
            residue_wind = 0.0
            rtm_solar = 0.0
            rtm_wind = 0.0

            if load_fulfilled and remaining_energy > 0.1:
                residue_solar, residue_wind = split_energy_by_contract(
                    remaining_energy,
                    remaining_solar_budget,
                    remaining_wind_budget,
                )
                rtm_solar = residue_solar
                rtm_wind = residue_wind

                logger.info(f"Block {block}: ✅ RTM allocation proceeding - all loads fulfilled")
                logger.info(
                    f"Block {block}: 🏭 RESIDUE SPLIT - GDAM: Solar=0.0, Wind=0.0 | "
                    f"RTM: Solar={rtm_solar:.1f}, Wind={rtm_wind:.1f}"
                )
            else:
                logger.warning(f"Block {block}: ❌ RTM allocation SKIPPED - Unfulfilled buyers: {unfulfilled_buyers} with loads: {[round(req_remaining.get(buyer, 0), 1) for buyer in unfulfilled_buyers]}")
                logger.info(
                    "Block %s: 🏭 RESIDUE SPLIT - GDAM: Solar=0.0, Wind=0.0 | RTM: Solar=0.0, Wind=0.0",
                    block,
                )

            if curr_date == da_date:
                raw_gdam_solar_output = rtm_solar * gdam_allocation_ratio
                raw_gdam_wind_output = rtm_wind * gdam_allocation_ratio
                gdam_solar_output = round_down_one_decimal(raw_gdam_solar_output)
                gdam_wind_output = round_down_one_decimal(raw_gdam_wind_output)
                rtm_solar = round(max(rtm_solar - gdam_solar_output, 0.0), 1)
                rtm_wind = round(max(rtm_wind - gdam_wind_output, 0.0), 1)
            else:
                gdam_solar_output = None
                gdam_wind_output = None

            result_entry = {
                "Date": row["date"],
                "Block": block,
                "Solar_Sch": solar_val,
                "Wind_Sch": wind_val,
                "Hyb_Sch": row["hyb_sch"],
                "Declared_Plant_Schedule": row.get("declared_plant_schedule", row["hyb_sch"]),
                "Plant_AvC": row.get("plant_avc_input", 0),
                "Declared_Hybrid_Schedule": row["hyb_sch"],
                "Total_Hybrid_Solar": solar_val,
                "Total_Hybrid_Wind": wind_val,
                "Solar_AvC": row.get("avc_solar", 0),
                "Wind_AvC": row.get("avc_wind", 0),
                "Hybrid_AvC_Input": row.get("hybrid_avc_input", 0),
                "Hybrid_AvC_Cap": row.get("avc_cap", DEFAULT_HYBRID_AVC_CAP),
                "Capped_Solar_AvC": row.get("solar_avc_capped", 0),
                "Capped_Wind_AvC": row.get("wind_avc_capped", 0),
                "Hybrid_AvC": row.get("hybrid_avc", 0),
                "GDAM": row["GDAM"],
                "Adjusted_Sch": adjusted_sch,
                "Linde_Contract_Total": linde_contract_total,
                "Linde_Actual_Demand": linde_actual_demand,
                "Linde_Additional_Energy": linde_additional_energy,
                "RTM_Available": round(rtm_solar + rtm_wind, 1),
                "RTM_Solar": rtm_solar,
                "RTM_Wind": rtm_wind,
                "G-DAM_Solar_Output": gdam_solar_output,
                "G-DAM_Wind_Output": gdam_wind_output,
            }
            for buyer in buyer_list:
                result_entry[f"{buyer}_Solar"] = allocation_row.get(f"{buyer}_Solar", 0)
                result_entry[f"{buyer}_Wind"] = allocation_row.get(f"{buyer}_Wind", 0)

            # Store result entry debug info
            if block in debug_blocks:
                buyer_entries = {k: v for k, v in result_entry.items() if any(buyer in k for buyer in buyer_list)}

                # Check for None values in result entry
                none_in_result = []
                for key, value in result_entry.items():
                    if is_missing_scalar(value):
                        none_in_result.append(key)

                result_debug = {
                    'type': 'result_entry',
                    'date': curr_date,
                    'block': block,
                    'buyer_allocations': buyer_entries,
                    'none_values_in_result': none_in_result,
                    'has_none_in_result': len(none_in_result) > 0
                }
                st.session_state.debug_logs.append(result_debug)

            allocation_results.append(result_entry)

        # No CUF% calculation needed here as we're using yesterday's values from the database

    result_df = pd.DataFrame(allocation_results)
    result_df['Block'] = result_df['Block'].astype(int)
    result_df = result_df.groupby(['Date', 'Block'], as_index=False).first()
    result_df = result_df.sort_values(['Date', 'Block']).round(2)

    result_df["Date_only"] = pd.to_datetime(result_df["Date"]).dt.date

    # Create a "Date + Block" column with Excel serial number + block
    # Step 1: Convert pandas datetime to datetime objects
    date_objects = pd.to_datetime(result_df["Date"])

    # Step 2: Define Excel base date
    excel_base = datetime.datetime(1899, 12, 30)

    # Step 3: Calculate Excel serial numbers
    excel_serials = [(date_obj - excel_base).days for date_obj in date_objects]

    # Step 4: Concatenate with block numbers
    result_df["Date + Block"] = [f"{serial}{block}" for serial, block in zip(excel_serials, result_df["Block"])]

    # Reorder columns to put "Date + Block" before "Date"
    cols = result_df.columns.tolist()
    date_idx = cols.index("Date")
    date_block_idx = cols.index("Date + Block")
    cols.pop(date_block_idx)
    cols.insert(date_idx, "Date + Block")
    result_df = result_df[cols]

    id_results = result_df[result_df["Date_only"] == id_date]
    da_results = result_df[result_df["Date_only"] == da_date]

    # Log output results for both ID and DA
    logger.info("=== BIFURCATION RESULTS SUMMARY ===")

    # Calculate summary statistics for ID results
    if not id_results.empty:
        id_summary = {
            'total_blocks': len(id_results),
            'total_hybrid_schedule': id_results['Hyb_Sch'].sum(),
            'total_gdam': id_results['GDAM'].sum(),
            'total_adjusted_schedule': id_results['Adjusted_Sch'].sum(),
            'total_rtm_available': id_results['RTM_Available'].sum(),
            'total_rtm_solar': id_results['RTM_Solar'].sum(),
            'total_rtm_wind': id_results['RTM_Wind'].sum(),
            'buyer_allocations': {}
        }

        # Calculate buyer allocations
        for buyer in buyer_list:
            solar_col = f"{buyer}_Solar"
            wind_col = f"{buyer}_Wind"
            if solar_col in id_results.columns and wind_col in id_results.columns:
                total_solar = id_results[solar_col].sum()
                total_wind = id_results[wind_col].sum()
                id_summary['buyer_allocations'][buyer] = {
                    'solar': total_solar,
                    'wind': total_wind,
                    'total': total_solar + total_wind
                }

        logger.info(f"ID (Intraday) results for {id_date}: {id_summary}")
        log_output_results("INTRADAY", id_date, id_summary, id_results)
    else:
        logger.warning(f"No ID (Intraday) results generated for {id_date}")

    # Calculate summary statistics for DA results
    if not da_results.empty:
        da_summary = {
            'total_blocks': len(da_results),
            'total_hybrid_schedule': da_results['Hyb_Sch'].sum(),
            'total_gdam': da_results['GDAM'].sum(),
            'total_adjusted_schedule': da_results['Adjusted_Sch'].sum(),
            'total_rtm_available': da_results['RTM_Available'].sum(),
            'total_rtm_solar': da_results['RTM_Solar'].sum(),
            'total_gdam_solar_output': da_results['G-DAM_Solar_Output'].sum(),
            'total_rtm_wind': da_results['RTM_Wind'].sum(),
            'total_gdam_wind_output': da_results['G-DAM_Wind_Output'].sum(),
            'buyer_allocations': {}
        }

        # Calculate buyer allocations
        for buyer in buyer_list:
            solar_col = f"{buyer}_Solar"
            wind_col = f"{buyer}_Wind"
            if solar_col in da_results.columns and wind_col in da_results.columns:
                total_solar = da_results[solar_col].sum()
                total_wind = da_results[wind_col].sum()
                da_summary['buyer_allocations'][buyer] = {
                    'solar': total_solar,
                    'wind': total_wind,
                    'total': total_solar + total_wind
                }

        logger.info(f"DA (Day-Ahead) results for {da_date}: {da_summary}")
        log_output_results("DAY_AHEAD", da_date, da_summary, da_results)
    else:
        logger.warning(f"No DA (Day-Ahead) results generated for {da_date} - load data may be missing")

    # Display ID results
    st.subheader(f"ID Bifurcation (Intraday) for {id_date}")
    if not id_results.empty:
        st.dataframe(
            id_results.drop(
                columns=["Date_only", "G-DAM_Solar_Output", "G-DAM_Wind_Output"],
                errors="ignore",
            )
        )
    else:
        st.warning(f"No Intraday results available for {id_date}")

    # Display DA results
    st.subheader(f"DA Bifurcation (Day-Ahead) for {da_date}")
    if not da_results.empty:
        st.dataframe(da_results.drop(columns=["Date_only"]))
    else:
        st.info(f"No Day-Ahead results available for {da_date}. Load data may not be available yet.")

    if energy_mismatch_records:
        mismatch_df = pd.DataFrame(energy_mismatch_records)
        st.error("Energy conservation mismatches were detected in the blocks below.")
        st.dataframe(mismatch_df)

    # --- Sidebar: Display Previous Day's CUF% (from block 96) --->
    if not cuf_data_df.empty:
        cuf_sidebar_df = cuf_data_df.set_index("buyer_name")
        st.sidebar.markdown("### Previous Day (Block 96) CUF%")
        st.sidebar.dataframe(cuf_sidebar_df)
    else:
        st.sidebar.write("No CUF% data available from previous dates.")

    csv_data = result_df.to_csv(index=False).encode('utf-8')
    st.download_button(
        label="Download full results as CSV",
        data=csv_data,
        file_name="allocation_results.csv",
        mime="text/csv"
    )

    # Debug summary
    debug_count = len(st.session_state.debug_logs) if 'debug_logs' in st.session_state else 0
    if debug_count > 0:
        st.info(f"🔍 **Debug Information Available**: {debug_count} debug entries captured. Go to 'Debug Logs' page to view detailed allocation process information.")
    else:
        st.warning("No debug information was captured during this run.")

    # -------------------------------
    # Create and Download REMC Template Files
    # -------------------------------
    st.subheader("📄 Creating REMC Template Files")

    try:
        # Reopen database connection for REMC template creation
        try:
            conn = get_db_connection()
            logger.info("Database connection established for REMC template creation")
        except Exception as db_error:
            st.error(f"❌ Failed to connect to database: {str(db_error)}")
            logger.error(f"Database connection failed for REMC templates: {str(db_error)}")
            raise

        # Create REMC template files
        id_file_path, id_revision, da_file_path, da_revision = create_remc_template_files(
            result_df, id_date, da_date, buyer_list, conn, logger
        )

        # Close the connection after template creation
        conn.close()
        logger.info("Database connection closed after REMC template creation")

        # Store files in session state for auto-download
        st.session_state.remc_files_ready = True

        # Provide download buttons for the created files
        col1, col2 = st.columns(2)

        if id_file_path and os.path.exists(id_file_path):
            with open(id_file_path, 'rb') as f:
                id_file_data = f.read()

            id_filename = os.path.basename(id_file_path)
            with col1:
                st.download_button(
                    label=f"📥 Download ID Template ({id_date.strftime('%d-%m-%Y')})",
                    data=id_file_data,
                    file_name=id_filename,
                    mime="text/csv",
                    key="download_id_template"
                )
            st.success(f"✅ ID template created: {id_filename}")
            logger.info(f"ID template file ready for download: {id_filename}")

            # Store ID file path in session state for manual submission
            st.session_state.id_file_path = id_file_path
            st.session_state.id_date = id_date
            st.session_state.id_revision = id_revision
        else:
            st.warning("⚠️ ID template file could not be created")
            logger.warning("ID template file creation failed")

        if da_file_path and os.path.exists(da_file_path):
            with open(da_file_path, 'rb') as f:
                da_file_data = f.read()

            da_filename = os.path.basename(da_file_path)
            with col2:
                st.download_button(
                    label=f"📥 Download DA Template ({da_date.strftime('%d-%m-%Y')})",
                    data=da_file_data,
                    file_name=da_filename,
                    mime="text/csv",
                    key="download_da_template"
                )
            st.success(f"✅ DA template created: {da_filename}")
            logger.info(f"DA template file ready for download: {da_filename}")

            # Store DA file path in session state for manual submission
            st.session_state.da_file_path = da_file_path
            st.session_state.da_date = da_date
            st.session_state.da_revision = da_revision
        else:
            if not da_results.empty:
                st.warning("⚠️ DA template file could not be created")
                logger.warning("DA template file creation failed")
            else:
                st.info("ℹ️ DA template file not created (no DA results available)")
                logger.info("DA template file skipped - no DA results")

        # Add manual portal submission buttons
        st.markdown("---")
        st.subheader("🚀 Submit to REMC Portal")
        st.info("Click the buttons below to submit templates to the REMC portal")

        col1, col2 = st.columns(2)

        # ID Template submission button
        with col1:
            if hasattr(st.session_state, 'id_file_path') and st.session_state.id_file_path and os.path.exists(st.session_state.id_file_path):
                if st.button("🚀 Submit ID Template to REMC Portal", key="submit_id_portal"):
                    st.info("🚀 Initiating REMC Portal Submission for Intraday...")
                    submit_remc_schedule(st.session_state.id_file_path, "Intraday", logger)
            else:
                st.button("🚀 Submit ID Template to REMC Portal", key="submit_id_portal_disabled", disabled=True)
                st.caption("ID template not available")

        # DA Template submission button
        with col2:
            if hasattr(st.session_state, 'da_file_path') and st.session_state.da_file_path and os.path.exists(st.session_state.da_file_path):
                if st.button("🚀 Submit DA Template to REMC Portal", key="submit_da_portal"):
                    st.info("🚀 Initiating REMC Portal Submission for Day Ahead...")
                    submit_remc_schedule(st.session_state.da_file_path, "Day Ahead", logger)
            else:
                st.button("🚀 Submit DA Template to REMC Portal", key="submit_da_portal_disabled", disabled=True)
                st.caption("DA template not available")

        # Add manual email sending buttons
        st.markdown("---")
        st.subheader("📧 Send Email Notifications")
        st.info("Click the buttons below to send template emails to stakeholders")
        st.caption(f"Duplicate send protection is active ({EMAIL_SEND_COOLDOWN_SECONDS // 60} minute cooldown per email type/date/revision).")

        if st.button("🔓 Clear Email Cooldowns", key="clear_email_cooldowns"):
            st.session_state.email_send_locks = {}
            st.success("✅ Email cooldown locks cleared.")

        col1, col2 = st.columns(2)

        # ID Template email button
        with col1:
            if hasattr(st.session_state, 'id_file_path') and st.session_state.id_file_path and os.path.exists(st.session_state.id_file_path):
                id_email_key = f"remc_id::{st.session_state.id_date}::R{st.session_state.id_revision}"
                id_cooldown_remaining = get_email_cooldown_remaining(id_email_key)

                if id_cooldown_remaining > 0:
                    st.caption(f"⏳ ID email cooldown active: {format_cooldown(id_cooldown_remaining)} remaining")

                if st.button("📧 Send ID Template Email", key="send_id_email", disabled=id_cooldown_remaining > 0):
                    st.info("📤 Sending Intraday template email...")
                    email_success = send_remc_template_email(
                        st.session_state.id_file_path, "Intraday",
                        st.session_state.id_date, st.session_state.id_revision, logger
                    )
                    if email_success:
                        mark_email_sent(id_email_key)
                        st.success("✅ Intraday template email sent successfully!")
                        st.info(f"📧 Email sent to: {', '.join(REMC_EMAIL_TO)}")
                        st.info(f"📋 CC: {', '.join(REMC_EMAIL_CC)}")
                    else:
                        st.error("❌ Failed to send Intraday template email. Check logs for details.")
            else:
                st.button("📧 Send ID Template Email", key="send_id_email_disabled", disabled=True)
                st.caption("ID template not available")

        # DA Template email button
        with col2:
            if hasattr(st.session_state, 'da_file_path') and st.session_state.da_file_path and os.path.exists(st.session_state.da_file_path):
                da_email_key = f"remc_da::{st.session_state.da_date}::R{st.session_state.da_revision}"
                da_cooldown_remaining = get_email_cooldown_remaining(da_email_key)

                if da_cooldown_remaining > 0:
                    st.caption(f"⏳ DA email cooldown active: {format_cooldown(da_cooldown_remaining)} remaining")

                if st.button("📧 Send DA Template Email", key="send_da_email", disabled=da_cooldown_remaining > 0):
                    st.info("📤 Sending Day Ahead template email...")
                    email_success = send_remc_template_email(
                        st.session_state.da_file_path, "Day Ahead",
                        st.session_state.da_date, st.session_state.da_revision, logger
                    )
                    if email_success:
                        mark_email_sent(da_email_key)
                        st.success("✅ Day Ahead template email sent successfully!")
                        st.info(f"📧 Email sent to: {', '.join(REMC_EMAIL_TO)}")
                        st.info(f"📋 CC: {', '.join(REMC_EMAIL_CC)}")
                    else:
                        st.error("❌ Failed to send Day Ahead template email. Check logs for details.")
            else:
                st.button("📧 Send DA Template Email", key="send_da_email_disabled", disabled=True)
                st.caption("DA template not available")

        # Clean up temporary files after a delay (they'll be cleaned up when user closes browser)
        # Note: Files will remain in temp directory until manually cleaned or system cleanup

    except Exception as e:
        st.error(f"❌ Error creating REMC template files: {str(e)}")
        logger.error(f"REMC template creation error: {str(e)}")

    # -------------------------------
    # RTM Downloads + Email
    # -------------------------------
    st.markdown("---")
    st.subheader("📧 RTM Files and Email")
    st.info("RTM bid email is manual. Download Solar RTM and Wind RTM files below, and use the wind email button if needed.")
    st.caption(f"Solar RTM now uses portfolio `{SOLAR_RTM_PORTFOLIO_ID}` inside the generated workbook. Wind RTM continues to use `{WIND_RTM_PORTFOLIO_ID}`.")

    current_time_ist = datetime.datetime.now(IST)
    cutoff_time = current_time_ist.replace(hour=22, minute=15, second=0, microsecond=0)
    midnight = current_time_ist.replace(hour=23, minute=59, second=59, microsecond=999999)
    use_da_values = cutoff_time <= current_time_ist <= midnight

    rtm_today = current_time_ist.date()
    rtm_tomorrow = rtm_today + datetime.timedelta(days=1)

    rtm_wind_values = []
    rtm_solar_values = []
    source_type = "None"
    delivery_date_for_email = rtm_today
    shared_rtm_validation_error = None

    if use_da_values:
        source_type = "DA (Day-Ahead)"
        delivery_date_for_email = rtm_tomorrow
        st.info(f"⏰ Time is between 10:15 PM and midnight ({current_time_ist.strftime('%I:%M %p')} IST)")
        st.info("📊 RTM files will use Day-Ahead RTM values")

        if not da_results.empty:
            da_results_sorted = da_results.sort_values('Block')
            if 'RTM_Wind' in da_results_sorted.columns:
                rtm_wind_values = da_results_sorted['RTM_Wind'].tolist()
                logger.info(f"Prepared DA RTM_Wind values for delivery date: {delivery_date_for_email}")
            if 'RTM_Solar' in da_results_sorted.columns:
                rtm_solar_values = da_results_sorted['RTM_Solar'].tolist()
                logger.info(f"Prepared DA RTM_Solar values for delivery date: {delivery_date_for_email}")
        else:
            shared_rtm_validation_error = "Day-Ahead results are required between 10:15 PM and midnight IST."
            st.warning(f"⚠️ {shared_rtm_validation_error}")
            logger.warning("No DA results available during 10:15 PM - midnight IST window for RTM files")
    else:
        source_type = "ID (Intraday)"
        delivery_date_for_email = rtm_today
        st.info(f"⏰ Current time: {current_time_ist.strftime('%I:%M %p')} IST")
        st.info("📊 RTM files will use Intraday RTM values")

        if not id_results.empty:
            id_results_sorted = id_results.sort_values('Block')
            if 'RTM_Wind' in id_results_sorted.columns:
                rtm_wind_values = id_results_sorted['RTM_Wind'].tolist()
                logger.info(f"Prepared ID RTM_Wind values for delivery date: {delivery_date_for_email}")
            if 'RTM_Solar' in id_results_sorted.columns:
                rtm_solar_values = id_results_sorted['RTM_Solar'].tolist()
                logger.info(f"Prepared ID RTM_Solar values for delivery date: {delivery_date_for_email}")
        else:
            shared_rtm_validation_error = "Intraday results are not available for RTM files."
            st.warning(f"⚠️ {shared_rtm_validation_error}")
            logger.warning("No ID results available for RTM files")

    if shared_rtm_validation_error is None:
        st.info(f"📅 Delivery date for RTM bid: {delivery_date_for_email.strftime('%d/%m/%Y')}")

    wind_rtm_validation_error = shared_rtm_validation_error
    solar_rtm_validation_error = shared_rtm_validation_error

    if wind_rtm_validation_error is None:
        if len(rtm_wind_values) == 96:
            st.info(f"📊 Extracted 96 RTM_Wind values from {source_type}")
            logger.info(f"Extracted 96 RTM_Wind values from {source_type} results")
        elif len(rtm_wind_values) > 0:
            wind_rtm_validation_error = f"Expected 96 RTM_Wind values, got {len(rtm_wind_values)}."
            st.warning(f"⚠️ {wind_rtm_validation_error}")
            logger.warning(f"Incorrect number of RTM_Wind values: {len(rtm_wind_values)}")
        else:
            wind_rtm_validation_error = f"No RTM_Wind values available from {source_type}."
            st.warning(f"⚠️ {wind_rtm_validation_error}")
            logger.warning("No RTM_Wind values available for RTM files")

    if solar_rtm_validation_error is None:
        if len(rtm_solar_values) == 96:
            st.info(f"📊 Extracted 96 RTM_Solar values from {source_type}")
            logger.info(f"Extracted 96 RTM_Solar values from {source_type} results")
        elif len(rtm_solar_values) > 0:
            solar_rtm_validation_error = f"Expected 96 RTM_Solar values, got {len(rtm_solar_values)}."
            st.warning(f"⚠️ {solar_rtm_validation_error}")
            logger.warning(f"Incorrect number of RTM_Solar values: {len(rtm_solar_values)}")
        else:
            solar_rtm_validation_error = f"No RTM_Solar values available from {source_type}."
            st.warning(f"⚠️ {solar_rtm_validation_error}")
            logger.warning("No RTM_Solar values available for RTM files")

    rtm_wind_file = None
    rtm_solar_file = None

    if wind_rtm_validation_error is None:
        rtm_wind_file = create_rtm_file(rtm_wind_values, delivery_date_for_email, "Wind")
        st.session_state.rtm_wind_file_path = rtm_wind_file
    else:
        st.session_state.rtm_wind_file_path = None

    if solar_rtm_validation_error is None:
        rtm_solar_file = create_rtm_file(rtm_solar_values, delivery_date_for_email, "Solar")
        st.session_state.rtm_solar_file_path = rtm_solar_file
    else:
        st.session_state.rtm_solar_file_path = None

    st.session_state.rtm_delivery_date = delivery_date_for_email
    st.session_state.rtm_source_type = source_type

    download_col1, download_col2 = st.columns(2)

    with download_col1:
        if rtm_solar_file and os.path.exists(rtm_solar_file):
            with open(rtm_solar_file, 'rb') as f:
                solar_rtm_data = f.read()
            st.download_button(
                label=f"📥 Download Solar RTM ({delivery_date_for_email.strftime('%d-%m-%Y')})",
                data=solar_rtm_data,
                file_name=os.path.basename(rtm_solar_file),
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                key="download_solar_rtm",
            )
        else:
            st.button("📥 Download Solar RTM", key="download_solar_rtm_disabled", disabled=True)
            if solar_rtm_validation_error:
                st.caption(solar_rtm_validation_error)

    with download_col2:
        if rtm_wind_file and os.path.exists(rtm_wind_file):
            with open(rtm_wind_file, 'rb') as f:
                wind_rtm_data = f.read()
            st.download_button(
                label=f"📥 Download Wind RTM ({delivery_date_for_email.strftime('%d-%m-%Y')})",
                data=wind_rtm_data,
                file_name=os.path.basename(rtm_wind_file),
                mime="application/vnd.ms-excel.sheet.macroEnabled.12",
                key="download_wind_rtm",
            )
        else:
            st.button("📥 Download Wind RTM", key="download_wind_rtm_disabled", disabled=True)
            if wind_rtm_validation_error:
                st.caption(wind_rtm_validation_error)

    rtm_email_key = f"rtm::{delivery_date_for_email}::{source_type}"
    rtm_cooldown_remaining = get_email_cooldown_remaining(rtm_email_key)

    if rtm_cooldown_remaining > 0:
        st.caption(f"⏳ RTM email cooldown active: {format_cooldown(rtm_cooldown_remaining)} remaining")

    send_rtm_clicked = st.button(
        "📧 Send Wind RTM Bid Email",
        key="send_rtm_email",
        disabled=(wind_rtm_validation_error is not None or rtm_cooldown_remaining > 0)
    )

    if send_rtm_clicked and wind_rtm_validation_error is None:
        rtm_file = rtm_wind_file or create_rtm_file(rtm_wind_values, delivery_date_for_email, "Wind")

        if rtm_file:
            st.info("📤 Sending Wind RTM bid email...")
            email_success = send_rtm_email(rtm_file, delivery_date_for_email, logger)

            if email_success:
                mark_email_sent(rtm_email_key)
                st.success("✅ Wind RTM email sent successfully!")
                st.success(f"📧 Email sent to: {', '.join(EMAIL_TO)}")
                st.info(f"📋 CC: {', '.join(EMAIL_CC[:3])}... (+{len(EMAIL_CC)-3} more)")
            else:
                st.error("❌ Failed to send Wind RTM email. Check logs for details.")

            try:
                if os.path.exists(rtm_file):
                    os.remove(rtm_file)
                    logger.info(f"Cleaned up temporary RTM file: {rtm_file}")
            except Exception as e:
                logger.warning(f"Failed to clean up temporary RTM file: {str(e)}")
        else:
            st.error("❌ Failed to create Wind RTM file. Email not sent.")
            logger.error("Wind RTM file creation failed - email not sent")

    # Final logging
    logger.info("=== BIFURCATION PROCESS COMPLETED SUCCESSFULLY ===")
    logger.info(f"Total allocation results generated: {len(result_df)} records")
    logger.info(f"ID results: {len(id_results)} blocks, DA results: {len(da_results)} blocks")
    logger.info(f"Debug entries captured: {debug_count}")

    # Display completion message based on what was processed
    if not id_results.empty and not da_results.empty:
        st.success(f"✅ Bifurcation completed successfully for both ID and DA! Logs saved to: {log_file}")
    elif not id_results.empty:
        st.success(f"✅ Bifurcation completed successfully for ID (Intraday) only! Logs saved to: {log_file}")
        st.info(f"ℹ️ Day-Ahead bifurcation was skipped due to missing load data for {da_date}")
    else:
        st.warning(f"⚠️ Bifurcation completed with no results. Check logs: {log_file}")

def show_debug_logs():
    """Display debug logs from bifurcation process"""
    st.subheader("🔍 Debug Logs")

    if 'debug_logs' not in st.session_state or not st.session_state.debug_logs:
        st.info("No debug logs available. Run bifurcation to generate debug information.")
        if st.button("Clear Debug Logs"):
            st.session_state.debug_logs = []
            st.success("Debug logs cleared!")
        return

    st.write(f"**Total debug entries:** {len(st.session_state.debug_logs)}")

    # Add clear button
    col1, col2 = st.columns([1, 4])
    with col1:
        if st.button("Clear Debug Logs"):
            st.session_state.debug_logs = []
            st.success("Debug logs cleared!")
            st.rerun()

    # Filter options
    st.write("### Filter Options")
    debug_types = list(set([log.get('type', 'unknown') for log in st.session_state.debug_logs]))
    selected_types = st.multiselect("Select debug types to show:", debug_types, default=debug_types)

    dates = list(set([str(log.get('date', 'unknown')) for log in st.session_state.debug_logs if log.get('date')]))
    selected_dates = st.multiselect("Select dates to show:", sorted(dates), default=dates)

    blocks = list(set([log.get('block') for log in st.session_state.debug_logs if log.get('block')]))
    if blocks:
        selected_blocks = st.multiselect("Select blocks to show:", sorted(blocks), default=sorted(blocks))
    else:
        selected_blocks = []

    # Display filtered logs
    st.write("### Debug Information")

    filtered_logs = []
    for log in st.session_state.debug_logs:
        if (log.get('type') in selected_types and
            str(log.get('date', 'unknown')) in selected_dates and
            (not blocks or log.get('block') in selected_blocks)):
            filtered_logs.append(log)

    if not filtered_logs:
        st.warning("No logs match the selected filters.")
        return

    for i, log in enumerate(filtered_logs):
        log_type = log.get('type', 'unknown')

        if log_type == 'date_processing':
            with st.expander(f"📅 Date Processing: {log.get('date')}", expanded=False):
                st.write(f"**Date:** {log.get('date')}")
                st.write(f"**High Priority Buyers:** {log.get('high_priority_buyers', [])}")
                st.write(f"**Low Priority Buyers:** {log.get('low_priority_buyers', [])}")
                st.write(f"**Schedule Records:** {log.get('schedule_records', 0)}")

        elif log_type == 'block_processing':
            with st.expander(f"🔢 Block {log.get('block')} - {log.get('date')}", expanded=False):
                st.write(f"**Date:** {log.get('date')}, **Block:** {log.get('block')}")
                st.write(f"**Schedule:** Solar={log.get('solar_val')}, Wind={log.get('wind_val')}, Total={log.get('total_sch')}")
                st.write(f"**GDAM:** {log.get('gdam')}, **Adjusted:** {log.get('adjusted_sch')}, **Remaining:** {log.get('remaining_energy')}")
                st.write(f"**Requisition:** {log.get('requisition', {})}")
                st.write(f"**Requisition Row Empty:** {log.get('req_row_empty', True)}")
                if log.get('raw_req_row'):
                    st.write(f"**Raw Requisition Row:** {log.get('raw_req_row')}")

        elif log_type == 'hp_allocation':
            with st.expander(f"⚡ HP Allocation - Block {log.get('block')} - {log.get('date')}", expanded=False):
                st.write(f"**Date:** {log.get('date')}, **Block:** {log.get('block')}")
                st.write(f"**HP Allocation Results:** {log.get('hp_alloc', {})}")
                st.write(f"**Total Allocations:** {log.get('total_allocations', {})}")
                st.write(f"**Remaining Energy after HP:** {log.get('remaining_energy_after_hp')}")

        elif log_type == 'no_hp_buyers':
            with st.expander(f"❌ No HP Buyers - Block {log.get('block')} - {log.get('date')}", expanded=False):
                st.write(f"**Date:** {log.get('date')}, **Block:** {log.get('block')}")
                st.write(f"**Message:** {log.get('message')}")

        elif log_type == 'solar_wind_split':
            with st.expander(f"🌞💨 Solar/Wind Split - Block {log.get('block')} - {log.get('date')}", expanded=False):
                st.write(f"**Date:** {log.get('date')}, **Block:** {log.get('block')}")
                st.write(f"**Solar/Wind Split:**")
                for buyer, values in log.get('solar_wind_split', {}).items():
                    st.write(f"  {buyer}: Solar={values.get('solar')}, Wind={values.get('wind')}")
                st.write(f"**Allocation Row Keys:** {log.get('allocation_row_keys', [])}")
                if log.get('has_none_values'):
                    st.error(f"❌ **Found None values:** {log.get('none_values', [])}")
                else:
                    st.success("✅ **No None values found**")

        elif log_type == 'result_entry':
            with st.expander(f"📊 Result Entry - Block {log.get('block')} - {log.get('date')}", expanded=False):
                st.write(f"**Date:** {log.get('date')}, **Block:** {log.get('block')}")
                st.write(f"**Buyer Allocations in Result:**")
                for key, value in log.get('buyer_allocations', {}).items():
                    st.write(f"  {key}: {value}")
                if log.get('has_none_in_result'):
                    st.error(f"❌ **Found None values in result:** {log.get('none_values_in_result', [])}")
                else:
                    st.success("✅ **No None values in result**")

        elif log_type == 'priority_classification':
            with st.expander(f"🎯 Priority Classification (from Excel)", expanded=True):
                st.write(f"**Source:** {log.get('source')}")
                st.write(f"**Totals:** {log.get('high_priority_count')} High Priority, {log.get('low_priority_count')} Low Priority")
                
                col1, col2 = st.columns(2)
                with col1:
                    st.write("**High Priority Buyers:**")
                    st.write(log.get('high_priority_buyers', []))
                with col2:
                    st.write("**Low Priority Buyers:**")
                    st.write(log.get('low_priority_buyers', []))

def clear_data_up_to_date():
    st.subheader("Clear Data Up To a Selected Date")
    st.warning("This will permanently delete all data up to and including the selected date from load_data, obligation, schedule_wind, schedule_solar, and hybrid_avc_cap tables.")
    clear_date = st.date_input("Clear all data up to and including this date:", value=today_ist())
    if st.button("Clear Data"):
        conn = get_db_connection()
        cursor = conn.cursor()
        for table in ["load_data", "obligation", "schedule_wind", "schedule_solar", "hybrid_avc_cap"]:
            cursor.execute(f"DELETE FROM {table} WHERE date <= %s", (clear_date,))
        conn.commit()
        cursor.close()
        conn.close()
        st.success(f"Data up to and including {clear_date} cleared from all relevant tables.")

# -------------------------------
# Main Application
# -------------------------------

def main():
    st.title("4th Partner Bifurcation Tool")

    # Add system time to sidebar (IST)
    import pytz
    ist = pytz.timezone('Asia/Kolkata')
    current_time = datetime.datetime.now(ist)
    st.sidebar.write(f"System Time (IST): {current_time.strftime('%Y-%m-%d %H:%M:%S')}")

    st.sidebar.header("Actions")
    action = st.sidebar.selectbox("Select Action", [
        "Create Tables",
        "Load Data",
        "Manual Data Entry",
        "View Tables",
        "Update Hybrid AVC Cap",
        "Edit G-DAM allocation",
        "TGNA/ GNA buyers",
        "Edit Contract Value",
        "Edit State",
        "Edit Tariff Difference",
        "Edit Final Revision Integration",
        "Test REMC Template Mapping",
        "Process Final Buyer Schedule",
        "Run Bifurcation",
        "Debug Logs",
        "Clear Data"
    ])

    # Date calculations (using IST)
    today = current_time.date()  # Use IST time from above
    tomorrow = today + datetime.timedelta(days=1)
    yesterday = today - datetime.timedelta(days=1)

    if action == "Create Tables":
        create_tables()
        create_config_tables()
        create_gdam_rtm_ratio_table()
        create_manual_priority_table()
        st.success("Data and configuration tables created successfully (if they did not already exist).")

    elif action == "Load Data":
        load_data_to_mysql()

    elif action == "Manual Data Entry":
        table_choice = st.sidebar.selectbox("Select Table to Edit",
                                             ["load_data", "obligation", "solar_and_wind_schedules", "gdam_rtm_ratio"])
        if table_choice == "load_data":
            manual_fill_load_data()
        elif table_choice == "obligation":
            manual_fill_obligation_data()
        elif table_choice == "solar_and_wind_schedules":
            manual_fill_schedule_combined()
        elif table_choice == "gdam_rtm_ratio":
            manual_fill_gdam_rtm_ratio()

    elif action == "View Tables":
        table_options = ["load_data", "obligation", "schedule_solar", "schedule_wind",
                        "contract_value", "state", "tariff_difference", "buyer_mapping", "buyer_access_type", "cuf_pct", "gdam_rtm_ratio", "hybrid_avc_cap", "gdam_allocation_config"]
        selected_table = st.sidebar.selectbox("Select Table to View", table_options)
        df = display_table(selected_table)

        # Calculate dates
        today = today_ist()
        tomorrow = today + datetime.timedelta(days=1)
        yesterday = today - datetime.timedelta(days=1)

        # Special handling for CUF% table
        if selected_table == "cuf_pct":
            # Always show yesterday's CUF% data
            date_filter = yesterday
            filter_text = "yesterday's"
        elif selected_table == "buyer_mapping":
            # Don't apply date filtering for buyer_mapping
            date_filter = None
            filter_text = "all"
        else:
            # For other tables, show today and tomorrow
            date_filter = [today, tomorrow]
            filter_text = "today's and tomorrow's"

        # Apply date filtering
        date_columns = [col for col in df.columns if col.lower() == "date"]
        if date_columns and date_filter is not None:
            date_col = date_columns[0]
            try:
                if selected_table == "cuf_pct":
                    df = df[pd.to_datetime(df[date_col]).dt.date == date_filter]
                else:
                    df = df[pd.to_datetime(df[date_col]).dt.date.isin(date_filter)]
            except Exception as e:
                st.warning(f"Could not filter dates: {e}")

        st.write(f"Showing {filter_text} data from the '{selected_table}' table:")
        st.dataframe(df)

    elif action == "Update Hybrid AVC Cap":
        update_hybrid_avc_cap()

    elif action == "Edit G-DAM allocation":
        update_gdam_allocation()

    elif action == "TGNA/ GNA buyers":
        update_buyer_access_type()

    elif action == "Edit Contract Value":
        update_contract_value()

    elif action == "Edit State":
        update_state()

    elif action == "Edit Tariff Difference":
        update_tariff_difference()

    elif action == "Edit Final Revision Integration":
        update_buyer_mapping()

    elif action == "Test REMC Template Mapping":
        test_remc_template_mapping()

    elif action == "Process Final Buyer Schedule":
        st.subheader("Process Final Buyer Schedule")

        # Date selection
        selected_date = st.date_input("Select Date to Process", value=today_ist())

        # Hard-coded values (not displayed to user)
        start_row = 17
        end_row = 115

        if st.button("Process Selected Date"):
            process_final_buyer_schedule(selected_date, start_row, end_row)

        st.write("---")
        st.write("### Process Multiple Dates")
        st.write("Use this option to process a range of dates at once")

        col1, col2 = st.columns(2)
        with col1:
            start_date = st.date_input("Start Date", value=today_ist() - datetime.timedelta(days=7))
        with col2:
            end_date = st.date_input("End Date", value=today_ist())

        if st.button("Process Date Range"):
            if start_date > end_date:
                st.error("Start date must be before or equal to end date")
            else:
                current_date = start_date
                success_count = 0
                fail_count = 0

                while current_date <= end_date:
                    st.write(f"Processing {current_date}...")
                    result = process_final_buyer_schedule(current_date, start_row, end_row)
                    if result:
                        success_count += 1
                    else:
                        fail_count += 1
                    current_date += datetime.timedelta(days=1)

                st.success(f"Processed {success_count} dates successfully. Failed: {fail_count}")

    elif action == "Run Bifurcation":
        st.subheader("Run Bifurcation")
        st.info("Bifurcation will run using automatic priority classification based on the 'Low_Priority_FP.xlsx' file on FTP.")

        if st.button("Run Bifurcation", type="primary"):
            run_bifurcation()

    elif action == "Debug Logs":
        show_debug_logs()

    elif action == "Clear Data":
        clear_data_up_to_date()

if __name__ == "__main__":
    main()
